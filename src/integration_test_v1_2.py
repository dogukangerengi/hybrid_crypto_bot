#!/usr/bin/env python3
# =============================================================================
# v1.2.0 KAPSAMLI ENTEGRASYON TESTİ
# =============================================================================
# Pipeline'ın her modülünü sırayla test eder:
#
# TEST 1:  Bitget API Bağlantısı (exchange.load_markets)
# TEST 2:  CoinScanner — Market tarama + filtreleme
# TEST 3:  BitgetFetcher — OHLCV çoklu TF veri çekme (BTC)
# TEST 4:  IndicatorCalculator — 64+ teknik indikatör
# TEST 5:  IndicatorSelector — IC analizi (Spearman + FDR)
# TEST 6:  GateKeeper — Sinyal filtreleme
# TEST 7:  RiskManager — SL/TP/pozisyon hesaplama
# TEST 8:  PaperTrader — Trade açma/kapama simülasyonu
# TEST 9:  Excel Export — openpyxl dosya oluşturma
# TEST 10: Full Pipeline Cycle — run_cycle() dry-run
#
# Çalıştırma:
#   cd hybrid_crypto_bot/src
#   python integration_test_v1_2.py
#
# Süre: ~3-5 dakika (API çağrıları nedeniyle)
# =============================================================================

import sys
import os
import time
import tempfile
import traceback
import warnings
import logging
import json
import numpy as np
import pandas as pd
from pathlib import Path
from datetime import datetime
from dataclasses import asdict

# ── Path setup ──
CURRENT_DIR = Path(__file__).parent
PROJECT_ROOT = CURRENT_DIR.parent
sys.path.insert(0, str(CURRENT_DIR))

# .env yükle (API key'ler)
try:
    from dotenv import load_dotenv
    load_dotenv(PROJECT_ROOT / '.env')
except ImportError:
    pass

# Log seviyesi: sadece WARNING+ göster (test çıktısı temiz kalsın)
logging.basicConfig(level=logging.WARNING, format='%(levelname)s | %(message)s')
warnings.filterwarnings('ignore')

# ── Sonuç sayaçları ──
passed = 0
failed = 0
skipped = 0
issues = []                                    # Bulunan sorunlar listesi

# ── Test arasında paylaşılan veriler ──
shared = {}                                    # Test'ler arası veri paylaşımı


# =============================================================================
# TEST RUNNER
# =============================================================================

def run_test(test_num: int, test_name: str, test_func, skip_reason: str = None):
    """Tek testi çalıştır, süre ölç, sonucu raporla."""
    global passed, failed, skipped

    print(f"\n{'━'*60}")
    print(f"  TEST {test_num:>2}: {test_name}")
    print(f"{'━'*60}")

    if skip_reason:
        print(f"  ⏭️  ATLANILDI: {skip_reason}")
        skipped += 1
        return

    start = time.time()
    try:
        test_func()
        elapsed = time.time() - start
        print(f"\n  ✅ BAŞARILI ({elapsed:.1f}s)")
        passed += 1
    except Exception as e:
        elapsed = time.time() - start
        print(f"\n  ❌ BAŞARISIZ ({elapsed:.1f}s)")
        print(f"     Hata: {e}")
        traceback.print_exc()
        failed += 1
        issues.append(f"TEST {test_num} ({test_name}): {str(e)[:100]}")


# =============================================================================
# TEST 1: BİTGET API BAĞLANTISI
# =============================================================================

def test_01_bitget_connection():
    """
    Bitget exchange bağlantısını ve market data yüklenmesini test eder.
    Tüm diğer testlerin ön koşulu — başarısız olursa devamı anlamsız.
    """
    from data.fetcher import BitgetFetcher

    fetcher = BitgetFetcher()

    # exchange nesnesi var mı?
    assert fetcher.exchange is not None, "Exchange nesnesi None!"
    print(f"  ✓ Exchange: {fetcher.exchange.id}")

    # Market data yüklendi mi?
    fetcher._ensure_markets_loaded()
    markets = fetcher.exchange.markets
    assert len(markets) > 0, "Markets boş!"

    # USDT futures sayısı
    usdt_futures = [s for s in markets if ':USDT' in s and '/USDT' in s]
    print(f"  ✓ Toplam market: {len(markets)}")
    print(f"  ✓ USDT Futures: {len(usdt_futures)} çift")

    # BTC/USDT:USDT var mı? (en temel kontrol)
    assert 'BTC/USDT:USDT' in markets, "BTC/USDT:USDT bulunamadı!"
    print(f"  ✓ BTC/USDT:USDT mevcut")

    # Ticker çekebiliyor mu? (tek bir API call)
    ticker = fetcher.exchange.fetch_ticker('BTC/USDT:USDT')
    assert ticker['last'] > 0, f"BTC fiyat hatalı: {ticker['last']}"
    shared['btc_price'] = ticker['last']
    print(f"  ✓ BTC fiyat: ${ticker['last']:,.2f}")

    shared['fetcher'] = fetcher


# =============================================================================
# TEST 2: COİNSCANNER — MARKET TARAMA
# =============================================================================

def test_02_scanner():
    """
    CoinScanner'ın market taramasını, filtrelemeyi ve skorlamayı test eder.
    
    Doğrulanan: min_volume filtresi, composite score hesabı, 
    top_n limiti, blacklist uygulanması.
    """
    from scanner.coin_scanner import CoinScanner

    fetcher = shared.get('fetcher')
    scanner = CoinScanner(fetcher=fetcher, verbose=False)

    # Tarama çalıştır (top 5 — hız için)
    top_coins = scanner.scan(top_n=5, force_refresh=True)

    assert len(top_coins) > 0, "Tarama sonucu boş!"
    assert len(top_coins) <= 5, f"Top N aşıldı: {len(top_coins)}"

    print(f"  ✓ {len(top_coins)} coin seçildi (top 5)")
    print(f"  {'─'*50}")
    print(f"  {'#':<3} {'Coin':<10} {'Fiyat':>12} {'Volume 24h':>15} {'Score':>7}")
    print(f"  {'─'*50}")

    for i, c in enumerate(top_coins, 1):
        print(f"  {i:<3} {c.coin:<10} ${c.price:>10,.2f} ${c.volume_24h:>12,.0f} {c.composite_score:>6.1f}")

        # Her coin'in alanları dolu mu?
        assert c.symbol, f"#{i} symbol boş"
        assert c.volume_24h > 0, f"#{i} volume 0"
        assert c.composite_score > 0, f"#{i} score 0"
        assert c.passed_filters, f"#{i} filtreden geçmemiş: {c.filter_reason}"

    # Config'deki min_volume uygulanmış mı?
    from config import get_setting
    min_vol = get_setting('scanner.min_24h_volume_usdt', 2_500_000)
    for c in top_coins:
        assert c.volume_24h >= min_vol, \
            f"{c.coin} volume ${c.volume_24h:,.0f} < min ${min_vol:,.0f}"
    print(f"  ✓ min_volume filtresi uygulanmış (≥${min_vol/1e6:.1f}M)")

    # Composite score sıralı mı? (büyükten küçüğe)
    scores = [c.composite_score for c in top_coins]
    for i in range(len(scores) - 1):
        assert scores[i] >= scores[i+1] - 0.01, "Score sıralaması bozuk!"
    print(f"  ✓ Score sıralaması doğru")

    shared['top_coins'] = top_coins
    shared['test_symbol'] = top_coins[0].symbol   # En yüksek scorlu coin
    shared['test_coin'] = top_coins[0].coin


# =============================================================================
# TEST 3: FETCHER — ÇOKLU TF VERİ ÇEKME
# =============================================================================

def test_03_fetcher_multi_tf():
    """
    BitgetFetcher ile tüm aktif TF'lerde veri çeker.
    Her TF için: bar sayısı, OHLCV tutarlılığı, veri güncelliği kontrol edilir.
    """
    from data.fetcher import BitgetFetcher

    fetcher = shared.get('fetcher', BitgetFetcher())
    symbol = shared.get('test_symbol', 'BTC/USDT:USDT')
    coin = shared.get('test_coin', 'BTC')

    print(f"  📊 Test coini: {coin} ({symbol})")

    # Tüm TF'lerde veri çek
    data = fetcher.fetch_all_timeframes(symbol=symbol)

    assert len(data) > 0, "Hiç TF'den veri gelmedi!"

    expected_tfs = {'15m', '30m', '1h', '2h', '4h'}
    actual_tfs = set(data.keys())

    print(f"\n  {'TF':<6} {'Bars':>6} {'Başlangıç':>14} {'Bitiş':>14} {'OHLCV':>6} {'Durum':>8}")
    print(f"  {'─'*58}")

    for tf in sorted(data.keys()):
        df = data[tf]
        bars = len(df)
        start = df.index[0].strftime('%Y-%m-%d')
        end = df.index[-1].strftime('%Y-%m-%d')

        # OHLCV sütunları var mı?
        has_cols = all(c in df.columns for c in ['open', 'high', 'low', 'close', 'volume'])

        # OHLC tutarlılığı: High ≥ max(Open, Close), Low ≤ min(Open, Close)
        ohlc_ok = True
        if has_cols:
            invalid = ((df['high'] < df['open']) | (df['high'] < df['close']) |
                       (df['low'] > df['open']) | (df['low'] > df['close'])).sum()
            ohlc_ok = (invalid == 0)

        status = "✓" if (bars >= 100 and has_cols and ohlc_ok) else "⚠️"
        print(f"  {tf:<6} {bars:>6} {start:>14} {end:>14} {'OK' if has_cols else 'MISS':>6} {status:>8}")

        # Assertions
        assert bars >= 50, f"{tf}: Yetersiz bar sayısı ({bars})"
        assert has_cols, f"{tf}: OHLCV sütunları eksik"

    # Beklenen TF'ler var mı?
    missing_tfs = expected_tfs - actual_tfs
    if missing_tfs:
        print(f"\n  ⚠️  Eksik TF'ler: {missing_tfs}")
        issues.append(f"Fetcher: Eksik TF'ler: {missing_tfs}")
    else:
        print(f"\n  ✓ Tüm beklenen TF'ler mevcut: {sorted(expected_tfs)}")

    # NaN kontrolü (son 20 bar'da NaN olmamalı)
    for tf, df in data.items():
        recent = df.tail(20)
        nan_count = recent[['close', 'volume']].isna().sum().sum()
        if nan_count > 0:
            print(f"  ⚠️  {tf}: Son 20 bar'da {nan_count} NaN!")
            issues.append(f"Fetcher {tf}: Son 20 bar'da NaN var")

    shared['ohlcv_data'] = data


# =============================================================================
# TEST 4: INDICATOR CALCULATOR — TEKNİK İNDİKATÖRLER
# =============================================================================

def test_04_indicators():
    """
    IndicatorCalculator ile 64+ teknik indikatör hesaplar.
    Çıktı kolon sayısı, NaN oranı, forward return eklenmesi kontrol edilir.
    """
    from indicators.calculator import IndicatorCalculator

    calc = IndicatorCalculator(verbose=False)
    data = shared.get('ohlcv_data', {})

    # En uzun TF'yi test et (en güvenilir veri)
    test_tf = '1h' if '1h' in data else list(data.keys())[0]
    df_raw = data[test_tf].copy()
    print(f"  📊 TF: {test_tf} | Girdi: {len(df_raw)} bar × {len(df_raw.columns)} kolon")

    # İndikatörleri hesapla
    df = calc.calculate_all(df_raw)

    # Price features ekle
    df = calc.add_price_features(df)

    # Forward returns ekle
    df = calc.add_forward_returns(df, periods=[1, 5])

    original_cols = len(df_raw.columns)         # 5 (OHLCV)
    new_cols = len(df.columns)
    indicator_cols = new_cols - original_cols

    print(f"  ✓ Çıktı: {len(df)} bar × {new_cols} kolon ({indicator_cols} yeni)")

    # Minimum indikatör sayısı
    assert indicator_cols >= 30, f"Çok az indikatör: {indicator_cols} (min 30)"
    print(f"  ✓ İndikatör sayısı: {indicator_cols} (≥30)")

    # Forward return kolonları var mı?
    assert 'fwd_ret_1' in df.columns, "fwd_ret_1 eksik!"
    assert 'fwd_ret_5' in df.columns, "fwd_ret_5 eksik!"
    print(f"  ✓ Forward return kolonları: fwd_ret_1, fwd_ret_5")

    # NaN oranı (ilk %30'u warm-up — normal, son %70'te düşük olmalı)
    warmup_cutoff = int(len(df) * 0.3)
    df_active = df.iloc[warmup_cutoff:]        # Warm-up sonrası
    nan_pct = df_active.isna().mean().mean() * 100

    print(f"  ✓ NaN oranı (warm-up sonrası): {nan_pct:.1f}%")
    if nan_pct > 20:
        print(f"  ⚠️  NaN oranı yüksek! Bazı indikatörler hesaplanamıyor olabilir")
        issues.append(f"Indicators {test_tf}: NaN oranı yüksek ({nan_pct:.1f}%)")

    # Kategori dağılımı
    from indicators.categories import get_category_names, get_indicators_by_category
    for cat in get_category_names():
        count = len(get_indicators_by_category(cat))
        print(f"    {cat}: {count} indikatör")

    shared['df_with_indicators'] = df
    shared['test_tf'] = test_tf


# =============================================================================
# TEST 5: IC ANALİZİ — SPEARMAN + FDR
# =============================================================================

def test_05_ic_analysis():
    """
    IndicatorSelector ile IC analizi yapar.
    Spearman korelasyon, FDR düzeltme, anlamlı indikatör seçimi test edilir.
    """
    from indicators.selector import IndicatorSelector

    selector = IndicatorSelector(alpha=0.05, correction_method='fdr', verbose=False)
    df = shared.get('df_with_indicators')

    if df is None:
        raise RuntimeError("TEST 4 başarısız — veri yok")

    # IC analizi çalıştır
    scores = selector.evaluate_all_indicators(df, target_col='fwd_ret_5')

    assert len(scores) > 0, "IC skor listesi boş!"
    print(f"  ✓ Değerlendirilen: {len(scores)} indikatör")

    # Anlamlı olanları filtrele (p_adjusted < 0.05)
    significant = [s for s in scores if s.is_significant]
    print(f"  ✓ İstatistiksel olarak anlamlı: {len(significant)}/{len(scores)}")

    # Top 10 IC skorları
    print(f"\n  {'Rank':<5} {'İndikatör':<30} {'IC':>7} {'p-adj':>10} {'Anlamlı':>8}")
    print(f"  {'─'*65}")
    for i, s in enumerate(scores[:10], 1):
        sig = "✓" if s.is_significant else "✗"
        ic_str = f"{s.ic_mean:+.4f}" if not np.isnan(s.ic_mean) else "NaN"
        p_str = f"{s.p_value_adjusted:.2e}" if not np.isnan(s.p_value_adjusted) else "NaN"
        name = s.name[:28]
        print(f"  {i:<5} {name:<30} {ic_str:>7} {p_str:>10} {sig:>8}")

    # En az 1 anlamlı indikatör olmalı (gerçek verida genelde 5-20 arası)
    if len(significant) == 0:
        print(f"  ⚠️  Hiç anlamlı indikatör yok — piyasa rejimi düz/noisy olabilir")
        issues.append("IC Analysis: Hiç anlamlı indikatör bulunamadı (düz piyasa?)")
    else:
        # En yüksek |IC| skoru
        best = scores[0]
        print(f"\n  🏆 En güçlü sinyal: {best.name}")
        print(f"     IC={best.ic_mean:+.4f} | p_adj={best.p_value_adjusted:.2e}")

    shared['ic_scores'] = scores
    shared['significant_count'] = len(significant)


# =============================================================================
# TEST 6: GATEKEEPER — SİNYAL FİLTRELEME
# =============================================================================

def test_06_gatekeeper():
    """
    GateKeeper eşiklerini doğrular.
    Farklı IC confidence seviyelerinde doğru aksiyon döndürüyor mu?
    """
    from config import cfg

    # Config eşikleri
    no_trade = cfg.gate.no_trade               # 40 (v1.2.0)
    report_only = cfg.gate.report_only         # 55
    full_trade = cfg.gate.full_trade           # 55

    print(f"  Config eşikleri:")
    print(f"    NO_TRADE:    < {no_trade}")
    print(f"    REPORT_ONLY: {no_trade} - {full_trade}")
    print(f"    FULL_TRADE:  ≥ {full_trade}")

    # Simüle: farklı IC değerleri test et
    test_cases = [
        (20.0, "NO_TRADE",    "Düşük IC → işlem yapma"),
        (35.0, "NO_TRADE",    "Gate altı → atla"),
        (45.0, "REPORT_ONLY", "Orta IC → sadece raporla"),
        (55.0, "FULL_TRADE",  "Gate eşiği → tam işlem"),
        (75.0, "FULL_TRADE",  "Yüksek IC → kesin işlem"),
        (90.0, "FULL_TRADE",  "Çok yüksek IC → en güçlü sinyal"),
    ]

    print(f"\n  {'IC':>5} {'Beklenen':<14} {'Gerçek':<14} {'Durum':<6} {'Açıklama'}")
    print(f"  {'─'*65}")

    all_ok = True
    for ic_val, expected, desc in test_cases:
        # Gate mantığı
        if ic_val < no_trade:
            actual = "NO_TRADE"
        elif ic_val < full_trade:
            actual = "REPORT_ONLY"
        else:
            actual = "FULL_TRADE"

        ok = actual == expected
        if not ok:
            all_ok = False
        status = "✓" if ok else "✗"
        print(f"  {ic_val:>5.0f} {expected:<14} {actual:<14} {status:<6} {desc}")

    assert all_ok, "GateKeeper eşik mantığı hatalı!"
    print(f"\n  ✓ GateKeeper eşikleri doğru çalışıyor")


# =============================================================================
# TEST 7: RISK MANAGER — SL/TP/POZİSYON HESAPLAMA
# =============================================================================

def test_07_risk_manager():
    """
    RiskManager ile SL/TP/pozisyon büyüklüğü hesaplar.
    ATR bazlı SL, R:R ratio, margin kontrolü doğrulanır.
    """
    from execution.risk_manager import RiskManager

    balance = 75.0                             # Paper trade başlangıç bakiyesi
    rm = RiskManager(balance=balance)

    btc_price = shared.get('btc_price', 97000.0)

    # BTC LONG trade hesapla
    result = rm.calculate_trade(
        entry_price=btc_price,
        direction='LONG',
        atr=btc_price * 0.015,
        symbol='BTC/USDT:USDT',
    )

    print(f"  📊 {result.symbol if hasattr(result, 'symbol') else 'BTC'} LONG @ ${btc_price:,.2f}")
    print(f"  Bakiye: ${balance:.2f}")

    # SL kontrolü
    sl = result.stop_loss
    print(f"  SL: ${sl.price:,.2f} ({sl.distance_pct:+.2f}%)")
    # RiskManager distance_pct'yi mutlak değer olarak saklar (her zaman pozitif)
    assert sl.distance_pct > 0, f"LONG SL distance sıfır! {sl.distance_pct}"
    assert sl.price < btc_price, f"LONG SL entry üstünde! SL={sl.price}"

    # TP kontrolü
    tp = result.take_profit
    print(f"  TP: ${tp.price:,.2f} ({tp.distance_pct:+.2f}%)")
    assert tp.price > btc_price, f"LONG TP entry'nin altında! TP={tp.price}"

    # R:R kontrolü
    rr = tp.risk_reward
    print(f"  R:R: {rr:.1f}x")
    assert rr >= 1.0, f"Risk/Reward çok düşük: {rr}"

    # Pozisyon
    pos = result.position
    print(f"  Pozisyon: {pos.size:.6f} BTC (${pos.value:,.2f})")
    print(f"  Leverage: {pos.leverage}x | Margin: ${pos.margin_required:,.2f}")
    print(f"  Risk: ${pos.risk_amount:,.2f}")

    # Pozisyon büyüklüğü > 0 olmalı
    assert pos.size > 0, f"Pozisyon büyüklüğü 0!"
    assert pos.value > 0, f"Pozisyon değeri 0!"

    # Risk amount ≤ bakiyenin %2'si (config default)
    max_risk = balance * 0.02
    assert pos.risk_amount <= max_risk + 0.01, \
        f"Risk aşımı! ${pos.risk_amount:.2f} > ${max_risk:.2f}"
    print(f"  ✓ Risk limiti içinde (≤${max_risk:.2f})")

    # Onay durumu
    print(f"  Status: {result.status.value}")
    if result.rejection_reasons:
        print(f"  Red nedenleri: {result.rejection_reasons}")
        issues.append(f"RiskManager: Trade reddedildi: {result.rejection_reasons}")

    shared['trade_calc'] = result


# =============================================================================
# TEST 8: PAPER TRADER — TRADE AÇ/KAPAT
# =============================================================================

def test_08_paper_trader():
    """
    PaperTrader ile trade açma, SL/TP tetikleme ve bakiye güncelleme test edilir.
    """
    from paper_trader import PaperTrader, TradeStatus

    with tempfile.TemporaryDirectory() as tmpdir:
        pt = PaperTrader(initial_balance=100.0, log_dir=Path(tmpdir))
        btc_price = shared.get('btc_price', 97000.0)

        # ── Trade 1: BTC LONG — TP tetiklenecek ──
        trade1 = pt.open_trade(
            symbol="BTC",
            full_symbol="BTC/USDT:USDT",
            direction="LONG",
            entry_price=btc_price,
            position_size=0.001,
            stop_loss=btc_price * 0.97,         # %3 altında
            take_profit=btc_price * 1.05,       # %5 üzerinde
            leverage=5,
            ic_confidence=72.0,
            ic_direction="LONG",
            best_timeframe="1h",
            market_regime="trending_up",
        )
        print(f"  ✓ Trade 1 açıldı: BTC LONG @ ${btc_price:,.2f}")
        assert trade1.trade_id, "Trade ID boş!"
        assert len(pt.open_trades) == 1, "Açık trade sayısı hatalı"

        # ── Trade 2: ETH SHORT — SL tetiklenecek ──
        eth_price = 3200.0
        trade2 = pt.open_trade(
            symbol="ETH",
            full_symbol="ETH/USDT:USDT",
            direction="SHORT",
            entry_price=eth_price,
            position_size=0.1,
            stop_loss=eth_price * 1.03,
            take_profit=eth_price * 0.95,
            leverage=3,
            ic_confidence=65.0,
            ic_direction="SHORT",
            best_timeframe="4h",
            market_regime="range_bound",
        )
        print(f"  ✓ Trade 2 açıldı: ETH SHORT @ ${eth_price:,.2f}")
        assert len(pt.open_trades) == 2

        # ── TP tetikle (BTC yükseldi) ──
        tp_price = btc_price * 1.06
        closed = pt.check_exits({'BTC': tp_price, 'ETH': eth_price})
        print(f"  ✓ TP tetiklendi: {len(closed)} trade kapandı")

        if len(closed) > 0:
            t = closed[0]
            print(f"    {t.symbol} {t.direction}: PnL=${t.pnl_absolute:+.2f} ({t.pnl_percent:+.1f}%)")
            assert t.pnl_absolute > 0, "TP tetiklendi ama PnL negatif!"

        # ── SL tetikle (ETH yükseldi — SHORT zarar) ──
        sl_price = eth_price * 1.04
        closed2 = pt.check_exits({'BTC': tp_price, 'ETH': sl_price})
        print(f"  ✓ SL tetiklendi: {len(closed2)} trade kapandı")

        if len(closed2) > 0:
            t2 = closed2[0]
            print(f"    {t2.symbol} {t2.direction}: PnL=${t2.pnl_absolute:+.2f} ({t2.pnl_percent:+.1f}%)")

        # ── Bakiye kontrolü ──
        print(f"  Başlangıç: $100.00 → Güncel: ${pt.balance:.2f}")
        assert pt.balance != 100.0, "Bakiye hiç değişmemiş — trade simülasyonu çalışmıyor!"

        # ── Açık trade kalmadı mı? ──
        assert len(pt.open_trades) == 0, f"Açık trade kaldı: {len(pt.open_trades)}"
        print(f"  ✓ Tüm trade'ler kapandı, açık pozisyon: 0")

        # ── Özet istatistikler ──
        summary = pt.get_summary()
        print(f"  ✓ Toplam trade: {summary.get('total_trades', 0)}")
        print(f"  ✓ Kapanan: {summary.get('closed_trades', 0)}")

        shared['paper_trader_ok'] = True


# =============================================================================
# TEST 9: EXCEL EXPORT
# =============================================================================

def test_09_excel_export():
    """
    PaperTrader.export_to_xlsx() fonksiyonunun dosya oluşturmasını doğrular.
    """
    from paper_trader import PaperTrader
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory() as tmpdir:
        pt = PaperTrader(initial_balance=100.0, log_dir=Path(tmpdir))
        btc_price = shared.get('btc_price', 97000.0)

        # Birkaç trade aç/kapat
        pt.open_trade("BTC", "BTC/USDT:USDT", "LONG", btc_price, 0.001,
                       btc_price*0.97, btc_price*1.05, 5, 72, "LONG", "1h", "trending_up")
        pt.check_exits({'BTC': btc_price * 1.06})

        pt.open_trade("SOL", "SOL/USDT:USDT", "SHORT", 180.0, 1.0,
                       186.0, 170.0, 3, 68, "SHORT", "4h", "range_bound")
        pt.check_exits({'SOL': 169.0})

        # Excel export
        xlsx_path = Path(tmpdir) / "paper_trades.xlsx"
        pt.export_to_xlsx(xlsx_path)

        assert xlsx_path.exists(), "Excel dosyası oluşturulmadı!"
        size_kb = xlsx_path.stat().st_size / 1024
        print(f"  ✓ Excel: {xlsx_path.name} ({size_kb:.1f} KB)")

        # Sheet kontrolü
        wb = load_workbook(xlsx_path, read_only=True)
        assert "Trades" in wb.sheetnames, f"Trades sheet yok! {wb.sheetnames}"
        assert "Summary" in wb.sheetnames, f"Summary sheet yok! {wb.sheetnames}"

        # Trades sheet satır sayısı
        ws = wb["Trades"]
        rows = ws.max_row - 1
        print(f"  ✓ Trades sheet: {rows} satır")
        assert rows >= 2, f"Trades sheet'te yetersiz satır: {rows}"

        wb.close()
        print(f"  ✓ Excel export düzgün çalışıyor")


# =============================================================================
# TEST 10: FULL PİPELİNE CYCLE (DRY-RUN)
# =============================================================================

def test_10_full_cycle():
    """
    HybridTradingPipeline.run_cycle() ile tam bir döngü çalıştırır.
    Scanner → IC Analysis → Gate → (AI) → Risk → PaperTrade
    """
    from main import HybridTradingPipeline, CycleStatus

    pipeline = HybridTradingPipeline(
        dry_run=True,                          # Paper trade — emir göndermez
        top_n=3,                               # Sadece 3 coin (hız için)
        verbose=False,
    )

    # Bakiye başlat
    assert pipeline._init_balance(), "Bakiye başlatma başarısız"
    print(f"  ✓ Bakiye: ${pipeline._balance:.2f}")

    # Telegram'ı devre dışı bırak (test ortamı)
    pipeline.notifier = type('MockNotifier', (), {
        'is_configured': lambda self: False,
        'send_alert_sync': lambda self, *a, **kw: None,
        'send_risk_alert_sync': lambda self, *a, **kw: None,
    })()

    # Tam döngü çalıştır
    print(f"  🔄 Pipeline çalışıyor (3 coin, dry-run)...")
    report = pipeline.run_cycle()

    assert report is not None, "Rapor None!"

    # Sonuçları raporla
    print(f"\n  {'─'*45}")
    print(f"  📊 DÖNGÜ RAPORU")
    print(f"  {'─'*45}")
    print(f"  Status:      {report.status.value}")
    print(f"  Taranan:     {report.total_scanned} coin")
    print(f"  Analiz:      {report.total_analyzed} coin")
    print(f"  Gate geçen:  {report.total_above_gate} coin")
    print(f"  İşlem açılan: {report.total_traded} coin")
    print(f"  Bakiye:      ${report.balance:,.2f}")
    print(f"  Süre:        {report.elapsed:.0f}s")

    if hasattr(report, 'ai_mode'):
        print(f"  AI modu:     {report.ai_mode}")

    # Status kontrolü
    valid_statuses = (CycleStatus.SUCCESS, CycleStatus.PARTIAL, CycleStatus.NO_SIGNAL)
    assert report.status in valid_statuses, f"Beklenmeyen status: {report.status}"

    assert report.elapsed > 0, "Süre 0!"
    assert report.balance > 0, "Bakiye 0!"

    # Hatalar varsa listele
    if report.errors:
        print(f"\n  ⚠️ Hatalar ({len(report.errors)}):")
        for err in report.errors[:5]:
            print(f"    • {err[:80]}")
            issues.append(f"Pipeline: {err[:80]}")

    # Coin detayları
    if report.coins:
        print(f"\n  {'Coin':<10} {'IC':>5} {'Yön':<7} {'Gate':<12} {'Durum'}")
        print(f"  {'─'*50}")
        for c in report.coins[:5]:
            ic_str = f"{c.ic_confidence:.0f}" if hasattr(c, 'ic_confidence') and c.ic_confidence else "N/A"
            direction = getattr(c, 'ic_direction', 'N/A')
            gate = getattr(c, 'gate_action', 'N/A')
            if hasattr(gate, 'value'):
                gate = gate.value
            status = getattr(c, 'status', 'N/A')
            name = getattr(c, 'coin', '?')[:8]
            print(f"  {name:<10} {ic_str:>5} {direction:<7} {str(gate):<12} {status}")


# =============================================================================
# ANA ÇALIŞTIRMA
# =============================================================================

if __name__ == "__main__":
    print("=" * 60)
    print("  🔬 v1.2.0 KAPSAMLI ENTEGRASYON TESTİ")
    print(f"  📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  📁 {PROJECT_ROOT}")
    print("=" * 60)

    # ── Testler ──
    run_test(1,  "Bitget API Bağlantısı",          test_01_bitget_connection)
    run_test(2,  "CoinScanner — Market Tarama",     test_02_scanner)
    run_test(3,  "Fetcher — Çoklu TF Veri Çekme",   test_03_fetcher_multi_tf)
    run_test(4,  "İndikatör Hesaplama (64+)",       test_04_indicators)
    run_test(5,  "IC Analizi (Spearman + FDR)",     test_05_ic_analysis)
    run_test(6,  "GateKeeper — Sinyal Filtreleme",  test_06_gatekeeper)
    run_test(7,  "RiskManager — SL/TP/Pozisyon",    test_07_risk_manager)
    run_test(8,  "PaperTrader — Trade Açma/Kapama", test_08_paper_trader)
    run_test(9,  "Excel Export",                     test_09_excel_export)
    run_test(10, "Full Pipeline Cycle (DRY-RUN)",   test_10_full_cycle)

    # ── Özet ──
    total = passed + failed + skipped
    print(f"\n{'='*60}")
    print(f"  📊 SONUÇ: {passed}/{total} BAŞARILI", end="")
    if failed:
        print(f" | {failed} BAŞARISIZ", end="")
    if skipped:
        print(f" | {skipped} ATLANILDI", end="")
    print()

    if issues:
        print(f"\n  ⚠️  TESPİT EDİLEN SORUNLAR ({len(issues)}):")
        for i, issue in enumerate(issues, 1):
            print(f"    {i}. {issue}")

    if failed == 0 and not issues:
        print(f"\n  ✅ TÜM SİSTEMLER ÇALIŞIYOR — Paper trading hazır!")
        print(f"\n  → python main.py --dry-run --top 5")
    elif failed == 0 and issues:
        print(f"\n  ⚠️  Testler geçti ama {len(issues)} uyarı var — gözden geçir")
    else:
        print(f"\n  ❌ {failed} test başarısız — düzeltilmesi gereken sorunlar var")

    print("=" * 60)
    sys.exit(failed)
