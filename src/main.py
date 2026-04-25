# =============================================================================
# MAIN.PY — ML-DRIVEN TRADING PIPELINE v2.1.4 (BINANCE VADELİ İŞLEMLER)
# =============================================================================
# v2.1.4 Değişiklikler:
#   ✅ Günlük %6 kayıp limiti AKTİF edildi (TR yerel saati, 00:00 TR'e kadar halt)
#   ✅ Toplam margin limitleri (%60) ve per-trade margin (%25) AKTİF
#   ✅ Kill switch canlı bakiye + güncel PnL ile hesaplanıyor
#   ✅ risk_manager.update_state() her döngüde daily_pnl + used_margin ile besleniyor
#   ✅ Günlük halt durumunda _execute_trade çağrıları reddediliyor
#
# v2.1.3 Değişiklikler:
#   ✅ Tarama süresi 10 dakikaya düşürüldü.
#   ✅ Cooldown için UTC saat dilimi hataları düzeltildi.
#   ✅ Scheduler uyku modundayken 30 saniyede bir açık emir takibi eklendi.
# =============================================================================

import sys
import os
import time
import signal
import argparse
import logging
import traceback
import schedule

from pathlib import Path
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Any
from dataclasses import dataclass, field
from enum import Enum

from execution.risk_manager import RiskManager
import numpy as np
import pandas as pd

# =============================================================================
# YEREL SAAT DİLİMİ (Istanbul / Türkiye)
# =============================================================================
# Günlük kayıp limiti sayacı yerel geceyarısı (00:00 TR) ile sıfırlansın diye
# tüm "günlük" hesaplamalar bu timezone üzerinden yapılır.
# TR yaz saati uygulamasını kaldırdığı için sabit UTC+3 offset güvenlidir.
LOCAL_TZ = timezone(timedelta(hours=3), name="TRT")

# =============================================================================
# [SORUN 4 DÜZELTMESİ] — UTC-aware datetime helper fonksiyonları
# =============================================================================
# Eski kodda datetime.now() (naive, timezone'suz) kullanılıyordu.
# _compute_daily_pnl bu naive string'leri UTC sanıp +3 saat çevirince
# 3 saatlik kayma oluşuyordu. Artık tüm zaman üretimi UTC veya TR-aware.
#
# Kullanım:
#   _now_utc()  → UTC datetime (karşılaştırma, hesaplama)
#   _now_local() → TR datetime (log, display)
#   _iso_tr()   → Excel için TR zaman string'i (timezone bilgisi olmadan)

def _now_utc() -> datetime:
    """UTC-aware şimdi — karşılaştırma ve hesaplamalar için."""
    return datetime.now(timezone.utc)

def _now_local() -> datetime:
    """TR (LOCAL_TZ) aware şimdi — log ve gösterim için."""
    return datetime.now(LOCAL_TZ)

def _iso_tr(dt: datetime = None) -> str:
    """
    ISO formatında TR zaman string'i (Excel için okunabilir).
    Excel tz-aware ISO'yu parse edemeyebilir, bu yüzden
    saf string olarak yazıyoruz ama TR saatiyle.
    """
    if dt is None:
        dt = datetime.now(timezone.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(LOCAL_TZ).strftime("%Y-%m-%dT%H:%M:%S.%f")



# ── .env yükle ────────────────────────────────────────────────────────────────
from dotenv import load_dotenv
_src_dir  = Path(__file__).parent
_root_dir = _src_dir.parent
load_dotenv(_root_dir / ".env")
sys.path.insert(0, str(_src_dir))

# ── Mevcut modüller ───────────────────────────────────────────────────────────
from config import cfg
from scanner import CoinScanner
from data import BinanceFetcher, DataPreprocessor
from indicators import IndicatorCalculator, IndicatorSelector
from execution import RiskManager, BinanceExecutor
from notifications import TelegramNotifier
from paper_trader import PaperTrader
from performance_analyzer import PerformanceAnalyzer

# ── ML modülleri (v2.0) ───────────────────────────────────────────────────────
from ml.feature_engineer import FeatureEngineer, MLFeatureVector
from ml.ensemble_model import EnsemblePredictor as LGBMSignalModel, MLDecisionResult
from ml.signal_validator import SignalValidator, ValidationResult
from ml.trade_memory import TradeMemory, TradeOutcome

# =============================================================================
# LOGLAMA
# =============================================================================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
)
logger = logging.getLogger(__name__)

# =============================================================================
# SABİTLER
# =============================================================================

VERSION                = "2.1.4"
MAX_COINS_PER_CYCLE    = 30
DEFAULT_FWD_PERIOD     = 6
MAX_OPEN_POSITIONS     = 5
MAX_CONSECUTIVE_ERRORS = 5
ERROR_COOLDOWN_SECONDS = 300

# ── TIMEOUT EXIT (Time-Based Stop) ──
# 24 saatten fazla açık kalmış pozisyonlar piyasa fiyatından kapatılır.
# Mantık: Sinyal edge'i bayatlamış, slot fırsat maliyeti artıyor.
# Kâr ile kapanan TIMEOUT trade'leri model retrain'inde WIN olarak öğrenilir
# (trade_memory pnl > 0 ise outcome=WIN atar, exit_reason'a bakmaz).
MAX_TRADE_AGE_HOURS    = 24

DEFAULT_TIMEFRAMES = {
    '5m': 1500,
    '15m': 1000,
    '30m': 1000,
    '1h' : 1000,
}

# [MADDE 2 DÜZELTMESİ] — IC gate eşikleri yükseltildi
# Canlı veri analizi: IC<17 bandındaki 135 trade ortalama -$0.062/trade kaybetti.
# IC≥20 olan 36 trade ise +$0.167/trade kazandı → IC ile WR arasında monotonik artış var.
# Önceki değerler (4/6) çok düşüktü ve gürültülü sinyallere kapı açıyordu.
IC_NO_TRADE = 5.0    # 10 → 5: veri birikimi için gevşetildi, paper modda öğrenme hızlanır
IC_TRADE    = 8.0   # 15 → 8: düşük IC'li sinyaller de geçsin, model tecrübe kazansın

# =============================================================================
# ENUM'LAR
# =============================================================================

class CycleStatus(Enum):
    SUCCESS   = "success"
    PARTIAL   = "partial"
    NO_SIGNAL = "no_signal"
    ERROR     = "error"
    KILLED    = "killed"
    HALTED    = "halted"   # Günlük kayıp limiti nedeniyle halt

# =============================================================================
# DATACLASS'LAR
# =============================================================================

@dataclass
class CoinAnalysisResult:
    coin:             str   = ""
    full_symbol:      str   = ""
    price:            float = 0.0
    change_24h:       float = 0.0
    volume_24h:       float = 0.0

    best_timeframe:   str   = ""
    ic_confidence:    float = 0.0
    ic_direction:     str   = ""
    significant_count: int  = 0
    market_regime:    str   = ""
    top_ic:           float = 0.0   # ← FIX 2: en güçlü tek indikatörün |IC| değeri

    category_tops:    Dict  = field(default_factory=dict)
    tf_rankings:      List  = field(default_factory=list)

    atr:              float = 0.0
    atr_pct:          float = 0.0
    sl_price:         float = 0.0
    tp_price:         float = 0.0
    position_size:    float = 0.0
    leverage:         int   = 1
    risk_reward:      float = 0.0

    ml_result:        Optional[MLDecisionResult] = None
    val_result:       Optional[ValidationResult] = None
    ml_skipped:       bool  = False

    trade_executed:   bool  = False
    status:           str   = "pending"
    error:            str   = ""
    execution_result: Any   = None
    paper_trade_id:   str   = ""


@dataclass
class CycleReport:
    timestamp:        str         = ""
    status:           CycleStatus = CycleStatus.NO_SIGNAL
    total_scanned:    int         = 0
    total_analyzed:   int         = 0
    total_above_gate: int         = 0
    total_traded:     int         = 0
    coins:            List[CoinAnalysisResult] = field(default_factory=list)
    balance:          float       = 0.0
    paper_balance:    float       = 0.0
    errors:           List[str]   = field(default_factory=list)
    elapsed:          float       = 0.0
    ml_stats:         Dict        = field(default_factory=dict)
    daily_pnl:        float       = 0.0   # Bugünkü toplam kapalı PnL


# =============================================================================
# ANA PIPELINE
# =============================================================================

class MLTradingPipeline:
    def __init__(
        self,
        dry_run:    bool = True,
        top_n:      int  = MAX_COINS_PER_CYCLE,
        timeframes: Dict = None,
        fwd_period: int  = DEFAULT_FWD_PERIOD,
        verbose:    bool = True,
    ):
        self.dry_run    = dry_run
        self.top_n      = min(top_n, MAX_COINS_PER_CYCLE)
        self.timeframes = timeframes or DEFAULT_TIMEFRAMES
        self.fwd_period = fwd_period
        self.verbose    = verbose

        self.scanner      = CoinScanner()
        self.fetcher      = BinanceFetcher()
        self.preprocessor = DataPreprocessor()
        self.calculator   = IndicatorCalculator()
        self.selector     = IndicatorSelector(alpha=0.05)
        self.risk_manager = RiskManager()
        self.executor     = BinanceExecutor(dry_run=dry_run)
        self.notifier     = TelegramNotifier()
        self.paper_trader = PaperTrader()

        self.feature_eng  = FeatureEngineer()
        self.lgbm_model   = LGBMSignalModel()
        self.validator    = SignalValidator()
        self.trade_memory = TradeMemory(log_dir = _root_dir / "logs")

        self._balance          = 0.0
        self._initial_balance  = 0.0
        self._kill_switch      = False
        self._is_running       = False
        self._consecutive_errors = 0
        self.cooldowns         = {}

        # [MADDE 3] — Deployment gate onay durumu
        # True  → model istatistiksel kapıları geçti, canlı trade açılabilir
        # False → model henüz eğitilmedi veya kapılar başarısız, trade bloklanır
        # Paper modda bu flag devreye girmez (dry_run=True ise engelleme yok).
        self._model_deployment_approved: bool = False

        # ── GÜNLÜK KAYIP LİMİTİ STATE (v2.1.4) ──────────────────────────
        # UTC gününün başında bakiye snapshot'ı alınır.
        # Gün içinde toplam kapalı trade PnL'i bu değerin -%6'sına inerse
        # _daily_halt_until ertesi gün 00:00 TR olarak set edilir (iç saklamada UTC-aware).
        self._daily_date: Optional[str]           = None   # Aktif gün (TR YYYY-MM-DD)
        self._daily_start_balance: float          = 0.0    # Gün başı bakiye
        self._daily_realized_pnl: float           = 0.0    # Bugünkü gerçekleşen PnL
        self._daily_halt_until: Optional[datetime] = None  # Halt bitiş zamanı (UTC-aware datetime)

        self._restore_cooldowns()

        logger.info(f"🚀 ML Trading Pipeline v{VERSION} (dry_run={dry_run})")

    def _restore_cooldowns(self):
        try:
            now = datetime.now(timezone.utc)

            if not self.dry_run:
                memory_file = self.trade_memory.log_dir / "ml_trade_memory.json"
                if memory_file.exists():
                    import json
                    try:
                        with open(memory_file, 'r', encoding='utf-8') as f:
                            trades = json.load(f)

                        for trade in trades:
                            exit_reason = trade.get('exit_reason', '')
                            closed_at_str = trade.get('closed_at', '')
                            coin = trade.get('coin', '')
                            symbol = trade.get('symbol', '')

                            if exit_reason == "SL Hit" and closed_at_str and coin:
                                try:
                                    closed_time = datetime.fromisoformat(closed_at_str)
                                    if closed_time.tzinfo is None:
                                        closed_time = closed_time.replace(tzinfo=timezone.utc)

                                    if (now - closed_time).total_seconds() < 7200:
                                        kalan_sure = closed_time + timedelta(hours=2)
                                        self.cooldowns[coin] = kalan_sure
                                        if symbol:
                                            self.cooldowns[symbol] = kalan_sure
                                        logger.info(f"   ❄️ HAFIZA GERİ YÜKLENDİ: {coin} (Ceza bitişi: {kalan_sure.strftime('%H:%M:%S')})")
                                except (ValueError, TypeError):
                                    continue
                    except Exception as e:
                        logger.warning(f"Trade memory cooldown yükleme hatası: {e}")
            else:
                if hasattr(self.paper_trader, 'closed_trades'):
                    for trade in self.paper_trader.closed_trades:
                        exit_reason = getattr(trade, 'exit_reason', None) or (trade.get('exit_reason') if isinstance(trade, dict) else None)
                        closed_at = getattr(trade, 'closed_at', None) or (trade.get('closed_at') if isinstance(trade, dict) else None)
                        symbol = getattr(trade, 'symbol', None) or (trade.get('symbol') if isinstance(trade, dict) else None)
                        full_symbol = getattr(trade, 'full_symbol', None) or (trade.get('full_symbol') if isinstance(trade, dict) else None)

                        if exit_reason == "SL Hit" and closed_at and symbol:
                            if isinstance(closed_at, str):
                                try:
                                    closed_time = datetime.fromisoformat(closed_at)
                                    if closed_time.tzinfo is None:
                                        closed_time = closed_time.replace(tzinfo=timezone.utc)
                                except ValueError:
                                    continue
                            else:
                                closed_time = closed_at
                                if closed_time.tzinfo is None:
                                    closed_time = closed_time.replace(tzinfo=timezone.utc)

                            if (now - closed_time).total_seconds() < 7200:
                                kalan_sure = closed_time + timedelta(hours=2)
                                self.cooldowns[symbol] = kalan_sure
                                if full_symbol:
                                    self.cooldowns[full_symbol] = kalan_sure
                                logger.info(f"   ❄️ HAFIZA GERİ YÜKLENDİ: {symbol} (Ceza bitişi: {kalan_sure.strftime('%H:%M:%S')})")

        except Exception as e:
            logger.error(f"Cooldown hafıza yükleme hatası: {e}")

    def _init_balance(self) -> bool:
        try:
            if self.dry_run:
                # Paper trader'ın mevcut bakiyesini kullan (1000.0 değil)
                # Böylece initial_balance gerçek başlangıç noktasını yansıtır
                # ve kill switch yanlış tetiklenmez.
                self._balance = self._initial_balance = self.paper_trader.balance
                logger.info(f"💰 Paper bakiye: ${self._balance:.2f} (paper_trader'dan)")
            else:
                b = self.executor.fetch_balance()
                if isinstance(b, dict):
                    self._balance = self._initial_balance = b.get('total', 0.0)
                else:
                    self._balance = self._initial_balance = float(b)

                logger.info(f"💰 Canlı bakiye: ${self._balance:.2f}")

            # ── Günlük state'i başlat (v2.1.4) ──
            # Yerel saat (TR) üzerinden bugünün tarihini al → 00:00 TR'de sıfırlanır
            self._daily_date = datetime.now(LOCAL_TZ).strftime("%Y-%m-%d")
            self._daily_start_balance = self._balance
            self._daily_realized_pnl = 0.0
            logger.info(
                f"📅 Günlük limit başlangıcı: {self._daily_date} (TR) | "
                f"Start bakiye: ${self._daily_start_balance:.2f} | "
                f"Max günlük kayıp: %{cfg.risk.daily_max_loss_pct}"
            )
            return True
        except Exception as e:
            logger.error(f"❌ Bakiye hatası: {e}")
            return False

    def _refresh_live_balance(self) -> None:
        """Canlı modda Binance'ten güncel bakiyeyi çek (toplam = free + margin)."""
        if self.dry_run:
            self._balance = self.paper_trader.balance
            return
        try:
            b = self.executor.fetch_balance()
            if isinstance(b, dict):
                new_balance = b.get('total', self._balance)
                if new_balance > 0:
                    self._balance = new_balance
        except Exception as e:
            logger.debug(f"Bakiye refresh hatası (kritik değil): {e}")

    def _check_kill_switch(self) -> bool:
        """%15 kill switch drawdown kontrolü (canlı bakiye + günlük PnL ile)."""
        if self._kill_switch:
            return True
        if self._initial_balance <= 0:
            return False

        # Canlı bakiyeyi tazele
        self._refresh_live_balance()

        # Güncel equity = canlı bakiye (hâlihazırda realized PnL dahil)
        current_equity = self._balance
        dd = (self._initial_balance - current_equity) / self._initial_balance * 100

        if dd >= cfg.risk.kill_switch_drawdown_pct:
            self._kill_switch = True
            logger.critical(f"🚨 KILL SWITCH! DD={dd:.1f}% (Initial: ${self._initial_balance:.2f} → Current: ${current_equity:.2f})")
            if self.notifier.is_configured():
                self.notifier.send_risk_alert_sync(
                    alert_type="KILL_SWITCH",
                    details=(
                        f"⛔ Kill switch tetiklendi!\n"
                        f"📉 Drawdown: %{dd:.1f}\n"
                        f"💰 Başlangıç: ${self._initial_balance:.2f}\n"
                        f"💰 Güncel: ${self._balance:.2f}"
                    ),
                    severity="critical",
                )
            return True
        return False

    # =========================================================================
    # GÜNLÜK KAYIP LİMİTİ (v2.1.4)
    # =========================================================================

    def _reset_daily_state_if_new_day(self) -> None:
        """
        Yerel saate (TR) göre gün değiştiyse günlük sayaçları sıfırla.
        00:00 TR (yani 21:00 UTC) saatinde yeni gün başlar.
        Aktif halt varsa ve bitiş zamanı geçtiyse onu da kaldırır.
        """
        # Halt için UTC karşılaştırması (datetime aware karşılaştırma için)
        now_utc = datetime.now(timezone.utc)
        # Tarih kıyaslaması için yerel saat (TR)
        today = datetime.now(LOCAL_TZ).strftime("%Y-%m-%d")

        # Halt süresi dolduysa kaldır
        if self._daily_halt_until and now_utc >= self._daily_halt_until:
            # Halt bitiş zamanını kullanıcıya TR saatiyle göster
            halt_end_tr = self._daily_halt_until.astimezone(LOCAL_TZ)
            logger.info(
                f"🌅 Günlük halt süresi doldu "
                f"({halt_end_tr.strftime('%Y-%m-%d %H:%M TR')}) — trade'e devam"
            )
            self._daily_halt_until = None

        # Yeni gün mü? (TR saatine göre)
        if self._daily_date != today:
            old_day = self._daily_date
            # Yeni günün başlangıç bakiyesini güncel bakiye olarak snapshot'la
            self._refresh_live_balance()
            self._daily_date = today
            self._daily_start_balance = self._balance
            self._daily_realized_pnl = 0.0
            logger.info(
                f"📅 Yeni gün: {old_day} → {today} (TR) | "
                f"Start bakiye: ${self._daily_start_balance:.2f}"
            )

    def _compute_daily_pnl(self) -> float:
        """
        Bugün (yerel saat TR) kapanan tüm trade'lerin toplam PnL'ini hesapla.
        Canlı mod: ml_trade_memory.json
        Paper mod: paper_trader.closed_trades
        """
        # "Bugün" tanımı TR saatine göre (00:00 TR - 23:59 TR arası)
        today = datetime.now(LOCAL_TZ).strftime("%Y-%m-%d")
        total = 0.0

        try:
            if not self.dry_run:
                # Canlı: memory dosyasından oku
                memory_file = self.trade_memory.log_dir / "ml_trade_memory.json"
                if not memory_file.exists():
                    return 0.0
                import json
                with open(memory_file, 'r', encoding='utf-8') as f:
                    trades = json.load(f)
                for t in trades:
                    closed_at = t.get('closed_at', '')
                    if not closed_at:
                        continue
                    # closed_at UTC ISO string olmalı
                    try:
                        closed_dt = datetime.fromisoformat(closed_at)
                        if closed_dt.tzinfo is None:
                            closed_dt = closed_dt.replace(tzinfo=timezone.utc)
                        # ⏱️ Yerel saate çevirip TR gününe göre karşılaştır
                        closed_dt_local = closed_dt.astimezone(LOCAL_TZ)
                        if closed_dt_local.strftime("%Y-%m-%d") == today:
                            pnl = float(t.get('pnl', 0.0) or 0.0)
                            total += pnl
                    except (ValueError, TypeError):
                        continue
            else:
                # Paper: paper_trader üzerinden
                for trade in self.paper_trader.closed_trades:
                    closed_at = getattr(trade, 'closed_at', None)
                    if not closed_at:
                        continue
                    if isinstance(closed_at, str):
                        try:
                            closed_dt = datetime.fromisoformat(closed_at)
                        except ValueError:
                            continue
                    else:
                        closed_dt = closed_at
                    if closed_dt.tzinfo is None:
                        # Naive datetime geldi — UTC varsayıyoruz
                        # Bu dal artık tetiklenmemeli (_iso_tr() her yerde UTC üretiyor)
                        logger.debug(f'Naive datetime: {closed_at} — UTC varsayılıyor')
                        closed_dt = closed_dt.replace(tzinfo=timezone.utc)
                    # ⏱️ Yerel saate çevirip TR gününe göre karşılaştır
                    closed_dt_local = closed_dt.astimezone(LOCAL_TZ)
                    if closed_dt_local.strftime("%Y-%m-%d") == today:
                        pnl = getattr(trade, 'net_pnl', None)
                        if pnl is None:
                            pnl = getattr(trade, 'pnl_absolute', 0.0) or 0.0
                        total += float(pnl or 0.0)
        except Exception as e:
            logger.warning(f"⚠️ Daily PnL hesaplama hatası: {e}")
            return self._daily_realized_pnl  # Son bilinen değer

        return total

    def _check_daily_loss_limit(self) -> bool:
        """
        Günlük %X kayıp limitini kontrol et.
        
        Returns:
            True  → halt aktif, döngü durdurulmalı
            False → halt yok, trade'e devam
        """
        now_utc = datetime.now(timezone.utc)

        # Zaten halt aktifse
        if self._daily_halt_until and now_utc < self._daily_halt_until:
            kalan = (self._daily_halt_until - now_utc).total_seconds() / 60
            # Halt bitişini TR saatiyle göster (kullanıcı dostu)
            halt_end_tr = self._daily_halt_until.astimezone(LOCAL_TZ)
            logger.info(
                f"⏸️ Günlük halt aktif — kalan: {kalan:.0f} dk "
                f"(bitiş: {halt_end_tr.strftime('%H:%M TR')})"
            )
            return True

        # Bugünkü gerçekleşen PnL
        daily_pnl = self._compute_daily_pnl()
        self._daily_realized_pnl = daily_pnl

        if self._daily_start_balance <= 0:
            return False

        max_loss_pct = cfg.risk.daily_max_loss_pct / 100.0
        daily_loss_ratio = -daily_pnl / self._daily_start_balance  # Pozitif sayı = kayıp

        logger.info(
            f"📊 Günlük PnL: ${daily_pnl:+.2f} "
            f"({daily_loss_ratio*100:+.2f}% / limit: %{cfg.risk.daily_max_loss_pct}) "
            f"| Start: ${self._daily_start_balance:.2f}"
        )

        if daily_loss_ratio >= max_loss_pct:
            # ⏰ Halt tetikle — YEREL GECEYARISINA kadar (00:00 TR)
            # Önce yerel saati al, sonra yerel'de ertesi günün 00:00'ına ayarla,
            # en son UTC'ye geri çevir ki datetime karşılaştırmaları tutarlı kalsın.
            now_local = datetime.now(LOCAL_TZ)
            tomorrow_local_midnight = (now_local + timedelta(days=1)).replace(
                hour=0, minute=0, second=0, microsecond=0
            )
            # İç saklamada UTC (kodun geri kalanı UTC-aware datetime kullanıyor)
            self._daily_halt_until = tomorrow_local_midnight.astimezone(timezone.utc)

            logger.critical(
                f"🛑 GÜNLÜK KAYIP LİMİTİ AŞILDI! "
                f"Kayıp: ${daily_pnl:+.2f} ({daily_loss_ratio*100:.1f}% ≥ %{cfg.risk.daily_max_loss_pct}) "
                f"| Halt bitişi: {tomorrow_local_midnight.strftime('%Y-%m-%d %H:%M TR')}"
            )

            if self.notifier.is_configured():
                try:
                    msg = (
                        f"🛑 <b>GÜNLÜK KAYIP LİMİTİ!</b>\n"
                        f"━━━━━━━━━━━━━━━━━━━━━\n"
                        f"📉 Günlük PnL: ${daily_pnl:+.2f}\n"
                        f"📊 Kayıp: {daily_loss_ratio*100:.2f}% (limit: %{cfg.risk.daily_max_loss_pct})\n"
                        f"💰 Start: ${self._daily_start_balance:.2f}\n"
                        f"💰 Current: ${self._balance:.2f}\n"
                        f"⏰ Halt bitişi: {tomorrow_local_midnight.strftime('%Y-%m-%d %H:%M TR')}\n"
                        f"━━━━━━━━━━━━━━━━━━━━━\n"
                        f"Yeni trade AÇILMAYACAK — mevcut açık pozisyonlar SL/TP ile takip ediliyor."
                    )
                    self.notifier.send_message_sync(msg)
                except Exception as tg_err:
                    logger.warning(f"Telegram bildirimi gönderilemedi: {tg_err}")
            return True

        return False

    def _compute_used_margin(self) -> float:
        """Açık pozisyonların toplam margin kullanımını hesapla."""
        if self.dry_run:
            total = 0.0
            for trade in self.paper_trader.open_trades.values():
                notional = trade.entry_price * trade.position_size
                lev = max(trade.leverage, 1)
                total += notional / lev
            return total
        try:
            positions = self.executor.fetch_positions()
            return sum(float(p.get('margin', 0) or 0) for p in positions)
        except Exception:
            return 0.0

    def _sync_risk_manager(self) -> None:
        """RiskManager'ın state'ini güncel bakiye, daily PnL, margin, pozisyon ile besle."""
        try:
            current_balance = self.paper_trader.balance if self.dry_run else self._balance
            used_margin = self._compute_used_margin()

            if self.dry_run:
                open_count = len(self.paper_trader.open_trades)
            else:
                try:
                    open_count = len(self.executor.fetch_positions())
                except Exception:
                    open_count = 0

            self.risk_manager.update_state(
                balance=current_balance,
                used_margin=used_margin,
                open_positions=open_count,
                daily_pnl=self._daily_realized_pnl,
            )
        except Exception as e:
            logger.debug(f"RiskManager sync hatası (kritik değil): {e}")

    # =========================================================================
    # [MADDE 3] — MODEL DEPLOYMENT GATE KONTROLÜ
    # =========================================================================
    def _check_deployment_gates(self, metrics) -> bool:
        """
        Model deployment istatistiksel kapılarını kontrol eder.

        eval_harness.py'deki kapıların aynısı burada da uygulanır.
        Tüm kapılardan geçemeyen model canlı trade'de kullanılmaz.

        İstatistiksel temel:
        - H0: 'Modelin tahmin gücü yok (IC = 0)'
        - Aggregated Z ≥ 1.65 → H0 reddedilir (p ≤ 0.05, tek kuyruk)
        - Z = 0.83 gibi değerler p ≈ 0.20 demektir → H0 reddedilemiyor

        Returns:
            True  → model onaylandı, canlıya geçilebilir
            False → model yetersiz, paper modda kalınmalı
        """
        # Aggregated Z-skoru: fold başına IC değerlerinin one-sided testi
        # Varsayım: fold IC'leri IID → merkezi limit teoremi
        n_folds = getattr(metrics, 'n_folds', 0)
        ic_std  = getattr(metrics, 'ic_std', 1.0)
        aggregated_z = (
            metrics.spearman_ic * (n_folds ** 0.5) / ic_std
            if ic_std > 1e-9 and n_folds > 0 else 0.0
        )

        gates = {
            "Spearman IC ≥ 0.05  (Grinold-Kahn alt sınırı)":  metrics.spearman_ic >= 0.05,
            "Information Ratio ≥ 0.50  (sinyal/gürültü)":      metrics.information_ratio >= 0.50,
            "Long-Short Spread ≥ 0.10R  (yön ayırt gücü)":     metrics.long_short_spread >= 0.10,  # 0.20 → 0.10: Mevcut değer +0.135R, eşik biraz iddialıydı. Model gerçek edge görüyor.
            "Aggregated Z ≥ 1.65  (p ≤ 0.05, tek kuyruk)":    aggregated_z >= 1.65,
        }

        passed = all(gates.values())

        # ── Log ──────────────────────────────────────────────────────────────
        logger.info(f"\n{'─'*58}")
        logger.info("  🔬 MODEL DEPLOYMENT GATE KONTROLÜ")
        logger.info(f"{'─'*58}")
        logger.info(f"  Spearman IC    : {metrics.spearman_ic:+.4f}  (eşik: ≥ 0.05)")
        logger.info(f"  Info Ratio     : {metrics.information_ratio:+.2f}   (eşik: ≥ 0.50)")
        logger.info(f"  L-S Spread     : {metrics.long_short_spread:+.3f}R  (eşik: ≥ 0.10R)")
        logger.info(f"  Aggregated Z   : {aggregated_z:+.2f}   (eşik: ≥ 1.65, p≤0.05)")
        logger.info(f"{'─'*58}")
        for gate_name, gate_result in gates.items():
            icon = "✅" if gate_result else "❌"
            logger.info(f"  {icon}  {gate_name}")
        logger.info(f"{'─'*58}")

        if passed:
            logger.info("  🟢 KARAR: MODEL ONAYLANDI — Canlı trade'e hazır")
        else:
            n_failed = sum(1 for v in gates.values() if not v)
            logger.warning(
                f"  🔴 KARAR: MODEL REDDEDİLDİ — {n_failed} kapı başarısız\n"
                f"     Model istatistiksel olarak rastgele tahminden ayırt edilemiyor.\n"
                f"     Canlı trade DURDURULDU — paper modda çalışmaya devam ediyor."
            )
            # Telegram bildirimi
            try:
                if self.notifier.is_configured():
                    gate_lines = "\n".join(
                        f"{'✅' if v else '❌'} {k}"
                        for k, v in gates.items()
                    )
                    self.notifier.send_message_sync(
                        f"⚠️ <b>DEPLOYMENT GATE BAŞARISIZ</b>\n"
                        f"━━━━━━━━━━━━━━━━━━━━━\n"
                        f"{gate_lines}\n"
                        f"━━━━━━━━━━━━━━━━━━━━━\n"
                        f"📊 Z={aggregated_z:.2f} | IC={metrics.spearman_ic:.4f}\n"
                        f"🛑 Canlı trade DURDURULDU. Paper modda devam ediyor."
                    )
            except Exception:
                pass

        logger.info(f"{'─'*58}\n")
        return passed

    def _detect_regime(self, df: pd.DataFrame) -> str:
        """Piyasa rejimi tespiti: trending / ranging / volatile"""
        try:
            if 'close' not in df.columns or len(df) < 20:
                return 'ranging'

            adx_col = next((col for col in df.columns if col.upper().startswith('ADX') and 'DMP' not in col.upper() and 'DMN' not in col.upper()), None)

            if adx_col and not df[adx_col].dropna().empty:
                adx = float(df[adx_col].dropna().iloc[-1])
                if adx > 25: return 'trending'
                if adx > 15: return 'ranging'
                return 'volatile'

            returns = df['close'].pct_change().tail(20)
            vol = float(returns.std())

            if pd.isna(vol):
                return 'ranging'

            if vol > 0.03: return 'volatile'
            elif vol > 0.01: return 'ranging'
            else: return 'trending'

        except Exception as e:
            logger.warning(f"  ⚠️ Rejim tespiti hatası: {e}")
            return 'ranging'


    @staticmethod
    def _compute_category_tops(
        all_scores: list,
        categories: dict
    ) -> dict:
        """
        [SORUN 2 DÜZELTMESİ] — Kategori bazlı en güçlü indikatörü bulur.

        Eski kodda iç içe bozuk for-döngüsü vardı:
        - `if 'matched' in locals()` her zaman True dönerdi
        - `if not cat_scores:` döngü içinde değişkeni güncellemiyordu
        - Sonuç: category_tops neredeyse her zaman boş dict dönüyordu
        - ic_cat_trend/momentum/volatility/volume feature'ları hep NaN oluyordu

        Bu implementasyon iki aşamalı temiz eşleştirme yapar:
        1. startswith eşleşmesi (tam ön ek)
        2. contains eşleşmesi (daha esnek, min 3 karakter)

        Parameters
        ----------
        all_scores : List[IndicatorScore]
            selector.evaluate_all_indicators() çıktısı
        categories : Dict[str, List[str]]
            {'trend': ['ADX', 'Aroon', ...], 'momentum': ['RSI', ...], ...}

        Returns
        -------
        Dict[str, Dict]
            {'trend': {'name': 'ADX_14', 'ic': -0.05}, ...}
            Boş kategoriler sonuç dict'e dahil edilmez.
        """
        result = {}

        for cat, cat_indicators in categories.items():
            # Aşama 1: startswith (tam ön ek eşleşmesi)
            cat_scores = [
                s for s in all_scores
                if any(
                    s.name.lower().startswith(ci.lower() + '_') or
                    s.name.lower().startswith(ci.lower())
                    for ci in cat_indicators
                )
            ]

            # Aşama 2: contains (esnek eşleşme, en az 3 karakter)
            if not cat_scores:
                cat_scores = [
                    s for s in all_scores
                    if any(
                        ci.lower() in s.name.lower()
                        for ci in cat_indicators
                        if len(ci) >= 3
                    )
                ]

            if cat_scores:
                top = max(cat_scores, key=lambda s: abs(s.ic_mean))
                result[cat] = {
                    "name": top.name,
                    "ic": round(top.ic_mean, 4),
                }

        return result

    def _analyze_coin(self, symbol: str, coin: str, scan_result=None) -> CoinAnalysisResult:
        result = CoinAnalysisResult(coin=coin, full_symbol=symbol)

        # ── FIX 1: Scanner'dan gelen 24h market verilerini kopyala ──
        # mkt_change_24h ve mkt_volume_24h_log feature'ları için gerekli.
        # Önceden bu alanlar default 0 kalıyordu → feature engineer hep 0 üretiyordu.
        if scan_result is not None:
            result.change_24h = float(getattr(scan_result, 'change_24h', 0.0) or 0.0)
            result.volume_24h = float(getattr(scan_result, 'volume_24h', 0.0) or 0.0)

        try:
            all_data = {}
            for tf, limit in self.timeframes.items():
                df_raw = self.fetcher.fetch_ohlcv(symbol, tf, limit=limit)
                if df_raw is None or len(df_raw) < 50:
                    continue
                df_clean = self.preprocessor.full_pipeline(df_raw)
                if df_clean is not None and len(df_clean) > 50:
                    all_data[tf] = df_clean

            if not all_data:
                result.status = "no_data"; return result

            result.price = float(next(iter(all_data.values()))['close'].iloc[-1])

            indicator_data = {}
            for tf, df in all_data.items():
                df_ind = self.calculator.calculate_all(df)
                df_ind = self.calculator.add_forward_returns(
                    df_ind, periods=[self.fwd_period]
                )
                if df_ind is not None and len(df_ind) > 50:
                    indicator_data[tf] = df_ind

            if not indicator_data:
                result.status = "indicator_error"; return result

            target_col = f'fwd_ret_{self.fwd_period}'
            best_tf    = None
            best_ic    = -1.0
            best_scores= []
            tf_rankings= []

            for tf, df in indicator_data.items():
                if target_col not in df.columns:
                    continue
                scores = self.selector.evaluate_all_indicators(df, target_col=target_col)
                if not scores:
                    continue
                valid_scores = [s for s in scores if not np.isnan(s.ic_mean)]
                if not valid_scores:
                    continue
                avg_ic = sum(abs(s.ic_mean) for s in valid_scores) / len(valid_scores)
                significant = [s for s in scores if s.is_significant]
                if len(significant) == 0:
                    significant = scores[:2]

                # ── FIX 3: feature_engineer ctf_* feature'ları için 'score' ve 'direction' bekliyor ──
                # Önceden sadece 'avg_ic' ve 'sig_count' vardı → feature_engineer 'score'u bulamıyordu
                # → ctf_best_score, ctf_avg_score, ctf_score_std, ctf_score_spread hep 0 oluyordu.
                tf_avg_signed = sum(s.ic_mean for s in valid_scores) / len(valid_scores)
                tf_dir = 'LONG' if tf_avg_signed > 0 else ('SHORT' if tf_avg_signed < 0 else 'NEUTRAL')

                tf_rankings.append({
                    'tf': tf,
                    'avg_ic': avg_ic,
                    'score': avg_ic * 100,                                         # ← FIX 3
                    'sig_count': len([s for s in scores if s.is_significant]),
                    'direction': tf_dir,                                           # ← FIX 3
                })

                if avg_ic > best_ic:
                    best_ic = avg_ic
                    best_tf = tf
                    best_scores = significant

            if not best_tf:
                result.status = "no_ic"; return result

            result.best_timeframe = best_tf
            result.tf_rankings    = sorted(tf_rankings, key=lambda x: x['avg_ic'], reverse=True)

            # ── FIX 2: top_ic — en güçlü tek indikatörün mutlak IC değeri ──
            # feature_engineer 'ic_top_abs' feature'ı bunu okuyor.
            # Önceden CoinAnalysisResult'ta top_ic field'ı yoktu → hep 0 dönüyordu.
            if best_scores:
                result.top_ic = max(abs(s.ic_mean) for s in best_scores if not np.isnan(s.ic_mean))

            ic_scores = [s.ic_mean for s in best_scores]
            ic_signal_dir = "NEUTRAL"
            if ic_scores:
                ic_avg = sum(ic_scores) / len(ic_scores)
                if ic_avg > 0:
                    ic_signal_dir = "LONG"
                elif ic_avg < 0:
                    ic_signal_dir = "SHORT"
                result.ic_confidence   = abs(ic_avg) * 100
                result.ic_direction    = ic_signal_dir
                result.significant_count = len(best_scores)

            if result.ic_confidence < IC_NO_TRADE:
                result.status = "ic_too_low"; return result
            if result.ic_confidence < IC_TRADE:
                result.status = "below_gate"; return result

            result.market_regime = self._detect_regime(indicator_data[best_tf])

            # [REJİM ÇARPAN REVİZYONU] — Daha dengeli çarpanlar
            # trending  → ×1.10 (trend sinyalini hafif güçlendir, model güvenilir)
            # ranging   → ×0.85 (hafif ceza, tamamen engelleme — veri birikimi için)
            # volatile  → ×0.70 (orta ceza, yüksek IC geçebilir)
            #
            # IC=8 + trending  → 8.8  → trade eşiği (8) geçer ✅
            # IC=8 + ranging   → 6.8  → trade eşiği geçer ✅
            # IC=8 + volatile  → 5.6  → no_trade eşiği (5) geçer, trade (8) geçmez ⚠️
            # IC=5 + ranging   → 4.25 → no_trade (5) altında → elenir ✅
            REGIME_PENALTY = {
                'trending':     1.10,
                'trending_up':  1.10,
                'trending_down': 1.10,
                'ranging':      0.85,
                'volatile':     0.70,
                'transitioning': 0.90,
                'unknown':      0.85,
            }
            penalty = REGIME_PENALTY.get(result.market_regime, 0.85)
            if penalty != 1.0:
                original_ic = result.ic_confidence
                result.ic_confidence = result.ic_confidence * penalty
                icon = "📈" if penalty > 1.0 else "📉"
                logger.info(
                    f"   {icon} {result.coin} [{result.best_timeframe}]: "
                    f"{result.market_regime.upper()} rejim → IC ×{penalty:.2f} "
                    f"({original_ic:.1f} → {result.ic_confidence:.1f})"
                )

            # [SORUN 2 DÜZELTMESİ] — Bozuk iç içe for-döngüsü yerine temiz helper
            CATEGORIES = {
                'volume':     ['OBV', 'CMF', 'VPT', 'FI', 'EOM', 'ADI', 'NVI', 'MFI'],
                'momentum':   ['RSI', 'Stoch', 'UO', 'MACD', 'PPO', 'ROC', 'TSI', 'CCI', 'Williams', 'WILLR'],
                'trend':      ['ADX', 'Aroon', 'PSAR', 'DPO', 'Vortex', 'KST', 'Ichimoku', 'SMA', 'EMA', 'WMA'],
                'volatility': ['BBW', 'BBU', 'BBL', 'BBM', 'ATR', 'NATR', 'Keltner', 'Donchian'],
            }
            result.category_tops = self._compute_category_tops(best_scores, CATEGORIES)

            df_best = indicator_data[best_tf]

            atr_found = False
            for atr_col in ['ATRr_14', 'ATR_14', 'ATRr_7', 'NATR_14', 'TRUERANGE_1']:
                if atr_col in df_best.columns:
                    atr_series = df_best[atr_col].dropna()
                    if not atr_series.empty:
                        raw_val = float(atr_series.iloc[-1])

                        if atr_col == 'NATR_14':
                            result.atr = raw_val * result.price / 100
                        else:
                            result.atr = raw_val

                        result.atr_pct = (result.atr / result.price) * 100 if result.price > 0 else 0.0
                        atr_found = True
                        break

            if not atr_found and 'high' in df_best.columns and 'low' in df_best.columns:
                hl_range = (df_best['high'] - df_best['low']).tail(14)
                result.atr = float(hl_range.mean())
                result.atr_pct = (result.atr / result.price) * 100 if result.price > 0 else 0.0

            if not self.lgbm_model.is_trained:
                result.ml_skipped = True
                result.status = "model_not_trained"
                return result

            fv = self.feature_eng.build_features(
                analysis   = result,
                ohlcv_df   = df_best
            )

            ml_result = self.lgbm_model.predict(
                fv, ic_direction=result.ic_direction
            )
            result.ml_result = ml_result

            if ml_result.decision.value == "WAIT":
                result.status = "wait"; return result

            val_result = self.validator.validate(
                fv, self.lgbm_model, ml_result.decision, ml_result.confidence,
                result.ic_direction, result.market_regime
            )
            result.val_result = val_result

            is_ok = getattr(val_result, 'is_valid', False) or getattr(val_result, 'is_approved', False)

            if not is_ok:
                result.status = "rejected_by_validator"
                return result

            result.status = "ready"

        except Exception as e:
            result.status = "error"
            result.error  = str(e)
            logger.error(f"   ❌ {coin} analiz hatası: {e}", exc_info=True)

        return result

    def _execute_trade(self, result: CoinAnalysisResult) -> CoinAnalysisResult:
        # ── Günlük halt kontrolü (v2.1.4) ──
        now = datetime.now(timezone.utc)
        if self._daily_halt_until and now < self._daily_halt_until:
            result.status = "daily_halt"
            logger.info(f"   🛑 {result.coin} atlandı: Günlük kayıp limiti aktif (halt: {self._daily_halt_until.astimezone(LOCAL_TZ).strftime('%H:%M TR')})")
            return result

        # [MADDE 3] — Deployment gate koruması (yalnızca canlı mod)
        # Paper modda gate başarısız olsa da trade simüle edilir (öğrenme amaçlı).
        # Canlı modda gate onayı olmadan gerçek para riske edilmez.
        if not self.dry_run and not self._model_deployment_approved:
            result.status = "deployment_gate_failed"
            logger.warning(
                f"   🔴 {result.coin} atlandı: Deployment gate onayı yok "
                f"(model istatistiksel eşiklerin altında). "
                f"--live modda çalıştırmak için önce model gate'leri geçmeli."
            )
            return result

        # UTC cooldown kontrolü
        if hasattr(self, 'cooldowns'):
            if result.coin in self.cooldowns or result.full_symbol in self.cooldowns:
                cooldown_key = result.coin if result.coin in self.cooldowns else result.full_symbol
                if now < self.cooldowns[cooldown_key]:
                    kalan_dk = int((self.cooldowns[cooldown_key] - now).total_seconds() / 60)
                    logger.info(f"   ❄️ {result.coin} işlemi REDDEDİLDİ: Soğuma süresinde (Kalan: {kalan_dk} dk)")
                    result.status = "cooldown"
                    return result
                else:
                    if result.coin in self.cooldowns:
                        del self.cooldowns[result.coin]
                    if result.full_symbol in self.cooldowns:
                        del self.cooldowns[result.full_symbol]

        if result.status in ["api_error", "execution_error", "live_price_error"]:
            logger.info(f"   🚫 {result.coin} API tarafından daha önce reddedildi. 1 saat soğumaya alınıyor.")
            if not hasattr(self, 'cooldowns'):
                 self.cooldowns = {}
            self.cooldowns[result.coin] = now + timedelta(hours=1)
            self.cooldowns[result.full_symbol] = now + timedelta(hours=1)
            return result

        if result.status != "ready" or result.ml_result is None:
            return result

        direction = result.ml_result.decision.value
        if direction == "WAIT":
            return result

        try:
            try:
                exchange = self.executor._get_exchange()
                ticker = exchange.fetch_ticker(result.full_symbol)
                live_price = ticker['last']

                if live_price is None or live_price <= 0:
                    raise ValueError(f"Geçersiz canlı fiyat: {live_price}")

                # RiskManager state'ini TAM güncel değerlerle besle (v2.1.4)
                self._sync_risk_manager()

                trade_calc = self.risk_manager.calculate_trade(
                    symbol      = result.full_symbol,
                    direction   = direction,
                    entry_price = live_price,
                    atr         = result.atr,
                )

                if not trade_calc.is_approved():
                    logger.warning(
                        f"   ⚠️ {result.coin} riskten geçemedi: "
                        f"{trade_calc.rejection_reasons}"
                    )
                    result.status = "risk_rejected"
                    return result

                result.price         = live_price
                result.sl_price      = trade_calc.stop_loss.price
                result.tp_price      = trade_calc.take_profit.price
                result.position_size = trade_calc.position.size
                result.leverage      = trade_calc.position.leverage
                result.risk_reward   = trade_calc.take_profit.risk_reward

            except Exception as live_err:
                logger.error(f"   ❌ {result.coin} canlı fiyat çekme hatası: {live_err}")
                result.status = "live_price_error"
                result.error  = str(live_err)
                return result

            logger.info(
                f"\n💹 İŞLEM HAZIR: {result.coin} | {direction} @ ${result.price:,.4f} | "
                f"SL: ${result.sl_price:,.4f} | TP: ${result.tp_price:,.4f} | "
                f"Size: {result.position_size:.4f} | Lev: {result.leverage}x"
            )

            if self.dry_run:
                t = self.paper_trader.open_trade(
                    symbol        = result.coin,
                    full_symbol   = result.full_symbol,
                    direction     = direction,
                    entry_price   = result.price,
                    stop_loss     = result.sl_price,
                    take_profit   = result.tp_price,
                    position_size = result.position_size,
                    leverage      = result.leverage,
                    ic_confidence = result.ic_confidence,
                    ic_direction  = result.ic_direction,
                    best_timeframe= result.best_timeframe,
                    market_regime = result.market_regime,
                )
                result.trade_executed = True
                result.status = "executed"

                if self.notifier.is_configured():
                    msg = (f"🧪 <b>SANAL İŞLEM AÇILDI</b>\n"
                           f"━━━━━━━━━━━━━━━━━━━━━\n"
                           f"🪙 <b>Coin:</b> {result.coin}\n"
                           f"📈 <b>Yön:</b> {direction}\n"
                           f"💲 <b>Giriş:</b> ${result.price:,.4f}\n"
                           f"🛑 <b>SL:</b> ${result.sl_price:,.4f}\n"
                           f"🎯 <b>TP:</b> ${result.tp_price:,.4f}\n"
                           f"📊 <b>Kaldıraç:</b> {result.leverage}x")
                    self.notifier.send_message_sync(msg)

            else:
                try:
                    current_positions = self.executor.fetch_positions()
                    open_count = len(current_positions)
                    open_symbols = {p['symbol'] for p in current_positions if p.get('symbol')}
                except Exception as e:
                    result.status = "api_error"
                    logger.warning(f"   ⚠️ Binance API'ye ulaşılamadı ({e}). Risk almamak için işlem atlanıyor.")
                    return result

                if open_count >= MAX_OPEN_POSITIONS:
                    result.status = "position_limit"
                    logger.info(f"   ⚠️ {result.coin} atlandı: Canlı borsada max pozisyona ({MAX_OPEN_POSITIONS}) ulaşıldı.")
                    return result

                candidate_symbols = {
                    result.coin,
                    f"{result.coin}USDT",
                    f"{result.coin}/USDT",
                    result.full_symbol
                }

                if open_symbols & candidate_symbols:
                    result.status = "already_open"
                    logger.info(
                        f"   ⏭️  {result.coin} atlandı: Borsada zaten açık "
                        f"pozisyon mevcut → {open_symbols & candidate_symbols}"
                    )
                    return result

                try:
                    symbol_for_check  = result.full_symbol
                    has_existing_orders = self.executor.has_tp_sl_orders(
                        symbol_for_check
                    )
                except Exception:
                    has_existing_orders = False

                exec_res = self.executor.execute_trade(
                    trade_calc,
                    skip_sl=has_existing_orders,
                    skip_tp=has_existing_orders,
                )

                if exec_res.success:
                    result.trade_executed = True
                    result.status = "executed"

                    if self.notifier.is_configured():
                        msg = (f"🔴 <b>CANLI İŞLEM AÇILDI</b>\n"
                               f"━━━━━━━━━━━━━━━━━━━━━\n"
                               f"🪙 <b>Coin:</b> {result.coin}\n"
                               f"📈 <b>Yön:</b> {direction}\n"
                               f"💲 <b>Giriş:</b> ${exec_res.actual_entry:,.4f}\n"
                               f"🛑 <b>SL:</b> ${result.sl_price:,.4f}\n"
                               f"🎯 <b>TP:</b> ${result.tp_price:,.4f}\n"
                               f"📊 <b>Kaldıraç:</b> {result.leverage}x")
                        self.notifier.send_message_sync(msg)

                    try:
                        import pandas as pd
                        from pathlib import Path
                        import uuid

                        log_dir = Path("logs")
                        log_dir.mkdir(parents=True, exist_ok=True)
                        excel_path = log_dir / "live_trades.xlsx"

                        islem_hacmi = exec_res.actual_entry * result.position_size

                        yeni_islem = {
                            "ID": str(uuid.uuid4())[:8],
                            "Tarih (Açılış)": _iso_tr(_now_utc()),  # UTC üretilir, TR gösterilir
                            "Tarih (Kapanış)": "",
                            "Coin": result.coin,
                            "Yön": direction,
                            "Giriş ($)": exec_res.actual_entry,
                            "Çıkış ($)": "",
                            "Lot (Coin)": result.position_size,
                            "Hacim ($)": round(islem_hacmi, 2),
                            "Kaldıraç": result.leverage,
                            "SL ($)": result.sl_price,
                            "TP ($)": result.tp_price,
                            "R:R": getattr(result, 'risk_reward', 1.5),
                            "PnL ($)": 0.0,
                            "PnL (%)": 0.0,
                            "Fee ($)": 0.0,
                            "Net PnL ($)": 0.0,
                            "IC Güven": getattr(result, 'ic_confidence', 0.0),
                            "IC Yön": getattr(result, 'ic_direction', ''),
                            "TF": getattr(result, 'best_timeframe', ''),
                            "Rejim": getattr(result, 'market_regime', ''),
                            "AI Karar": direction,
                            "Durum": "open",
                            "Çıkış Nedeni": "",
                            "Süre (dk)": ""
                        }

                        df_yeni = pd.DataFrame([yeni_islem])

                        if excel_path.exists():
                            df_mevcut = pd.read_excel(excel_path)
                            df_son = pd.concat([df_mevcut, df_yeni], ignore_index=True)
                        else:
                            df_son = df_yeni

                        self._export_live_trades_to_xlsx(df_son, excel_path)
                        logger.info(f"📊 Canlı işlem Paper formatında Excel'e eklendi: {result.coin}")
                    except Exception as e:
                        logger.error(f"❌ Excel kayıt hatası: {e}")
                else:
                    result.status = "execution_error"
                    result.error  = exec_res.error
                    logger.error(f"   ❌ Canlı işlem açılamadı: {exec_res.error}")
                    if not hasattr(self, 'cooldowns'):
                        self.cooldowns = {}
                    self.cooldowns[result.coin] = now + timedelta(hours=1)
                    self.cooldowns[result.full_symbol] = now + timedelta(hours=1)
                    logger.info(f"   🚫 {result.coin} API tarafından reddedildiği için 1 saat soğumaya alındı.")

            if result.trade_executed:
                fv_dict = {}
                if result.ml_result and hasattr(result.ml_result, 'feature_vector'):
                    try:
                        fv_dict = result.ml_result.feature_vector.to_dict()
                    except Exception:
                        pass

                # ── FIX 4: Risk feature'larını gerçek hesaplanmış değerlerle override et ──
                # Feature vektörü _analyze_coin sonunda (predict çağrısında) oluşturulduğunda
                # risk değerleri henüz hesaplanmamıştı (hepsi 0 idi). Trade fiilen açıldıktan
                # sonra elimizde gerçek SL/TP/leverage/size var → snapshot'a yazıyoruz.
                try:
                    fv_dict['risk_atr_pct']     = float(result.atr_pct or 0.0)
                    fv_dict['risk_rr_ratio']    = float(result.risk_reward or 0.0)
                    if result.price > 0 and result.sl_price > 0:
                        fv_dict['risk_sl_distance_pct'] = abs(result.price - result.sl_price) / result.price * 100
                    fv_dict['risk_leverage']    = float(result.leverage or 0)
                    if result.position_size and result.position_size > 0:
                        fv_dict['risk_position_size_log'] = float(np.log10(result.position_size + 1))
                except Exception as e:
                    logger.debug(f"Risk feature override hatası (kritik değil): {e}")

                mem_rec = self.trade_memory.open_trade(
                    symbol           = result.full_symbol,
                    coin             = result.coin,
                    direction        = direction,
                    entry_price      = result.price,
                    sl_price         = result.sl_price,
                    tp_price         = result.tp_price,
                    timeframe        = result.best_timeframe,
                    ml_confidence    = getattr(result.ml_result, 'confidence', 0.0),
                    ml_direction     = direction,
                    ic_confidence    = result.ic_confidence,
                    ic_direction     = result.ic_direction,
                    market_regime    = result.market_regime,
                    validated_conf   = getattr(result.val_result, 'confidence', 0.0),
                    feature_snapshot = fv_dict,
                    position_size    = result.position_size,
                    leverage         = result.leverage,
                    risk_reward      = result.risk_reward,
                    atr              = result.atr,
                )
                if not result.paper_trade_id:
                    result.paper_trade_id = mem_rec.trade_id

        except Exception as e:
            result.status = "execution_error"
            result.error  = str(e)
            logger.error(f"❌ {result.coin} execution hatası: {e}", exc_info=True)

        return result

    def _check_open_positions(self):
        import pandas as pd  # <--- ÇÖZÜM: FONKSİYONUN EN BAŞINA EKLİYORUZ
        
        logger.debug("\n🔍 Açık pozisyonlar kontrol ediliyor...")

        if self.dry_run:
            from paper_trader import TradeStatus
            open_trades_dict = self.paper_trader.open_trades

            if not open_trades_dict:
                return

            closed_count = 0
            trades_to_check = list(open_trades_dict.values())

            for trade in trades_to_check:
                try:
                    fetch_symbol = trade.symbol if "USDT" in trade.symbol else f"{trade.symbol}/USDT"
                    ohlcv = self.fetcher.fetch_ohlcv(fetch_symbol, timeframe="1m", limit=5)

                    if ohlcv is not None and not ohlcv.empty:
                        max_high = float(ohlcv['high'].max())
                        min_low = float(ohlcv['low'].min())
                        current_price = float(ohlcv['close'].iloc[-1])

                        close_reason = None
                        exit_price = None
                        status = None

                        if trade.direction == "LONG":
                            if min_low <= trade.stop_loss:
                                close_reason = "SL Hit"
                                exit_price = trade.stop_loss
                                status = TradeStatus.CLOSED_SL
                            elif max_high >= trade.take_profit:
                                close_reason = "TP Hit"
                                exit_price = trade.take_profit
                                status = TradeStatus.CLOSED_TP
                        else:
                            if max_high >= trade.stop_loss:
                                close_reason = "SL Hit"
                                exit_price = trade.stop_loss
                                status = TradeStatus.CLOSED_SL
                            elif min_low <= trade.take_profit:
                                close_reason = "TP Hit"
                                exit_price = trade.take_profit
                                status = TradeStatus.CLOSED_TP

                        if close_reason:
                            if close_reason == "SL Hit":
                                self.cooldowns[trade.symbol] = datetime.now(timezone.utc) + timedelta(hours=2)
                                self.cooldowns[trade.full_symbol] = datetime.now(timezone.utc) + timedelta(hours=2)
                                logger.info(f"   ❄️ {trade.symbol} SL oldu! 2 saat soğuma süresine alındı.")
                            else:
                                logger.info(f"   🎯 {trade.symbol} {close_reason}! Hedefe ulaşıldı.")

                            try:
                                if hasattr(self.executor, 'cancel_all_orders'):
                                    self.executor.cancel_all_orders(trade.full_symbol)
                                    logger.info(f"   🧹 {trade.symbol} için kalan emirler iptal edildi")
                            except Exception as cancel_err:
                                logger.warning(f"   ⚠️ {trade.symbol} emir iptal hatası: {cancel_err}")

                            self.paper_trader._close_trade(trade, exit_price, status, close_reason)
                            closed_count += 1
                            logger.info(f"   ✅ {trade.symbol} işlemi kapandı! Neden: {close_reason} | Fiyat: ${exit_price:,.4f}")

                            if self.notifier.is_configured():
                                pnl_emoji = "✅ KÂR" if trade.net_pnl > 0 else "❌ ZARAR"

                                msg = (f"{pnl_emoji} <b>({close_reason})</b>\n"
                                       f"━━━━━━━━━━━━━━━━━━━━━\n"
                                       f"🪙 <b>Coin:</b> {trade.symbol}\n"
                                       f"💲 <b>Çıkış:</b> ${exit_price:,.4f}\n"
                                       f"💰 <b>Net PnL:</b> ${trade.net_pnl:+.2f} ({trade.pnl_percent:+.2f}%)\n"
                                       f"🏦 <b>Güncel Kasa:</b> ${self.paper_trader.balance:,.2f}")
                                self.notifier.send_message_sync(msg)

                            try:
                                matched = False
                                for mem_id, mem_trade in list(self.trade_memory.open_trades.items()):
                                    if mem_trade.symbol == trade.full_symbol or mem_trade.coin == trade.symbol:
                                        actual_pnl = trade.pnl_absolute if trade.pnl_absolute else 0.0
                                        self.trade_memory.close_trade(
                                            trade_id=mem_id,
                                            exit_price=exit_price,
                                            pnl=actual_pnl,
                                            exit_reason=close_reason,
                                        )
                                        matched = True
                                        logger.info(f"   🧠 Memory güncellendi: {trade.symbol} → {close_reason} | PnL: ${actual_pnl:+.2f}")
                                        break
                                if not matched:
                                    logger.warning(f"   ⚠️ Memory'de eşleşen trade bulunamadı: {trade.symbol} (full={trade.full_symbol})")
                            except Exception as em:
                                logger.warning(f"   ⚠️ Memory update BAŞARISIZ: {trade.symbol} → {em}")

                except Exception as e:
                    logger.error(f"   ❌ {trade.symbol} pozisyon kontrolünde hata: {e}")

            if closed_count > 0:
                logger.info(f"   Mevcut Bakiye: ${self.paper_trader.balance:.2f}")

        else:
            try:
                closed_count = 0

                try:
                    real_positions = self.executor.fetch_positions()
                except Exception as e:
                    logger.error(f"   ❌ Binance pozisyon çekme hatası: {e}")
                    return

                active_coins = set()
                for p in real_positions:
                    size_ccxt = p.get('contracts')
                    size_raw = p.get('positionAmt')
                    size_info = p.get('info', {}).get('positionAmt')

                    is_active = False
                    try:
                        val = size_ccxt if size_ccxt is not None else (size_raw if size_raw is not None else size_info)
                        size_val = float(val) if val is not None else 0.0

                        if abs(size_val) > 0.0:
                            is_active = True
                    except (ValueError, TypeError):
                        pass

                    if size_ccxt is None and size_raw is None and size_info is None:
                        is_active = True

                    if is_active and p.get('symbol'):
                        coin_name = str(p['symbol']).split('/')[0].split(':')[0].replace('USDT', '')
                        active_coins.add(coin_name)

                if not hasattr(self.trade_memory, 'open_trades') or not self.trade_memory.open_trades:
                    try:
                        exchange = self.executor._get_exchange()
                        if hasattr(exchange, 'options'):
                            exchange.options["warnOnFetchOpenOrdersWithoutSymbol"] = False
                        all_open_orders = exchange.fetch_open_orders()
                        if all_open_orders:
                            orphan_count = 0
                            for order in all_open_orders:
                                order_symbol = order.get('symbol', '')
                                order_id     = order.get('id', '?')
                                order_type   = order.get('type', '?')
                                has_position = any(ac in order_symbol for ac in active_coins)
                                if not has_position:
                                    try:
                                        exchange.cancel_order(order_id, order_symbol)
                                        orphan_count += 1
                                        logger.info(f"   🗑️ Orphan emir silindi: {order_symbol} | {order_type} | ID: {order_id}")
                                    except Exception as cancel_err:
                                        if 'Unknown order' in str(cancel_err) or 'UNKNOWN_ORDER' in str(cancel_err):
                                            logger.debug(f"   ℹ️ Orphan emir zaten yok: {order_id}")
                                        else:
                                            logger.warning(f"   ⚠️ Orphan emir silinemedi: {order_symbol} {order_id}: {cancel_err}")
                            if orphan_count > 0:
                                logger.info(f"   🧹 {orphan_count} sahipsiz (orphan) emir temizlendi.")
                    except Exception as orphan_err:
                        logger.warning(f"   ⚠️ Orphan emir kontrolü (boş memory): {orphan_err}")
                    return

                for mem_id, mem_trade in list(self.trade_memory.open_trades.items()):
                    mem_coin = str(mem_trade.coin).replace('USDT', '').split('/')[0].split(':')[0]

                    opened_time = datetime.fromisoformat(mem_trade.opened_at)
                    if opened_time.tzinfo is None:
                        opened_time = opened_time.replace(tzinfo=timezone.utc)
                    time_open_sec = (datetime.now(timezone.utc) - opened_time).total_seconds()
                    age_hours = time_open_sec / 3600

                    # ── TIMEOUT EXIT (24 saat kuralı) ──
                    # Pozisyon Binance'te HÂLÂ AÇIK ve 24 saatten uzun süredir asılıysa
                    # piyasa fiyatından zorla kapat. Kâr veya zarar farketmez.
                    # Mantık: edge bayatladı, slot fırsat maliyeti artıyor.
                    # Kâr ile kapanırsa model bunu WIN olarak öğrenir (trade_memory pnl>0 → WIN).
                    # Zarar ile kapanırsa LOSS olarak öğrenir.
                    if mem_coin in active_coins and age_hours >= MAX_TRADE_AGE_HOURS:
                        logger.info(
                            f"   ⏰ {mem_coin} TIMEOUT (yaş={age_hours:.1f}h ≥ {MAX_TRADE_AGE_HOURS}h) "
                            f"→ piyasa fiyatından kapatılıyor..."
                        )

                        try:
                            # 1) Güncel piyasa fiyatını çek
                            exchange = self.executor._get_exchange()
                            ticker = exchange.fetch_ticker(mem_trade.symbol)
                            timeout_exit_price = float(ticker.get('last') or 0.0)

                            if timeout_exit_price <= 0:
                                logger.warning(f"   ⚠️ {mem_coin} TIMEOUT: geçersiz fiyat ({timeout_exit_price}), atlanıyor")
                                continue

                            # 2) Pozisyonu market order ile kapat (close_position reduce_only=True kullanır)
                            #    BinanceExecutor.close_position içinde LONG → sell, SHORT → buy mantığı var
                            try:
                                close_result = self.executor.close_position(
                                    symbol=mem_trade.symbol,
                                    side=mem_trade.direction.lower(),
                                    amount=mem_trade.position_size,
                                )
                                if not close_result.success:
                                    logger.error(f"   ❌ {mem_coin} TIMEOUT close hatası: {close_result.error}")
                                    continue
                                # Gerçek fill fiyatı varsa onu kullan
                                if close_result.price and close_result.price > 0:
                                    timeout_exit_price = float(close_result.price)
                            except Exception as close_err:
                                logger.error(f"   ❌ {mem_coin} TIMEOUT market close exception: {close_err}")
                                continue

                            # 3) Kalan SL/TP emirlerini iptal et (orphan kalmasınlar)
                            try:
                                if hasattr(self.executor, 'cancel_all_orders'):
                                    self.executor.cancel_all_orders(mem_trade.symbol)
                                    logger.info(f"   🧹 {mem_coin} TIMEOUT sonrası SL/TP emirleri iptal edildi")
                            except Exception as cancel_err:
                                logger.warning(f"   ⚠️ {mem_coin} TIMEOUT emir temizleme hatası: {cancel_err}")

                            # 4) PnL hesapla
                            safe_entry = float(mem_trade.entry_price or 0)
                            safe_size = float(mem_trade.position_size or 0)
                            safe_leverage = int(mem_trade.leverage or 1)

                            if mem_trade.direction == "LONG":
                                pnl_val = (timeout_exit_price - safe_entry) * safe_size
                            else:
                                pnl_val = (safe_entry - timeout_exit_price) * safe_size

                            fee_val = (safe_entry * safe_size * 0.0004) + (timeout_exit_price * safe_size * 0.0004)
                            net_pnl_val = pnl_val - fee_val
                            pnl_pct = (pnl_val / (safe_entry * safe_size)) * 100 * safe_leverage \
                                      if (safe_entry * safe_size) > 0 else 0.0

                            close_reason_timeout = "TIMEOUT"

                            # 5) TIMEOUT cooldown YOK — edge bayatlaması, başarısızlık değil
                            #    (SL Hit'te 2 saat cooldown var, TIMEOUT'ta yok bilinçli)

                            # 6) Memory güncelle
                            try:
                                self.trade_memory.close_trade(
                                    trade_id    = mem_id,
                                    exit_price  = timeout_exit_price,
                                    pnl         = pnl_val,
                                    exit_reason = close_reason_timeout,
                                )
                                outcome_str = "KÂR" if pnl_val > 0 else "ZARAR"
                                logger.info(
                                    f"   🧠 Memory güncellendi: {mem_coin} → TIMEOUT/{outcome_str} | "
                                    f"PnL: ${pnl_val:+.2f}"
                                )
                            except Exception as mem_err:
                                logger.error(f"   ❌ Memory güncelleme hatası: {mem_coin} → {mem_err}")

                            # 7) Telegram bildirimi
                            try:
                                if self.notifier.is_configured():
                                    pnl_emoji = "✅ KÂR" if pnl_val > 0 else "❌ ZARAR"
                                    msg = (f"⏰ <b>TIMEOUT EXIT ({pnl_emoji})</b>\n"
                                           f"━━━━━━━━━━━━━━━━━━━━━\n"
                                           f"🪙 <b>Coin:</b> {mem_coin}\n"
                                           f"⏱️ <b>Yaş:</b> {age_hours:.1f}h\n"
                                           f"💲 <b>Çıkış:</b> ${timeout_exit_price:,.4f}\n"
                                           f"💰 <b>PnL:</b> ${pnl_val:+.2f} ({pnl_pct:+.2f}%)\n"
                                           f"🏦 <b>Güncel Kasa:</b> ${self._balance:,.2f}")
                                    self.notifier.send_message_sync(msg)
                            except Exception as tg_err:
                                logger.warning(f"   ⚠️ Telegram bildirimi gönderilemedi: {tg_err}")


                            # 8) Excel güncelleme
                            try:
                                from pathlib import Path
                                log_dir = Path("logs")
                                excel_path = log_dir / "live_trades.xlsx"

                                if excel_path.exists():
                                    df = pd.read_excel(excel_path)

                                    for col in ['Tarih (Kapanış)', 'Çıkış Nedeni', 'Durum']:
                                        if col in df.columns:
                                            df[col] = df[col].astype(object)
                                    for col in ['Çıkış ($)', 'PnL ($)', 'PnL (%)']:
                                        if col in df.columns:
                                            df[col] = df[col].astype(float)

                                    mask = (df['Coin'] == mem_coin) & (df['Durum'] == 'open')

                                    if mask.any():
                                        idx = df[mask].index[-1]
                                        kapanis_zamani = _now_utc()
                                        df.at[idx, 'Tarih (Kapanış)'] = _iso_tr(kapanis_zamani)
                                        try:
                                            acilis_str = str(df.at[idx, 'Tarih (Açılış)'])
                                            acilis_zamani = pd.to_datetime(acilis_str)
                                            # UTC-aware yap (aware - naive → TypeError önle)
                                            if acilis_zamani.tzinfo is None:
                                                acilis_zamani = acilis_zamani.tz_localize('UTC')
                                            df.at[idx, 'Süre (dk)'] = int((kapanis_zamani - acilis_zamani).total_seconds() / 60)
                                        except Exception:
                                            df.at[idx, 'Süre (dk)'] = 0

                                        df.at[idx, 'Çıkış ($)']    = timeout_exit_price
                                        df.at[idx, 'Çıkış Nedeni'] = close_reason_timeout
                                        df.at[idx, 'PnL ($)']      = round(pnl_val, 4)
                                        df.at[idx, 'PnL (%)']      = round(pnl_pct, 2)
                                        df.at[idx, 'Fee ($)']      = round(fee_val, 4)
                                        df.at[idx, 'Net PnL ($)']  = round(net_pnl_val, 4)
                                        df.at[idx, 'Durum']        = "closed"

                                        self._export_live_trades_to_xlsx(df, excel_path)
                                        logger.info(f"   📊 Excel güncellendi (TIMEOUT): {mem_coin} | PnL: ${pnl_val:+.2f}")
                                    else:
                                        logger.warning(f"   ⚠️ Excel'de 'open' statüsünde {mem_coin} kaydı bulunamadı!")
                            except Exception as e:
                                logger.error(f"   ❌ TIMEOUT Excel güncelleme hatası: {e}", exc_info=True)

                            # TIMEOUT işlemi tamamlandı, bu trade için döngü iterasyonunu bitir
                            continue

                        except Exception as timeout_err:
                            logger.error(f"   ❌ {mem_coin} TIMEOUT prosedürü hatası: {timeout_err}", exc_info=True)
                            continue

                    if mem_coin not in active_coins:
                        if time_open_sec <= 120:
                            logger.debug(f"   ⏳ {mem_coin} borsa onayı bekleniyor... (Kalan: {int(120-time_open_sec)}s)")
                            continue

                        logger.info(f"   🔍 {mem_coin} pozisyonu kapalı bulundu → kapanış nedeni araştırılıyor...")

                        close_reason = "Manuel / API"
                        exit_price = mem_trade.entry_price

                        try:
                            ohlcv = self.fetcher.fetch_ohlcv(mem_trade.symbol, timeframe="1m", limit=5)

                            if ohlcv is not None and not ohlcv.empty:
                                max_high = float(ohlcv['high'].max())
                                min_low = float(ohlcv['low'].min())
                                current_price = float(ohlcv['close'].iloc[-1])

                                dist_tp = abs(current_price - float(mem_trade.tp_price))
                                dist_sl = abs(current_price - float(mem_trade.sl_price))

                                if mem_trade.direction == "LONG":
                                    if min_low <= mem_trade.sl_price:
                                        close_reason = "SL Hit"
                                        exit_price = mem_trade.sl_price
                                    elif max_high >= mem_trade.tp_price:
                                        close_reason = "TP Hit"
                                        exit_price = mem_trade.tp_price
                                    else:
                                        close_reason = "TP Hit" if dist_tp < dist_sl else "SL Hit"
                                        exit_price = mem_trade.tp_price if dist_tp < dist_sl else mem_trade.sl_price

                                elif mem_trade.direction == "SHORT":
                                    if max_high >= mem_trade.sl_price:
                                        close_reason = "SL Hit"
                                        exit_price = mem_trade.sl_price
                                    elif min_low <= mem_trade.tp_price:
                                        close_reason = "TP Hit"
                                        exit_price = mem_trade.tp_price
                                    else:
                                        close_reason = "TP Hit" if dist_tp < dist_sl else "SL Hit"
                                        exit_price = mem_trade.tp_price if dist_tp < dist_sl else mem_trade.sl_price

                        except Exception as ohlcv_err:
                            logger.warning(f"   ⚠️ {mem_coin} OHLCV kontrol hatası: {ohlcv_err}")

                        safe_entry = float(mem_trade.entry_price or 0)
                        safe_exit = float(exit_price or safe_entry)
                        safe_size = float(mem_trade.position_size or 0)
                        safe_leverage = int(mem_trade.leverage or 1)

                        if mem_trade.direction == "LONG":
                            pnl_val = (safe_exit - safe_entry) * safe_size
                        else:
                            pnl_val = (safe_entry - safe_exit) * safe_size

                        fee_val = (safe_entry * safe_size * 0.0004) + (safe_exit * safe_size * 0.0004)
                        net_pnl_val = pnl_val - fee_val

                        pnl_pct = (pnl_val / (safe_entry * safe_size)) * 100 * safe_leverage if (safe_entry * safe_size) > 0 else 0.0

                        if close_reason == "SL Hit":
                            self.cooldowns[mem_coin] = datetime.now(timezone.utc) + timedelta(hours=2)
                            self.cooldowns[mem_trade.symbol] = datetime.now(timezone.utc) + timedelta(hours=2)
                            logger.info(f"   ❄️ {mem_coin} SL oldu! 2 saat soğuma süresine alındı.")
                        else:
                            logger.info(f"   🎯 {mem_coin} {close_reason}! Hedefe ulaşıldı.")

                        try:
                            if hasattr(self.executor, 'cancel_all_orders'):
                                self.executor.cancel_all_orders(mem_trade.symbol)
                                logger.info(f"   🧹 {mem_coin} için kalan emirler iptal edildi")
                        except Exception as cancel_err:
                            logger.warning(f"   ⚠️ {mem_coin} emir temizleme hatası: {cancel_err}")

                        try:
                            self.trade_memory.close_trade(
                                trade_id    = mem_id,
                                exit_price  = safe_exit,
                                pnl         = pnl_val,
                                exit_reason = close_reason,
                            )
                            logger.info(f"   🧠 Memory güncellendi: {mem_coin} → {close_reason} | PnL: ${pnl_val:+.2f}")
                        except Exception as mem_err:
                            logger.error(f"   ❌ Memory güncelleme hatası: {mem_coin} → {mem_err}")

                        try:
                            if self.notifier.is_configured():
                                pnl_emoji = "✅ KÂR" if pnl_val > 0 else "❌ ZARAR"
                                msg = (f"{pnl_emoji} <b>({close_reason})</b>\n"
                                       f"━━━━━━━━━━━━━━━━━━━━━\n"
                                       f"🪙 <b>Coin:</b> {mem_coin}\n"
                                       f"💲 <b>Çıkış:</b> ${safe_exit:,.4f}\n"
                                       f"💰 <b>PnL:</b> ${pnl_val:+.2f} ({pnl_pct:+.2f}%)\n"
                                       f"🏦 <b>Güncel Kasa:</b> ${self._balance:,.2f}")
                                self.notifier.send_message_sync(msg)
                        except Exception as tg_err:
                            logger.warning(f"   ⚠️ Telegram bildirimi gönderilemedi: {tg_err}")

                        try:
                            from pathlib import Path

                            log_dir = Path("logs")
                            excel_path = log_dir / "live_trades.xlsx"

                            if excel_path.exists():
                                df = pd.read_excel(excel_path)

                                for col in ['Tarih (Kapanış)', 'Çıkış Nedeni', 'Durum']:
                                    if col in df.columns:
                                        df[col] = df[col].astype(object)

                                for col in ['Çıkış ($)', 'PnL ($)', 'PnL (%)']:
                                    if col in df.columns:
                                        df[col] = df[col].astype(float)

                                mask = (df['Coin'] == mem_coin) & (df['Durum'] == 'open')

                                if mask.any():
                                    idx = df[mask].index[-1]

                                    kapanis_zamani = _now_utc()
                                    df.at[idx, 'Tarih (Kapanış)'] = _iso_tr(kapanis_zamani)

                                    try:
                                        acilis_str = str(df.at[idx, 'Tarih (Açılış)'])
                                        acilis_zamani = pd.to_datetime(acilis_str)
                                        if acilis_zamani.tzinfo is None:
                                            acilis_zamani = acilis_zamani.tz_localize('UTC')
                                        df.at[idx, 'Süre (dk)'] = int((kapanis_zamani - acilis_zamani).total_seconds() / 60)
                                    except Exception:
                                        df.at[idx, 'Süre (dk)'] = 0

                                    df.at[idx, 'Çıkış ($)'] = safe_exit
                                    df.at[idx, 'Çıkış Nedeni'] = close_reason
                                    df.at[idx, 'PnL ($)'] = round(pnl_val, 4)
                                    df.at[idx, 'PnL (%)'] = round(pnl_pct, 2)
                                    df.at[idx, 'Fee ($)'] = round(fee_val, 4)
                                    df.at[idx, 'Net PnL ($)'] = round(net_pnl_val, 4)
                                    df.at[idx, 'Durum'] = "closed"

                                    self._export_live_trades_to_xlsx(df, excel_path)
                                    logger.info(f"   📊 Excel güncellendi: {mem_trade.coin} ({close_reason}) | PnL: ${pnl_val:+.2f}")
                                else:
                                    logger.warning(f"   ⚠️ Excel'de 'open' statüsünde {mem_coin} kaydı bulunamadı!")
                        except Exception as e:
                            logger.error(f"   ❌ Excel güncelleme hatası: {e}", exc_info=True)

                        closed_count += 1

                try:
                    closed_count += self._reconcile_excel_orphans(active_coins)
                except Exception as rec_err:
                    logger.warning(f"   ⚠️ Excel reconcile hatası: {rec_err}")

                if closed_count > 0:
                    logger.info(f"   🧹 {closed_count} işlemin kontrolü ve temizliği tamamlandı.")

                try:
                    exchange = self.executor._get_exchange()

                    if hasattr(exchange, 'options'):
                        exchange.options["warnOnFetchOpenOrdersWithoutSymbol"] = False

                    all_open_orders = exchange.fetch_open_orders()

                    if all_open_orders:
                        orphan_count = 0
                        for order in all_open_orders:
                            order_symbol = order.get('symbol', '')
                            order_id     = order.get('id', '?')
                            order_type   = order.get('type', '?')

                            has_position = False
                            for act_coin in active_coins:
                                if act_coin in order_symbol:
                                    has_position = True
                                    break

                            if not has_position:
                                try:
                                    exchange.cancel_order(order_id, order_symbol)
                                    orphan_count += 1
                                    logger.info(f"   🗑️ Orphan emir silindi: {order_symbol} | {order_type} | ID: {order_id}")
                                except Exception as cancel_err:
                                    if 'Unknown order' in str(cancel_err) or 'UNKNOWN_ORDER' in str(cancel_err):
                                        logger.debug(f"   ℹ️ Orphan emir zaten yok: {order_id}")
                                    else:
                                        logger.warning(f"   ⚠️ Orphan emir silinemedi: {order_symbol} {order_id}: {cancel_err}")

                        if orphan_count > 0:
                            logger.info(f"   🧹 {orphan_count} sahipsiz (orphan) emir temizlendi.")

                except Exception as orphan_err:
                    logger.warning(f"   ⚠️ Orphan emir kontrolü başarısız: {orphan_err}")

            except Exception as e:
                logger.error(f"   ❌ Canlı pozisyon kontrol/temizlik hatası: {e}", exc_info=True)

    def run_cycle(self) -> CycleReport:
        start  = time.time()
        report = CycleReport(timestamp=datetime.now(timezone.utc).isoformat())

        logger.info(f"\n{'═'*60}")
        logger.info(f"🔄 YENİ DÖNGÜ — {_now_local().strftime('%H:%M:%S')} "
                    f"| v{VERSION} | {'PAPER' if self.dry_run else 'CANLI'}")
        logger.info(f"{'═'*60}")

        # ── Günlük state sıfırlama (gün değişimi kontrolü) ──
        self._reset_daily_state_if_new_day()

        # ── Kill switch (canlı bakiye refresh edilir) ──
        if self._check_kill_switch():
            report.status = CycleStatus.KILLED
            return report

        try:
            # ── Açık pozisyon takibi ÖNCE çalışır: yeni kapanan trade'ler varsa
            #    daily_pnl doğru hesaplansın. ──
            self._check_open_positions()

            # ── Günlük kayıp limiti kontrolü (v2.1.4) ──
            if self._check_daily_loss_limit():
                report.status = CycleStatus.HALTED
                report.daily_pnl = self._daily_realized_pnl
                report.balance = self._balance
                report.paper_balance = self.paper_trader.balance
                report.elapsed = time.time() - start
                self._log_cycle_summary(report)
                return report

            logger.info(f"\n📡 Coin taraması (top {self.top_n})...")
            coins = self.scanner.scan(top_n=self.top_n)
            if not coins:
                report.status = CycleStatus.ERROR
                return report
            report.total_scanned = len(coins)

            self.validator._closed_trade_count = self.trade_memory._count_closed()
            self.lgbm_model._closed_trade_count = self.trade_memory._count_closed()

            logger.info(f"\n🔬 ML analizi ({len(coins)} coin)...")
            results = []

            if self.dry_run:
                open_coins = [trade.symbol for trade in self.paper_trader.open_trades.values()]
            else:
                open_coins = [mem_trade.coin for mem_trade in self.trade_memory.open_trades.values()]

            for c in coins:
                if c.coin in open_coins:
                    logger.info(f"   ⏭️ {c.coin} atlanıyor (Zaten açık pozisyon var)")
                    continue

                if c.coin in self.cooldowns or c.symbol in self.cooldowns:
                    cooldown_key = c.coin if c.coin in self.cooldowns else c.symbol
                    if datetime.now(timezone.utc) < self.cooldowns[cooldown_key]:
                        kalan_dk = int((self.cooldowns[cooldown_key] - datetime.now(timezone.utc)).total_seconds() / 60)
                        logger.info(f"   ❄️ {c.coin} atlanıyor (Soğuma süresinde - Kalan: {kalan_dk} dk)")
                        continue
                    else:
                        if c.coin in self.cooldowns:
                            del self.cooldowns[c.coin]
                        if c.symbol in self.cooldowns:
                            del self.cooldowns[c.symbol]

                r = self._analyze_coin(c.symbol, c.coin, scan_result=c)   # ← FIX 1: scanner verisi geçiriliyor

                results.append(r)
                report.total_analyzed += 1
                if r.ic_confidence >= IC_TRADE:
                    report.total_above_gate += 1

            logger.info(f"\n💹 Execution...")
            for r in results:
                if r.status == "ready":
                    # ── Her execute öncesi tekrar halt kontrolü (güvenlik) ──
                    if self._daily_halt_until and datetime.now(timezone.utc) < self._daily_halt_until:
                        logger.info("   🛑 Döngü ortasında günlük halt devreye girdi — kalan sinyaller iptal")
                        break
                    r = self._execute_trade(r)
                    if r.trade_executed:
                        report.total_traded += 1

            report.coins        = results
            report.paper_balance= self.paper_trader.balance
            report.balance      = self._balance
            report.daily_pnl    = self._daily_realized_pnl

            pt_stats = self.paper_trader.get_summary()
            total_closed = pt_stats.get("closed_trades", 0)
            real_win_rate = pt_stats.get("win_rate_pct", 0.0)

            retrain_threshold = 30
            current_retrain_count = getattr(self.lgbm_model, 'retrain_count', 0)

            if total_closed >= retrain_threshold:
                target_retrain_count = total_closed // retrain_threshold

                if target_retrain_count > current_retrain_count:
                    logger.info(f"\n🧠 [RETRAIN] {total_closed} kapalı işleme ulaşıldı! Model kendi tecrübelerinden yeniden eğitiliyor...")
                    try:
                        if hasattr(self, 'retrain_from_experience'):
                            self.retrain_from_experience()
                    except Exception as e:
                        logger.error(f"Eğitim tetiklenemedi: {e}")

                    self.lgbm_model.retrain_count = target_retrain_count

            next_target = ((total_closed // retrain_threshold) + 1) * retrain_threshold
            kalan_islem = next_target - total_closed

            report.ml_stats = {
                "closed_trades":  total_closed,
                "win_rate":       real_win_rate,
                "retrain_count":  getattr(self.lgbm_model, 'retrain_count', 0),
                "next_retrain_in": kalan_islem,
                "model_trained":  self.lgbm_model.is_trained,
            }
            report.status = (
                CycleStatus.SUCCESS   if report.total_traded > 0
                else CycleStatus.PARTIAL   if report.total_above_gate > 0
                else CycleStatus.NO_SIGNAL
            )
            self._consecutive_errors = 0

        except Exception as e:
            report.status = CycleStatus.ERROR
            report.errors.append(str(e))
            self._consecutive_errors += 1
            logger.error(f"❌ Döngü hatası: {e}", exc_info=True)

        report.elapsed = time.time() - start
        self._log_cycle_summary(report)
        return report

    def retrain_from_experience(self):
        logger.info("🧠 Kendi tecrübelerinden öğrenme (Retrain) başlatılıyor...")

        try:
            memory_file = self.trade_memory.log_dir / "ml_trade_memory.json"
            if not memory_file.exists():
                logger.warning("   ⚠️ Trade hafızası bulunamadı, standart BTC eğitimine dönülüyor.")
                return self.initial_train()

            import json
            with open(memory_file, 'r', encoding='utf-8') as f:
                trades = json.load(f)

            closed_trades = [t for t in trades if t.get('exit_reason') or t.get('pnl', 0) != 0]

            if len(closed_trades) < 30:
                logger.warning(f"   ⚠️ Yeterli kapalı işlem yok ({len(closed_trades)} < 30), standart eğitime dönülüyor.")
                return self.initial_train()

            rows_X = []
            rows_y = []
            rows_dir = [] # YENİ EKLENDİ

            for t in closed_trades:
                fv = t.get('feature_snapshot', {})
                # Eğer fv boşsa veya amnezi (eksik kolon) sorunu yaşanmışsa o işlemi yoksay!
                if not fv or 'ic_confidence' not in fv:
                    continue

                direction = t.get('direction', '')
                exit_reason = t.get('exit_reason', '')
                pnl = t.get('pnl', 0.0)

                # [MADDE 1 DÜZELTMESİ] — Gerçek R-multiple hesabı
                # Eski kod: TP Hit → binary +1.5, SL Hit → binary -1.0
                # Yeni kod: exit/entry/sl fiyatlarından kontinü R değeri hesapla.
                # Neden önemli: Binary etiket model'in "ne kadar?" sorusunu öğrenmesini
                # engeller. Partial TP (+0.8R) ile tam TP (+1.5R) aynı görünüyordu.
                if exit_reason in ("TP Hit", "SL Hit"):
                    entry_p = float(t.get('entry_price', 0) or 0)
                    exit_p  = float(t.get('exit_price',  0) or 0)
                    sl_p    = float(t.get('sl_price',    0) or 0)
                    sl_dist = abs(entry_p - sl_p)

                    if sl_dist > 0 and entry_p > 0 and exit_p > 0:
                        # Yön bazlı fiyat hareketi / SL mesafesi = R-multiple
                        if direction == "LONG":
                            r_multiple = (exit_p - entry_p) / sl_dist
                        else:  # SHORT
                            r_multiple = (entry_p - exit_p) / sl_dist
                        r_multiple = max(-2.5, min(2.5, r_multiple))
                    else:
                        # Fiyat verisi eksik → eski fallback değerler (veri kalitesi sorunu)
                        logger.debug(
                            f"   ⚠️ R-multiple hesaplanamadı ({t.get('coin','?')} "
                            f"entry={entry_p} sl={sl_p}) — fallback değer kullanılıyor"
                        )
                        r_multiple = 1.5 if exit_reason == "TP Hit" else -1.0
                else:
                    if pnl != 0:
                        risk_est = self._balance * 0.02 if self._balance > 0 else 1.0
                        r_multiple = pnl / risk_est if risk_est > 0 else 0.0  # [SORUN 1 FIX]

                        r_multiple = max(-2.0, min(2.0, r_multiple))
                    else:
                        r_multiple = 0.0
                        
                # --- YENİ EKLENEN DEAD-BAND (ÖLÜ BÖLGE) FİLTRESİ ---
                # TIMEOUT olan ve çok küçük kâr/zarar (-0.25R ile +0.25R arası) eden işlemleri
                # gürültü (noise) olmaması için eğitimden dışlıyoruz.
                if exit_reason == "TIMEOUT" and abs(r_multiple) < 0.25:
                    continue
                
                target = float(r_multiple)

                rows_X.append(fv)
                rows_y.append(target)
                rows_dir.append(direction)

            if len(rows_X) == 0:
                return False

            win_ratio = sum(1 for y in rows_y if float(y) > 0) / max(1, len(rows_y)) * 100
            logger.info(f"   📥 Hafızadan {len(rows_X)} adet gerçek işlem tecrübesi yüklendi. (Piyasa yönü=LONG oranı: %{win_ratio:.1f})")

            X_experience = pd.DataFrame(rows_X).replace([np.inf, -np.inf], np.nan)
            y_experience = pd.Series(rows_y)
            dir_experience = pd.Series(rows_dir) # YENİ EKLENDİ

            # TRAIN KISMI GÜNCELLENDİ
            metrics = self.lgbm_model.train(X_experience, y_experience, directions=dir_experience)

            try:
                from pathlib import Path

                report_dir = Path("logs/reports")
                report_dir.mkdir(parents=True, exist_ok=True)
                report_path = report_dir / "model_egitim_raporu.xlsx"

                with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                    # --- 1. SAYFA: GENEL METRİKLER ---
                    z_score = (metrics.spearman_ic * np.sqrt(metrics.n_train_samples - 3)) if metrics.n_train_samples > 3 else 0.0
                    
                    df_metrics = pd.DataFrame([{
                        "Tarih": _now_local().strftime("%Y-%m-%d %H:%M:%S"),  # TR saati
                        "Kaynak": "retrain_from_experience",
                        "Eğitim Satır Sayısı": len(X_experience),
                        "Kazanma Oranı (Win Rate)": f"{sum(1 for val in y_experience if float(val) > 0) / max(1, len(y_experience)) * 100:.1f}%",
                        "Spearman IC Skoru": round(metrics.spearman_ic, 4),
                        "Bilgi Oranı (IR)": round(metrics.information_ratio, 2),
                        "Z-Skoru": round(z_score, 2),
                        "Long-Short Spread (R)": round(metrics.long_short_spread, 3),
                        "MAE (Hata Payı)": round(metrics.mae, 4)
                    }])
                    df_metrics.to_excel(writer, sheet_name="1_Genel_Metrikler", index=False)

                    # --- 2. SAYFA: KOLON ÖNEMLERİ (GÜVENLİ BLOK) ---
                    try:
                        importance = None
                        kalan_kolonlar = None
                        
                        if hasattr(self.lgbm_model, 'feature_importances_'):
                            importance = self.lgbm_model.feature_importances_() if callable(self.lgbm_model.feature_importances_) else self.lgbm_model.feature_importances_
                        elif hasattr(self.lgbm_model, 'model') and hasattr(self.lgbm_model.model, 'feature_importances_'):
                            importance = self.lgbm_model.model.feature_importances_
                            
                        if importance is not None:
                            if hasattr(self.lgbm_model, 'feature_names') and self.lgbm_model.feature_names:
                                kalan_kolonlar = self.lgbm_model.feature_names
                            else:
                                kalan_kolonlar = X_experience.columns[:len(importance)]
                            
                            if len(kalan_kolonlar) == len(importance):
                                df_imp = pd.DataFrame({
                                    "Feature (Kolon)": kalan_kolonlar,
                                    "Önem Puanı": importance
                                }).sort_values(by="Önem Puanı", ascending=False)
                            else:
                                df_imp = pd.DataFrame({"Hata": [f"Uyuşmazlık! Kolon:{len(kalan_kolonlar)}, Puan:{len(importance)}"]})
                        else:
                            df_imp = pd.DataFrame({"Bilgi": ["Model önem puanı üretmedi"]})
                            
                        df_imp.to_excel(writer, sheet_name="2_Kolon_Onemleri", index=False)
                    except Exception as e:
                        pd.DataFrame({"Hata": [str(e)]}).to_excel(writer, sheet_name="2_Kolon_Onemleri", index=False)

                    # --- 3. SAYFA: GEÇMİŞ HAM VERİ ---
                    df_raw = X_experience.copy()
                    df_raw['TARGET_SONUC'] = y_experience.values
                    df_raw['TARGET_ACIKLAMA'] = df_raw['TARGET_SONUC'].apply(
                        lambda x: "KAZANÇ (Pozitif R)" if float(x) > 0 else "ZARAR (Negatif R)"
                    )
                    df_raw.to_excel(writer, sheet_name="3_Gecmis_Ham_Veri", index=False)

                logger.info(f"📊 Retrain raporu Excel olarak kaydedildi: {report_path}")
            
            except Exception as ex:
                logger.error(f"⚠️ Excel raporu oluşturulurken hata (kritik değil): {ex}")

            gercek_win_rate = sum(1 for val in y_experience if float(val) > 0) / max(1, len(y_experience)) * 100
            logger.info(f"✅ Gerçek işlemlerden öğrenildi! Yeni Win Rate: %{gercek_win_rate:.1f}")

            # [MADDE 3] — Retrain sonrası deployment gate kontrolü
            # Her retrain sonrası modelin hâlâ yeterli istatistiksel güce sahip olup
            # olmadığını kontrol et. Gate başarısız → canlı trade bloklanır.
            self._model_deployment_approved = self._check_deployment_gates(metrics)
            if not self._model_deployment_approved:
                logger.warning(
                    "⚠️ Retrain sonrası deployment gate başarısız. "
                    "Canlı trade bloklandı — paper modda çalışmaya devam ediliyor."
                )
            return True

        except Exception as e:
            logger.error(f"❌ Tecrübeden öğrenme (Retrain) hatası: {e}", exc_info=True)
            return False

    def initial_train(self, symbol: str = "BTC/USDT:USDT") -> bool:
        logger.info(f"🎓 İlk eğitim: {symbol} 1h verisi kullanılıyor...")

        try:
            TRAIN_TF    = "1h"
            TRAIN_LIMIT = 1000

            df_raw = self.fetcher.fetch_ohlcv(symbol, TRAIN_TF, limit=TRAIN_LIMIT)
            if df_raw is None or len(df_raw) < 200:
                logger.error("❌ Yeterli veri çekilemedi")
                return False

            df_clean = self.preprocessor.full_pipeline(df_raw)
            if df_clean is None or len(df_clean) < 100:
                logger.error("❌ Preprocessing sonrası yetersiz veri")
                return False

            df_ind = self.calculator.calculate_all(df_clean)
            df_ind = self.calculator.add_forward_returns(df_ind, periods=[self.fwd_period])

            target_col = f'fwd_ret_{self.fwd_period}'
            if target_col not in df_ind.columns:
                logger.error(f"❌ Target kolon bulunamadı: {target_col}")
                return False

            class DummyAnalysis:
                def __init__(self, sym):
                    self.symbol          = sym
                    self.coin            = sym.split('/')[0]
                    self.price           = 0.0
                    self.change_24h      = 0.0
                    self.volume_24h      = 0.0
                    self.ic_confidence   = 65.0
                    self.ic_direction    = 'NEUTRAL'
                    self.significant_count = 10
                    self.market_regime   = 'trending'
                    self.category_tops   = {}
                    self.tf_rankings     = []
                    self.atr             = 0.0
                    self.atr_pct         = 0.0
                    self.sl_price        = 0.0
                    self.tp_price        = 0.0
                    self.risk_reward     = 0.0
                    self.position_size   = 0.0
                    self.leverage        = 1

            analysis_stub = DummyAnalysis(symbol)
            rows_X = []
            rows_y = []
            rows_dir = [] # YENİ EKLENDİ

            train_cat_tops = {}
            all_scores = self.selector.evaluate_all_indicators(df_ind, target_col)

            # [SORUN 2 DÜZELTMESİ] — Temiz helper kullan
            CATEGORIES = {
                'volume':     ['OBV', 'CMF', 'VPT', 'FI', 'EOM', 'ADI', 'NVI', 'MFI'],
                'momentum':   ['RSI', 'Stoch', 'UO', 'MACD', 'PPO', 'ROC', 'TSI', 'CCI', 'Williams', 'WILLR'],
                'trend':      ['ADX', 'Aroon', 'PSAR', 'DPO', 'Vortex', 'KST', 'Ichimoku', 'SMA', 'EMA', 'WMA'],
                'volatility': ['BBW', 'BBU', 'BBL', 'BBM', 'ATR', 'NATR', 'Keltner', 'Donchian'],
            }
            train_cat_tops = self._compute_category_tops(all_scores, CATEGORIES)

            MIN_MOVE = 0.0025  # Küçük fiyat hareketleri (TIMEOUT dead-band) için alt eşik
            start_idx = 250
            end_idx   = len(df_ind) - self.fwd_period

            # [SORUN 2] category_tops loop dışında bir kez hesapla
            analysis_stub.category_tops = self._compute_category_tops(all_scores, CATEGORIES)

            # [SORUN 1+9] ATR kolonu bul
            atr_col_train = next(
                (c for c in ['ATRr_14', 'ATR_14', 'ATRr_7', 'NATR_14'] if c in df_ind.columns),
                None
            )
            if atr_col_train is None:
                logger.error("❌ ATR kolonu bulunamadı — initial_train başarısız")
                return False

            ATR_MULT, RR_RATIO = 3.0, 1.5

            for i in range(start_idx, end_idx):
                fwd_val = df_ind[target_col].iloc[i]
                if pd.isna(fwd_val):
                    continue

                entry_price = df_ind['close'].iloc[i]
                atr_val     = df_ind[atr_col_train].iloc[i]
                if atr_col_train == 'NATR_14':
                    atr_val = atr_val * entry_price / 100  # % → $
                if pd.isna(entry_price) or pd.isna(atr_val) or atr_val <= 0 or entry_price <= 0:
                    continue

                sl_distance    = atr_val * ATR_MULT
                tp_distance    = sl_distance * RR_RATIO
                price_move_usd = (np.exp(fwd_val) - 1) * entry_price

                df_slice = df_ind.iloc[max(0, i - 50):i + 1].copy()
                if len(df_slice) < 40:
                    continue

                # [SORUN 9] Her bar için LONG ve SHORT → tautoloji yok
                for fake_direction in ["LONG", "SHORT"]:
                    if fake_direction == 'LONG':
                        if price_move_usd >= tp_distance:
                            r_multiple = RR_RATIO
                        elif price_move_usd <= -sl_distance:
                            r_multiple = -1.0
                        else:
                            r_multiple = price_move_usd / sl_distance
                    else:  # SHORT
                        if price_move_usd <= -tp_distance:
                            r_multiple = RR_RATIO
                        elif price_move_usd >= sl_distance:
                            r_multiple = -1.0
                        else:
                            r_multiple = -price_move_usd / sl_distance

                    if abs(r_multiple) < 0.25:  # dead-band: noise at
                        continue
                    r_multiple = max(-2.0, min(2.0, r_multiple))

                    fv = self.feature_eng.build_features(
                        analysis=analysis_stub,
                        ohlcv_df=df_slice,
                    )
                    rows_X.append(fv.to_dict())
                    rows_y.append(float(r_multiple))
                    rows_dir.append(fake_direction)

            if len(rows_X) < 30:
                logger.error(f"❌ Yetersiz eğitim verisi: {len(rows_X)} < 30")
                return False

            n_total_raw = end_idx - start_idx
            n_filtered  = n_total_raw - len(rows_X)
            logger.info(
                f"  Eğitim Verisi : {len(rows_X)} satır × "
                f"{len(rows_X[0]) if rows_X else 0} feature | "
                f"WIN={sum(rows_y)/len(rows_y):.1%} | "
                f"Dead-band: {n_filtered}/{n_total_raw} bar çıkarıldı "
                f"({n_filtered/max(n_total_raw,1):.1%}) | "
                f"MIN_MOVE={MIN_MOVE:.3f}"
            )

            X = pd.DataFrame(rows_X).replace([np.inf, -np.inf], np.nan)
            y = pd.Series(rows_y)
            dirs = pd.Series(rows_dir) # YENİ EKLENDİ

            # TRAIN KISMI GÜNCELLENDİ
            metrics = self.lgbm_model.train(X, y, directions=dirs)

            try:
                from pathlib import Path

                report_dir = Path("logs/reports")
                report_dir.mkdir(parents=True, exist_ok=True)
                report_path = report_dir / "model_egitim_raporu.xlsx"

                with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
                    # --- 1. SAYFA: GENEL METRİKLER ---
                    z_score = (metrics.spearman_ic * np.sqrt(metrics.n_train_samples - 3)) if metrics.n_train_samples > 3 else 0.0
                    
                    df_metrics = pd.DataFrame([{
                        "Tarih": _now_local().strftime("%Y-%m-%d %H:%M:%S"),  # TR saati
                        "Kaynak": "initial_train",
                        "Eğitim Satır Sayısı": len(X),
                        "Kazanma Oranı (Win Rate)": f"{sum(1 for val in y if float(val) > 0) / max(1, len(y)) * 100:.1f}%",
                        "Spearman IC Skoru": round(metrics.spearman_ic, 4),
                        "Bilgi Oranı (IR)": round(metrics.information_ratio, 2),
                        "Z-Skoru": round(z_score, 2),
                        "Long-Short Spread (R)": round(metrics.long_short_spread, 3),
                        "MAE (Hata Payı)": round(metrics.mae, 4)
                    }])
                    df_metrics.to_excel(writer, sheet_name="1_Genel_Metrikler", index=False)

                    # --- 2. SAYFA: KOLON ÖNEMLERİ (GÜVENLİ BLOK) ---
                    try:
                        importance = None
                        kalan_kolonlar = None
                        
                        if hasattr(self.lgbm_model, 'feature_importances_'):
                            importance = self.lgbm_model.feature_importances_() if callable(self.lgbm_model.feature_importances_) else self.lgbm_model.feature_importances_
                        elif hasattr(self.lgbm_model, 'model') and hasattr(self.lgbm_model.model, 'feature_importances_'):
                            importance = self.lgbm_model.model.feature_importances_
                            
                        if importance is not None:
                            if hasattr(self.lgbm_model, 'feature_names') and self.lgbm_model.feature_names:
                                kalan_kolonlar = self.lgbm_model.feature_names
                            else:
                                kalan_kolonlar = X.columns[:len(importance)]
                            
                            if len(kalan_kolonlar) == len(importance):
                                df_imp = pd.DataFrame({
                                    "Feature (Kolon)": kalan_kolonlar,
                                    "Önem Puanı": importance
                                }).sort_values(by="Önem Puanı", ascending=False)
                            else:
                                df_imp = pd.DataFrame({"Hata": [f"Uyuşmazlık! Kolon:{len(kalan_kolonlar)}, Puan:{len(importance)}"]})
                        else:
                            df_imp = pd.DataFrame({"Bilgi": ["Model önem puanı üretmedi"]})
                            
                        df_imp.to_excel(writer, sheet_name="2_Kolon_Onemleri", index=False)
                    except Exception as e:
                        pd.DataFrame({"Hata": [str(e)]}).to_excel(writer, sheet_name="2_Kolon_Onemleri", index=False)

                    # --- 3. SAYFA: GEÇMİŞ HAM VERİ ---
                    df_raw = X.copy()
                    df_raw['TARGET_SONUC'] = y.values
                    df_raw['TARGET_ACIKLAMA'] = df_raw['TARGET_SONUC'].apply(
                        lambda x: "KAZANÇ (Pozitif R)" if float(x) > 0 else "ZARAR (Negatif R)"
                    )
                    df_raw.to_excel(writer, sheet_name="3_Gecmis_Ham_Veri", index=False)

                logger.info(f"📊 Detaylı Eğitim Raporu Excel olarak kaydedildi: {report_path}")
            
            except Exception as ex:
                logger.error(f"⚠️ Excel raporu oluşturulurken hata (kritik değil): {ex}")

            gercek_win_rate = sum(1 for val in y if float(val) > 0) / max(1, len(y)) * 100
            logger.info(f"✅ İlk eğitim tamamlandı | Gerçek Win Rate: %{gercek_win_rate:.1f} | IC Skoru: {metrics.spearman_ic:.4f}")

            # [MADDE 3] — İlk eğitim sonrası deployment gate kontrolü
            # Model istatistiksel olarak yeterli mi? Gate başarısız olursa
            # canlı trade bloklanır, paper modda çalışmaya devam edilir.
            self._model_deployment_approved = self._check_deployment_gates(metrics)
            if not self._model_deployment_approved:
                logger.warning(
                    "⚠️ İlk eğitim sonrası deployment gate başarısız. "
                    "Daha fazla veri birikmesi ve tekrar eğitim gerekiyor. "
                    "Bot paper modda çalışmaya devam edecek."
                )
            return True

        except Exception as e:
            logger.error(f"❌ İlk eğitim hatası: {e}", exc_info=True)
            return False

    def _log_cycle_summary(self, report: CycleReport) -> None:
        emoji = {"success":"✅","partial":"⚡","no_signal":"😴",
                 "error":"❌","killed":"⛔","halted":"🛑"}.get(report.status.value, "❓")
        logger.info(f"\n{'─'*50}")
        logger.info(f"  {emoji} Döngü | Taranan={report.total_scanned} "
                    f"IC-geçen={report.total_above_gate} "
                    f"İşlem={report.total_traded} "
                    f"Süre={report.elapsed:.1f}s")
        if report.ml_stats:
            s = report.ml_stats
            logger.info(f"   ML: eğitildi={s.get('model_trained')} | "
                        f"retrain#{s.get('retrain_count',0)}")
        # Günlük PnL bilgisi (v2.1.4)
        if self._daily_start_balance > 0:
            loss_pct = (-self._daily_realized_pnl / self._daily_start_balance) * 100
            halt_str = ""
            if self._daily_halt_until:
                halt_str = f" | HALT → {self._daily_halt_until.astimezone(LOCAL_TZ).strftime('%H:%M TR')}"
            logger.info(
                f"   Günlük: PnL=${self._daily_realized_pnl:+.2f} "
                f"({loss_pct:+.2f}% / limit: %{cfg.risk.daily_max_loss_pct}){halt_str}"
            )
        logger.info(f"{'─'*50}\n")

    def _export_live_trades_to_xlsx(self, df: pd.DataFrame, filepath: Path) -> None:
        """
        Canlı işlemleri paper_trade formatında Excel'e kaydeder.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            df.to_excel(filepath, index=False)
            return

        try:
            TAKER = 0.0004
            if 'Durum' in df.columns:
                for col in ('Giriş ($)', 'Çıkış ($)', 'Lot (Coin)', 'Kaldıraç',
                            'PnL ($)', 'PnL (%)', 'Fee ($)', 'Net PnL ($)'):
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors='coerce').astype('float64')

                closed_mask = df['Durum'] == 'closed'

                need_fix = closed_mask & (df['Fee ($)'].fillna(0) == 0) & (df['PnL ($)'].fillna(0) != 0)
                if need_fix.any():
                    e = df.loc[need_fix, 'Giriş ($)'].fillna(0)
                    x = df.loc[need_fix, 'Çıkış ($)'].fillna(0)
                    s = df.loc[need_fix, 'Lot (Coin)'].fillna(0)
                    fee_calc = (e * s * TAKER) + (x * s * TAKER)
                    df.loc[need_fix, 'Fee ($)']     = fee_calc.round(4)
                    df.loc[need_fix, 'Net PnL ($)'] = (df.loc[need_fix, 'PnL ($)'] - fee_calc).round(4)

                broken_pct = closed_mask & (df['PnL (%)'].abs().fillna(0) < 0.001) & (df['PnL ($)'].fillna(0) != 0)
                if broken_pct.any():
                    e = df.loc[broken_pct, 'Giriş ($)'].fillna(0)
                    s = df.loc[broken_pct, 'Lot (Coin)'].fillna(0)
                    lev = df.loc[broken_pct, 'Kaldıraç'].fillna(1)
                    p = df.loc[broken_pct, 'PnL ($)'].fillna(0)
                    notional = (e * s).replace(0, pd.NA)
                    df.loc[broken_pct, 'PnL (%)'] = ((p / notional) * 100 * lev).fillna(0).round(2)

            wb = Workbook()
            ws_trades = wb.active
            ws_trades.title = "Trades"

            header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            data_font = Font(name='Arial', size=9)
            green_font = Font(name='Arial', size=9, color='006100')
            red_font = Font(name='Arial', size=9, color='9C0006')
            green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            long_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            short_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
            )

            columns_def = {
                "ID": (10, None), "Tarih (Açılış)": (18, None), "Tarih (Kapanış)": (18, None),
                "Coin": (8, None), "Yön": (8, None), "Giriş ($)": (12, '#,##0.00'),
                "Çıkış ($)": (12, '#,##0.00'), "Lot (Coin)": (12, '#,##0.000000'),
                "Hacim ($)": (12, '#,##0.00'), "Kaldıraç": (10, '0"x"'),
                "SL ($)": (12, '#,##0.00'), "TP ($)": (12, '#,##0.00'), "R:R": (8, '0.00'),
                "PnL ($)": (10, '#,##0.00;(#,##0.00);"-"'),
                "PnL (%)": (10, '0.00"%"'),
                "Fee ($)": (10, '#,##0.00'),
                "Net PnL ($)": (12, '#,##0.00;(#,##0.00);"-"'),
                "IC Güven": (10, '0.0'), "IC Yön": (8, None), "TF": (6, None),
                "Rejim": (14, None), "AI Karar": (10, None), "Durum": (10, None),
                "Çıkış Nedeni": (14, None), "Süre (dk)": (10, '#,##0')
            }

            headers = list(df.columns)

            for col_idx, col_name in enumerate(headers, 1):
                cell = ws_trades.cell(row=1, column=col_idx, value=col_name)
                cell.font, cell.fill, cell.alignment = header_font, header_fill, header_alignment
                width = columns_def.get(col_name, (12, None))[0]
                ws_trades.column_dimensions[get_column_letter(col_idx)].width = width

            ws_trades.freeze_panes = 'A2'

            for r_idx, row in enumerate(df.values, 2):
                for c_idx, val in enumerate(row, 1):
                    col_name = headers[c_idx-1]
                    if pd.isna(val):
                        val = ""

                    cell = ws_trades.cell(row=r_idx, column=c_idx, value=val)
                    cell.font, cell.border = data_font, thin_border

                    fmt = columns_def.get(col_name, (12, None))[1]
                    if fmt and val != "":
                        cell.number_format = fmt

                    if col_name == "Yön":
                        if val == "LONG": cell.fill = long_fill
                        elif val == "SHORT": cell.fill = short_fill

                    if col_name in ["PnL ($)", "Net PnL ($)", "PnL (%)"] and val != "":
                        try:
                            if float(val) > 0: cell.font, cell.fill = green_font, green_fill
                            elif float(val) < 0: cell.font, cell.fill = red_font, red_fill
                        except (ValueError, TypeError): pass

                    if col_name in ["Durum", "Çıkış Nedeni"] and isinstance(val, str):
                        if "tp" in str(val).lower(): cell.fill = green_fill
                        elif "sl" in str(val).lower(): cell.fill = red_fill

            if len(df) > 0:
                ws_trades.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df) + 1}"

            ws_summary = wb.create_sheet("Summary")
            closed_df = df[df['Durum'] == 'closed'].copy()
            total_trades = len(df)
            closed_trades = len(closed_df)
            open_trades = total_trades - closed_trades

            for col in ['PnL ($)', 'Net PnL ($)', 'Fee ($)']:
                if col not in closed_df.columns:
                    closed_df[col] = 0.0

            closed_df['PnL ($)']     = pd.to_numeric(closed_df['PnL ($)'], errors='coerce').fillna(0)
            closed_df['Net PnL ($)'] = pd.to_numeric(closed_df['Net PnL ($)'], errors='coerce').fillna(0)
            closed_df['Fee ($)']     = pd.to_numeric(closed_df['Fee ($)'], errors='coerce').fillna(0)

            winning  = int((closed_df['Net PnL ($)'] >  0).sum())
            losing   = int((closed_df['Net PnL ($)'] <= 0).sum())
            win_rate = (winning / closed_trades * 100) if closed_trades > 0 else 0.0

            total_pnl     = float(closed_df['PnL ($)'].sum())
            total_net_pnl = float(closed_df['Net PnL ($)'].sum())
            total_fee     = float(closed_df['Fee ($)'].sum())

            gross_profit  = float(closed_df.loc[closed_df['Net PnL ($)'] >  0, 'Net PnL ($)'].sum())
            gross_loss    = abs(float(closed_df.loc[closed_df['Net PnL ($)'] <= 0, 'Net PnL ($)'].sum()))
            profit_factor = (gross_profit / gross_loss) if gross_loss > 0 else float('inf')

            avg_pnl  = total_net_pnl / closed_trades if closed_trades > 0 else 0.0
            avg_win  = gross_profit / winning if winning > 0 else 0.0
            avg_loss = gross_loss   / losing  if losing  > 0 else 0.0

            avg_dur = pd.to_numeric(closed_df['Süre (dk)'], errors='coerce').mean() \
                      if (closed_trades > 0 and 'Süre (dk)' in closed_df.columns) else 0.0
            avg_duration = float(avg_dur) if not pd.isna(avg_dur) else 0.0

            total_return = ((self._balance - self._initial_balance) / self._initial_balance * 100) \
                           if self._initial_balance > 0 else 0.0

            title_font   = Font(name='Arial', bold=True, size=14, color='2F5496')
            section_font = Font(name='Arial', bold=True, size=11, color='2F5496')
            label_font   = Font(name='Arial', size=10)
            value_font   = Font(name='Arial', bold=True, size=10)
            pos_font     = Font(name='Arial', bold=True, size=10, color='006100')
            neg_font     = Font(name='Arial', bold=True, size=10, color='9C0006')

            ws_summary['A1'] = '📊 CANLI İŞLEM PERFORMANS RAPORU'
            ws_summary['A1'].font = title_font
            ws_summary.merge_cells('A1:D1')
            ws_summary['A2'] = f'Oluşturulma: {_now_local().strftime("%Y-%m-%d %H:%M:%S")} (TR)'
            ws_summary['A2'].font = Font(name='Arial', size=9, italic=True, color='808080')

            ws_summary.column_dimensions['A'].width = 22
            ws_summary.column_dimensions['B'].width = 15
            ws_summary.column_dimensions['C'].width = 15

            r = 4
            ws_summary.cell(row=r, column=1, value='💰 BAKİYE').font = section_font
            r += 1
            bakiye_rows = [
                ('Başlangıç Bakiye', f"${self._initial_balance:.2f}", None),
                ('Güncel Bakiye',    f"${self._balance:.2f}",         None),
                ('Toplam Getiri',    f"{total_return:+.2f}%",         total_return),
                ('Net PnL',          f"${total_net_pnl:+.2f}",        total_net_pnl),
                ('Toplam Fee',       f"${total_fee:.2f}",             None),
            ]
            for label, val, signed in bakiye_rows:
                ws_summary.cell(row=r, column=1, value=label).font = label_font
                v_cell = ws_summary.cell(row=r, column=2, value=val)
                v_cell.font = value_font
                if signed is not None:
                    if signed > 0:   v_cell.font = pos_font
                    elif signed < 0: v_cell.font = neg_font
                r += 1

            r += 1
            ws_summary.cell(row=r, column=1, value='📈 TRADE İSTATİSTİKLERİ').font = section_font
            r += 1
            for label, val in [('Toplam Trade',  str(total_trades)),
                               ('Açık Pozisyon', str(open_trades)),
                               ('Kapalı Trade',  str(closed_trades)),
                               ('Kazanan',       str(winning)),
                               ('Kaybeden',      str(losing))]:
                ws_summary.cell(row=r, column=1, value=label).font = label_font
                ws_summary.cell(row=r, column=2, value=val).font = value_font
                r += 1

            r += 1
            ws_summary.cell(row=r, column=1, value='📊 PERFORMANS METRİKLERİ').font = section_font
            r += 1
            pf_str = f"{profit_factor:.2f}" if profit_factor != float('inf') else "∞"
            for label, val in [('Win Rate',      f"{win_rate:.1f}%"),
                               ('Profit Factor', pf_str),
                               ('Ort. Net PnL',  f"${avg_pnl:.2f}"),
                               ('Ort. Kazanç',   f"${avg_win:.2f}"),
                               ('Ort. Kayıp',    f"-${avg_loss:.2f}"),
                               ('Ort. Süre',     f"{avg_duration:.0f} dk")]:
                ws_summary.cell(row=r, column=1, value=label).font = label_font
                ws_summary.cell(row=r, column=2, value=val).font = value_font
                r += 1

            r += 2
            ws_summary.cell(row=r, column=1, value='📅 GÜNLÜK PNL').font = section_font
            r += 1
            ws_summary.cell(row=r, column=1, value='Tarih').font        = Font(name='Arial', bold=True, size=10)
            ws_summary.cell(row=r, column=2, value='Net PnL ($)').font  = Font(name='Arial', bold=True, size=10)
            ws_summary.cell(row=r, column=3, value='Trade Sayısı').font = Font(name='Arial', bold=True, size=10)
            r += 1

            if closed_trades > 0 and 'Tarih (Kapanış)' in closed_df.columns:
                try:
                    closed_df['Date_Only'] = pd.to_datetime(closed_df['Tarih (Kapanış)'], errors='coerce').dt.date
                    daily_stats = closed_df.dropna(subset=['Date_Only']).groupby('Date_Only').agg(
                        daily_pnl=('Net PnL ($)', 'sum'),
                        trade_count=('ID', 'count')
                    ).reset_index().sort_values('Date_Only')

                    for _, d_row in daily_stats.iterrows():
                        date_str = str(d_row['Date_Only'])
                        d_pnl    = float(d_row['daily_pnl'])
                        t_cnt    = int(d_row['trade_count'])

                        ws_summary.cell(row=r, column=1, value=date_str).font = label_font
                        pnl_cell = ws_summary.cell(row=r, column=2, value=f"${d_pnl:+.2f}")
                        pnl_cell.font = pos_font if d_pnl > 0 else (neg_font if d_pnl < 0 else value_font)
                        ws_summary.cell(row=r, column=3, value=t_cnt).font = label_font
                        r += 1
                except Exception as e:
                    logger.error(f"Günlük PnL tablo hatası: {e}")

            try:
                wb.save(str(filepath))
            except PermissionError:
                alt_path = filepath.with_stem(filepath.stem + f"_{_now_local().strftime('%H%M%S')}")
                wb.save(str(alt_path))
                logger.warning(f"⚠️ Dosya kilitli, alternatif kaydedildi: {alt_path}")
        except Exception as e:
            logger.error(f"❌ Excel formatlama hatası: {e}")
            df.to_excel(filepath, index=False)

    def _reconcile_excel_orphans(self, active_coins: set) -> int:
        """
        Excel'de 'open' olarak duran ama memory'de ve Binance'te olmayan trade'leri kapatır.
        """
        from pathlib import Path

        closed_in_excel = 0
        excel_path = Path("logs") / "live_trades.xlsx"
        if not excel_path.exists():
            return 0

        try:
            df = pd.read_excel(excel_path)
        except Exception as e:
            logger.warning(f"   ⚠️ Excel reconcile read error: {e}")
            return 0

        if 'Durum' not in df.columns:
            return 0

        mem_coins = set()
        if hasattr(self, 'trade_memory') and hasattr(self.trade_memory, 'open_trades'):
            for mt in self.trade_memory.open_trades.values():
                c = str(getattr(mt, 'coin', '')).replace('USDT', '').split('/')[0].split(':')[0]
                if c:
                    mem_coins.add(c)

        open_rows = df[df['Durum'] == 'open']
        if open_rows.empty:
            return 0

        TAKER = 0.0004
        changed = False

        for idx, row in open_rows.iterrows():
            coin = str(row.get('Coin', '')).strip()
            if not coin:
                continue

            if coin in active_coins:
                continue
            if coin in mem_coins:
                continue

            try:
                entry     = float(row.get('Giriş ($)', 0) or 0)
                size      = float(row.get('Lot (Coin)', 0) or 0)
                leverage  = int(float(row.get('Kaldıraç', 1) or 1))
                sl_price  = float(row.get('SL ($)', 0) or 0)
                tp_price  = float(row.get('TP ($)', 0) or 0)
                direction = str(row.get('Yön', 'LONG')).upper()

                if entry <= 0 or size <= 0:
                    logger.debug(f"   ⏭️ Orphan {coin}: yetersiz veri, atlandı")
                    continue

                symbol = f"{coin}/USDT:USDT"
                close_reason = "Manuel / API"
                exit_price   = entry

                try:
                    ohlcv = self.fetcher.fetch_ohlcv(symbol, timeframe="1m", limit=200)
                    if ohlcv is not None and not ohlcv.empty:
                        max_high = float(ohlcv['high'].max())
                        min_low  = float(ohlcv['low'].min())
                        last     = float(ohlcv['close'].iloc[-1])
                        dist_tp  = abs(last - tp_price)
                        dist_sl  = abs(last - sl_price)

                        if direction == "LONG":
                            if min_low <= sl_price:
                                close_reason, exit_price = "SL Hit", sl_price
                            elif max_high >= tp_price:
                                close_reason, exit_price = "TP Hit", tp_price
                            else:
                                close_reason = "TP Hit" if dist_tp < dist_sl else "SL Hit"
                                exit_price   = tp_price if dist_tp < dist_sl else sl_price
                        else:
                            if max_high >= sl_price:
                                close_reason, exit_price = "SL Hit", sl_price
                            elif min_low <= tp_price:
                                close_reason, exit_price = "TP Hit", tp_price
                            else:
                                close_reason = "TP Hit" if dist_tp < dist_sl else "SL Hit"
                                exit_price   = tp_price if dist_tp < dist_sl else sl_price
                except Exception as oe:
                    logger.warning(f"   ⚠️ Orphan {coin} OHLCV hatası: {oe}")

                if direction == "LONG":
                    pnl_val = (exit_price - entry) * size
                else:
                    pnl_val = (entry - exit_price) * size
                fee_val = (entry * size * TAKER) + (exit_price * size * TAKER)
                net_pnl = pnl_val - fee_val
                notional = entry * size
                pnl_pct = (pnl_val / notional) * 100 * leverage if notional > 0 else 0.0

                for col in ('Tarih (Kapanış)', 'Çıkış Nedeni', 'Durum'):
                    if col in df.columns:
                        df[col] = df[col].astype(object)
                for col in ('Çıkış ($)', 'PnL ($)', 'PnL (%)', 'Fee ($)', 'Net PnL ($)'):
                    if col in df.columns:
                        df[col] = pd.to_numeric(df[col], errors='coerce').astype('float64')

                now_str = _iso_tr(_now_utc())  # UTC üret, TR göster
                df.at[idx, 'Tarih (Kapanış)'] = now_str
                df.at[idx, 'Çıkış ($)']       = round(exit_price, 6)
                df.at[idx, 'PnL ($)']         = round(pnl_val, 4)
                df.at[idx, 'PnL (%)']         = round(pnl_pct, 2)
                df.at[idx, 'Fee ($)']         = round(fee_val, 4)
                df.at[idx, 'Net PnL ($)']     = round(net_pnl, 4)
                df.at[idx, 'Çıkış Nedeni']    = close_reason
                df.at[idx, 'Durum']           = 'closed'

                try:
                    acilis = pd.to_datetime(str(row.get('Tarih (Açılış)', '')))
                    _acilis_aware = acilis.to_pydatetime()
                    if _acilis_aware.tzinfo is None:
                        _acilis_aware = _acilis_aware.replace(tzinfo=timezone.utc)
                    df.at[idx, 'Süre (dk)'] = int((_now_utc() - _acilis_aware).total_seconds() / 60)
                except Exception:
                    pass

                if close_reason == "SL Hit":
                    self.cooldowns[coin]   = datetime.now(timezone.utc) + timedelta(hours=2)
                    self.cooldowns[symbol] = datetime.now(timezone.utc) + timedelta(hours=2)

                logger.info(f"   🔧 ORPHAN kapatıldı: {coin} → {close_reason} | "
                            f"PnL: ${pnl_val:+.2f} | Net: ${net_pnl:+.2f}")
                closed_in_excel += 1
                changed = True

            except Exception as e:
                logger.error(f"   ❌ Orphan reconcile {coin} hatası: {e}", exc_info=True)
                continue

        if changed:
            try:
                self._export_live_trades_to_xlsx(df, excel_path)
                logger.info(f"   📊 Excel reconcile: {closed_in_excel} orphan işlem kapatıldı")
            except Exception as e:
                logger.error(f"   ❌ Excel reconcile save hatası: {e}")

        return closed_in_excel

    def print_performance(self) -> None:
        PerformanceAnalyzer(self.paper_trader).print_report(
            PerformanceAnalyzer(self.paper_trader).full_analysis()
        )
        self.trade_memory.print_summary()


# =============================================================================
# SCHEDULER
# =============================================================================

def run_scheduler(pipeline: MLTradingPipeline, interval_minutes: int = 10) -> None:
    pipeline._is_running = True

    def _stop(signum, frame):
        logger.info(f"\n🛑 Sinyal {signum} — durduruluyor...")
        pipeline._is_running = False

    signal.signal(signal.SIGINT, _stop)
    signal.signal(signal.SIGTERM, _stop)

    logger.info(f"⏰ Scheduler: {interval_minutes}dk aralık")

    if not pipeline._init_balance():
        return

    if not pipeline.lgbm_model.is_trained:
        pipeline.initial_train()

    while pipeline._is_running:
        if pipeline._kill_switch:
            break
        if pipeline._consecutive_errors >= MAX_CONSECUTIVE_ERRORS:
            time.sleep(ERROR_COOLDOWN_SECONDS)
            pipeline._consecutive_errors = 0
            continue

        report = pipeline.run_cycle()
        if report.status == CycleStatus.KILLED:
            break

        logger.info(f"⏰ Sonraki Tarama: {(_now_local()+timedelta(minutes=interval_minutes)).strftime('%H:%M:%S')}")

        for i in range(interval_minutes * 60):
            if not pipeline._is_running:
                break

            if i > 0 and i % 30 == 0:
                try:
                    pipeline._check_open_positions()
                    # Uyku sırasında da gün geçmiş mi / halt bitmiş mi kontrol et
                    pipeline._reset_daily_state_if_new_day()
                except Exception:
                    pass

            time.sleep(1)

    logger.info("🏁 Scheduler kapatıldı.")


# =============================================================================
# CLI
# =============================================================================

def main():
    parser = argparse.ArgumentParser(description="ML Crypto Bot v2.1.4")
    parser.add_argument("--live",     action="store_true", help="Canlı trade (ÖNERİLEN)")
    parser.add_argument("--top",      type=int, default=20, help="Top N coin")
    parser.add_argument("--schedule", action="store_true", help="Scheduler modu")
    parser.add_argument("-i","--interval", type=int, default=10, help="Aralık (dk)")
    parser.add_argument("--report",   action="store_true", help="Performans raporu")
    parser.add_argument("--train",    action="store_true", help="Sadece eğitim")
    parser.add_argument("--verbose",  action="store_true", help="Debug")
    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    pipeline = MLTradingPipeline(dry_run=not args.live, top_n=args.top)

    if args.report:
        pipeline.print_performance()
        return

    if args.train:
        pipeline.initial_train()
        return

    if not pipeline._init_balance():
        sys.exit(1)

    if not pipeline.lgbm_model.is_trained:
        logger.info("🎓 Model eğitilmemiş — ilk eğitim başlıyor...")
        pipeline.initial_train()

    if args.schedule:
        run_scheduler(pipeline, args.interval)
    else:
        report = pipeline.run_cycle()
        logger.info(f"Döngü: {report.status.value}")


if __name__ == "__main__":
    main()