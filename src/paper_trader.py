# =============================================================================
# PAPER TRADER — Trade Kayıt ve Simülasyon Modülü (ADIM 10)
# =============================================================================
# Bu modül paper trade modunda tüm işlemleri kaydeder ve takip eder.
# Gerçek para kullanmadan stratejinin performansını ölçmeye yarar.
#
# Özellikler:
# - Trade kayıtlarını JSON formatında sakla
# - Açık pozisyonları takip et
# - SL/TP tetiklenme simülasyonu
# - Günlük/haftalık PnL özeti
# - CSV export (analiz için)
#
# Kullanım:
#   from paper_trader import PaperTrader
#   pt = PaperTrader(initial_balance=75.0)
#   pt.open_trade(symbol='BTC', direction='LONG', entry=95000, ...)
#   pt.check_exits(current_prices={'BTC': 96000})
#   pt.get_summary()
#
# =============================================================================

import json                                    # Trade kayıtlarını JSON formatında sakla
import csv                                     # CSV export için
import os                                      # Dosya/klasör işlemleri
import logging                                 # Loglama
from pathlib import Path                       # Platform-bağımsız path
from datetime import datetime, timedelta      # Zaman damgaları
from typing import Dict, List, Optional, Tuple, Any  # Tip belirteçleri
from dataclasses import dataclass, field, asdict  # Yapılandırılmış veri
from enum import Enum                          # Trade durumları
import uuid                                    # Benzersiz trade ID

# Logger yapılandırması
logger = logging.getLogger(__name__)


# =============================================================================
# SABİTLER VE ENUM'LAR
# =============================================================================

class TradeStatus(Enum):
    """Trade durumu enumları."""
    OPEN = "open"                              # Pozisyon açık
    CLOSED_TP = "closed_tp"                    # Take-profit ile kapandı
    CLOSED_SL = "closed_sl"                    # Stop-loss ile kapandı
    CLOSED_MANUAL = "closed_manual"            # Manuel kapandı
    CLOSED_KILL = "closed_kill"                # Kill switch ile kapandı
    CLOSED_TIMEOUT = "closed_timeout"          # 24 saat TIMEOUT ile kapandı


class TradeDirection(Enum):
    """Trade yönü enumları."""
    LONG = "LONG"                              # Alış pozisyonu
    SHORT = "SHORT"                            # Satış pozisyonu


# Varsayılan kayıt dizini
DEFAULT_LOG_DIR = Path(__file__).parent.parent / "logs" / "paper_trades"


# =============================================================================
# TRADE KAYIT DATACLASS'I
# =============================================================================

@dataclass
class PaperTrade:
    """
    Tek bir paper trade kaydı.
    
    Tüm trade bilgilerini içerir: giriş, çıkış, PnL, metadata.
    JSON serializable olması için dataclass kullanılıyor.
    """
    # ---- Benzersiz Tanımlayıcılar ----
    trade_id: str                              # UUID formatında benzersiz ID
    symbol: str                                # Coin sembolü (örn: 'BTC', 'ETH')
    full_symbol: str                           # Tam sembol (örn: 'BTC/USDT:USDT')
    
    # ---- Trade Parametreleri ----
    direction: str                             # 'LONG' veya 'SHORT'
    entry_price: float                         # Giriş fiyatı ($)
    position_size: float                       # Pozisyon büyüklüğü (coin miktarı)
    position_value: float                      # Pozisyon değeri ($)
    leverage: int                              # Kaldıraç (örn: 5x)
    
    # ---- Risk Parametreleri ----
    stop_loss: float                           # Stop-loss fiyatı
    take_profit: float                         # Take-profit fiyatı
    risk_amount: float                         # Risk edilen miktar ($)
    risk_reward: float                         # Risk/Reward oranı
    
    # ---- IC & AI Metadata ----
    ic_confidence: float                       # IC güven skoru (0-100)
    ic_direction: str                          # IC'nin önerdiği yön
    best_timeframe: str                        # En iyi zaman dilimi
    market_regime: str                         # Piyasa rejimi (trending/ranging)
    ai_decision: Optional[str] = None          # ML Modeli kararı (LONG/SHORT/WAIT)
    ai_confidence: Optional[float] = None      # AI güven skoru
    
    # ---- Zaman Damgaları ----
    opened_at: str = ""                        # Açılış zamanı (ISO format)
    closed_at: Optional[str] = None            # Kapanış zamanı
    duration_minutes: Optional[int] = None     # Trade süresi (dakika)
    
    # ---- Çıkış Bilgileri ----
    status: str = TradeStatus.OPEN.value       # Trade durumu
    exit_price: Optional[float] = None         # Çıkış fiyatı
    exit_reason: Optional[str] = None          # Çıkış sebebi (TP/SL/Manual)
    
    # ---- PnL Hesaplamaları ----
    pnl_absolute: Optional[float] = None       # Kar/Zarar ($)
    pnl_percent: Optional[float] = None        # Kar/Zarar (%)
    fees: float = 0.0                          # Tahmini işlem ücreti
    net_pnl: Optional[float] = None            # Net kar (ücret düşülmüş)
    
    # ---- Ek Bilgiler ----
    notes: str = ""                            # Notlar
    tags: List[str] = field(default_factory=list)  # Etiketler (filtreleme için)

    def __post_init__(self):
        """Trade açıldığında zaman damgası ata."""
        if not self.opened_at:
            self.opened_at = datetime.now().isoformat()
        if not self.trade_id:
            self.trade_id = str(uuid.uuid4())[:8]  # Kısa UUID

    def to_dict(self) -> Dict[str, Any]:
        """Trade'i sözlük formatına çevir (JSON için)."""
        return asdict(self)

    def calculate_pnl(self, exit_price: float) -> Tuple[float, float]:
        """
        Belirtilen çıkış fiyatı için PnL hesapla.
        
        Parameters:
        ----------
        exit_price : float
            Çıkış fiyatı
            
        Returns:
        -------
        Tuple[float, float]
            (pnl_absolute, pnl_percent)
        """
        if self.direction == TradeDirection.LONG.value:
            # LONG: Fiyat yükselirse kâr
            price_change = exit_price - self.entry_price
        else:
            # SHORT: Fiyat düşerse kâr
            price_change = self.entry_price - exit_price
        
        # Mutlak PnL = fiyat değişimi × pozisyon büyüklüğü
        pnl_absolute = price_change * self.position_size
        
        # Yüzde PnL = (mutlak PnL / risk edilen miktar) × 100
        # veya basitçe: (çıkış - giriş) / giriş × kaldıraç × 100
        pnl_percent = (price_change / self.entry_price) * self.leverage * 100
        
        return pnl_absolute, pnl_percent

    def close(
        self,
        exit_price: float,
        status: TradeStatus,
        reason: str = "",
        fee_rate: float = 0.0006              # Binance maker fee: %0.06
    ) -> None:
        """
        Trade'i kapat ve PnL hesapla.
        
        Parameters:
        ----------
        exit_price : float
            Kapanış fiyatı
        status : TradeStatus
            Kapanış durumu (TP, SL, Manual, Kill)
        reason : str
            Kapanış sebebi açıklaması
        fee_rate : float
            İşlem ücreti oranı (varsayılan: %0.06)
        """
        self.exit_price = exit_price
        self.status = status.value
        self.exit_reason = reason or status.value
        self.closed_at = datetime.now().isoformat()
        
        # Süre hesapla
        opened = datetime.fromisoformat(self.opened_at)
        closed = datetime.fromisoformat(self.closed_at)
        self.duration_minutes = int((closed - opened).total_seconds() / 60)
        
        # PnL hesapla
        self.pnl_absolute, self.pnl_percent = self.calculate_pnl(exit_price)
        
        # Ücret hesapla (giriş + çıkış)
        self.fees = self.position_value * fee_rate * 2
        
        # Net PnL
        self.net_pnl = self.pnl_absolute - self.fees


# =============================================================================
# PAPER TRADER ANA SINIFI
# =============================================================================

class PaperTrader:
    """
    Paper trading yönetici sınıfı.
    
    Tüm paper trade'leri yönetir:
    - Yeni trade aç
    - Açık pozisyonları takip et
    - SL/TP tetiklenmesini simüle et
    - Trade geçmişini sakla ve raporla
    """

    def __init__(
        self,
        initial_balance: float = 1000.0,         # Başlangıç bakiyesi ($)
        log_dir: Path = DEFAULT_LOG_DIR,       # Kayıt dizini
        fee_rate: float = 0.0006,              # İşlem ücreti (%0.06)
        auto_save: bool = True,                # Her trade'de otomatik kaydet
    ):
        """
        Paper Trader'ı başlat.
        
        Parameters:
        ----------
        initial_balance : float
            Başlangıç bakiyesi (varsayılan: $75)
        log_dir : Path
            Trade loglarının kaydedileceği dizin
        fee_rate : float
            İşlem ücreti oranı (giriş + çıkış için 2x uygulanır)
        auto_save : bool
            Her işlemde otomatik JSON kaydı
        """
        self.initial_balance = initial_balance  # Başlangıç sermayesi
        self.balance = initial_balance          # Güncel bakiye
        self.log_dir = Path(log_dir)           # Log dizini
        self.fee_rate = fee_rate               # Fee oranı
        self.auto_save = auto_save             # Otomatik kayıt
        
        # ---- Trade Koleksiyonları ----
        self.open_trades: Dict[str, PaperTrade] = {}    # Açık pozisyonlar {trade_id: trade}
        self.closed_trades: List[PaperTrade] = []       # Kapatılan tradeler
        self.all_trades: List[PaperTrade] = []          # Tüm tradeler (açık + kapalı)
        
        # ---- İstatistikler ----
        self.total_trades = 0                  # Toplam trade sayısı
        self.winning_trades = 0                # Kazanan trade sayısı
        self.losing_trades = 0                 # Kaybeden trade sayısı
        self.total_pnl = 0.0                   # Toplam PnL ($)
        self.total_fees = 0.0                  # Toplam ödenen ücret
        self.peak_balance = initial_balance    # En yüksek bakiye (drawdown için)
        self.max_drawdown = 0.0                # Maksimum drawdown (%)
        
        # ---- Günlük İstatistikler ----
        self.daily_pnl: Dict[str, float] = {}  # {tarih: pnl}
        self.daily_trades: Dict[str, int] = {} # {tarih: trade_sayısı}
        
        # ---- Kayıt Dizinini Oluştur ----
        self.log_dir.mkdir(parents=True, exist_ok=True)
        
        # Mevcut logları yükle (varsa)
        self._load_existing_trades()
        
        logger.info(
            f"📝 PaperTrader başlatıldı | "
            f"Bakiye: ${self.balance:.2f} | "
            f"Log: {self.log_dir}"
        )

    # =========================================================================
    # TRADE AÇMA
    # =========================================================================

    def open_trade(
        self,
        symbol: str,                           # Kısa sembol (örn: 'BTC')
        full_symbol: str,                      # Tam sembol (örn: 'BTC/USDT:USDT')
        direction: str,                        # 'LONG' veya 'SHORT'
        entry_price: float,                    # Giriş fiyatı
        position_size: float,                  # Pozisyon büyüklüğü (coin)
        stop_loss: float,                      # Stop-loss fiyatı
        take_profit: float,                    # Take-profit fiyatı
        leverage: int = 5,                     # Kaldıraç
        ic_confidence: float = 0.0,            # IC skoru
        ic_direction: str = "",                # IC yönü
        best_timeframe: str = "",              # En iyi TF
        market_regime: str = "",               # Piyasa rejimi
        ai_decision: Optional[str] = None,     # AI kararı
        ai_confidence: Optional[float] = None, # AI güveni
        notes: str = "",                       # Notlar
        tags: List[str] = None,                # Etiketler
    ) -> PaperTrade:
        """
        Yeni bir paper trade aç.
        
        Parameters:
        ----------
        (Tüm parametreler PaperTrade dataclass'ındaki alanları doldurur)
        
        Returns:
        -------
        PaperTrade
            Açılan trade objesi
        """
        # Pozisyon değeri hesapla
        position_value = entry_price * position_size
        
        # Risk miktarı hesapla (SL'ye kadar olan kayıp)
        if direction == TradeDirection.LONG.value:
            risk_per_unit = entry_price - stop_loss
        else:
            risk_per_unit = stop_loss - entry_price
        risk_amount = risk_per_unit * position_size
        
        # Risk/Reward hesapla
        if direction == TradeDirection.LONG.value:
            reward_per_unit = take_profit - entry_price
        else:
            reward_per_unit = entry_price - take_profit
        risk_reward = reward_per_unit / risk_per_unit if risk_per_unit > 0 else 0
        
        # Trade objesi oluştur
        trade = PaperTrade(
            trade_id=str(uuid.uuid4())[:8],
            symbol=symbol,
            full_symbol=full_symbol,
            direction=direction,
            entry_price=entry_price,
            position_size=position_size,
            position_value=position_value,
            leverage=leverage,
            stop_loss=stop_loss,
            take_profit=take_profit,
            risk_amount=risk_amount,
            risk_reward=risk_reward,
            ic_confidence=ic_confidence,
            ic_direction=ic_direction,
            best_timeframe=best_timeframe,
            market_regime=market_regime,
            ai_decision=ai_decision,
            ai_confidence=ai_confidence,
            notes=notes,
            tags=tags or [],
        )
        
        # Koleksiyonlara ekle
        self.open_trades[trade.trade_id] = trade
        self.all_trades.append(trade)
        self.total_trades += 1
        
        # Otomatik kaydet
        if self.auto_save:
            self._save_trades()
        
        logger.info(
            f"📈 Trade açıldı: {trade.trade_id} | "
            f"{symbol} {direction} @ ${entry_price:,.2f} | "
            f"Size: {position_size:.4f} | "
            f"SL: ${stop_loss:,.2f} | TP: ${take_profit:,.2f} | "
            f"RR: {risk_reward:.1f}"
        )
        
        return trade

    # =========================================================================
    # EXIT KONTROL (SL/TP SİMÜLASYONU)
    # =========================================================================

    def check_exits(
        self,
        current_prices: Dict[str, float],      # {symbol: fiyat}
    ) -> List[PaperTrade]:
        """
        Açık pozisyonlar için SL/TP tetiklenmesini kontrol et.
        
        Bu metod her fiyat güncellemesinde çağrılmalı.
        Simülasyonda: scheduler her çalıştığında güncel fiyatlar ile.
        
        Parameters:
        ----------
        current_prices : dict
            Güncel fiyatlar {symbol: price}
            
        Returns:
        -------
        List[PaperTrade]
            Bu çağrıda kapanan tradeler
        """
        closed_this_cycle = []
        
        for trade_id, trade in list(self.open_trades.items()):
            current_price = current_prices.get(trade.symbol)
            
            if current_price is None:
                continue                       # Bu coin için fiyat yok, atla
            
            # ---- LONG POZİSYON KONTROLÜ ----
            if trade.direction == TradeDirection.LONG.value:
                # Stop-loss tetiklendi mi?
                if current_price <= trade.stop_loss:
                    # SL tetiklendiğinde SL fiyatından çık (limit emir gibi)
                    self._close_trade(trade, trade.stop_loss, TradeStatus.CLOSED_SL)
                    closed_this_cycle.append(trade)
                # Take-profit tetiklendi mi?
                elif current_price >= trade.take_profit:
                    # TP tetiklendiğinde TP fiyatından çık (limit emir gibi)
                    self._close_trade(trade, trade.take_profit, TradeStatus.CLOSED_TP)
                    closed_this_cycle.append(trade)
            
            # ---- SHORT POZİSYON KONTROLÜ ----
            else:  # SHORT
                # Stop-loss tetiklendi mi?
                if current_price >= trade.stop_loss:
                    # SL tetiklendiğinde SL fiyatından çık
                    self._close_trade(trade, trade.stop_loss, TradeStatus.CLOSED_SL)
                    closed_this_cycle.append(trade)
                # Take-profit tetiklendi mi?
                elif current_price <= trade.take_profit:
                    # TP tetiklendiğinde TP fiyatından çık
                    self._close_trade(trade, trade.take_profit, TradeStatus.CLOSED_TP)
                    closed_this_cycle.append(trade)
        
        return closed_this_cycle

    def close_trade_manual(
        self,
        trade_id: str,
        exit_price: float,
        reason: str = "Manual close"
    ) -> Optional[PaperTrade]:
        """
        Trade'i manuel olarak kapat.
        
        Parameters:
        ----------
        trade_id : str
            Kapatılacak trade ID
        exit_price : float
            Çıkış fiyatı
        reason : str
            Kapanış sebebi
            
        Returns:
        -------
        PaperTrade or None
            Kapatılan trade (bulunamazsa None)
        """
        trade = self.open_trades.get(trade_id)
        if trade:
            self._close_trade(trade, exit_price, TradeStatus.CLOSED_MANUAL, reason)
            return trade
        return None

    def close_all_trades(
        self,
        current_prices: Dict[str, float],
        reason: str = "Kill switch"
    ) -> List[PaperTrade]:
        """
        Tüm açık pozisyonları kapat (kill switch).
        
        Parameters:
        ----------
        current_prices : dict
            Güncel fiyatlar
        reason : str
            Kapanış sebebi
            
        Returns:
        -------
        List[PaperTrade]
            Kapatılan tradeler
        """
        closed = []
        for trade_id, trade in list(self.open_trades.items()):
            price = current_prices.get(trade.symbol, trade.entry_price)
            self._close_trade(trade, price, TradeStatus.CLOSED_KILL, reason)
            closed.append(trade)
        return closed

    def _close_trade(
        self,
        trade: PaperTrade,
        exit_price: float,
        status: TradeStatus,
        reason: str = ""
    ) -> None:
        """
        Dahili trade kapatma metodu.
        
        PnL hesaplar, istatistikleri günceller, bakiyeyi ayarlar.
        """
        # Trade'i kapat
        trade.close(exit_price, status, reason, self.fee_rate)
        
        # Koleksiyonlardan taşı
        del self.open_trades[trade.trade_id]
        self.closed_trades.append(trade)
        
        # ---- İSTATİSTİKLERİ GÜNCELLE ----
        
        # Win/Loss sayaçları
        if trade.net_pnl > 0:
            self.winning_trades += 1
        else:
            self.losing_trades += 1
        
        # Toplam PnL ve ücretler
        self.total_pnl += trade.net_pnl
        self.total_fees += trade.fees
        
        # Bakiyeyi güncelle
        self.balance += trade.net_pnl
        
        # Peak balance ve drawdown
        if self.balance > self.peak_balance:
            self.peak_balance = self.balance
        current_drawdown = (self.peak_balance - self.balance) / self.peak_balance * 100
        if current_drawdown > self.max_drawdown:
            self.max_drawdown = current_drawdown
        
        # Günlük istatistikler
        today = datetime.now().strftime('%Y-%m-%d')
        self.daily_pnl[today] = self.daily_pnl.get(today, 0.0) + trade.net_pnl
        self.daily_trades[today] = self.daily_trades.get(today, 0) + 1
        
        # Otomatik kaydet
        if self.auto_save:
            self._save_trades()
        
        # Log
        emoji = "✅" if trade.net_pnl > 0 else "❌"
        logger.info(
            f"{emoji} Trade kapandı: {trade.trade_id} | "
            f"{trade.symbol} {trade.direction} | "
            f"Exit: ${exit_price:,.2f} ({status.value}) | "
            f"PnL: ${trade.net_pnl:+.2f} ({trade.pnl_percent:+.1f}%) | "
            f"Bakiye: ${self.balance:.2f}"
        )

    # =========================================================================
    # RAPORLAMA
    # =========================================================================

    def get_summary(self) -> Dict[str, Any]:
        """
        Genel performans özeti döndür.
        
        Returns:
        -------
        dict
            Tüm performans metrikleri
        """
        # Win rate hesapla
        total_closed = len(self.closed_trades)
        win_rate = (self.winning_trades / total_closed * 100) if total_closed > 0 else 0
        
        # Ortalama trade metrikleri
        if total_closed > 0:
            avg_pnl = self.total_pnl / total_closed
            avg_win = sum(t.net_pnl for t in self.closed_trades if t.net_pnl > 0) / max(self.winning_trades, 1)
            avg_loss = sum(t.net_pnl for t in self.closed_trades if t.net_pnl <= 0) / max(self.losing_trades, 1)
            avg_duration = sum(t.duration_minutes or 0 for t in self.closed_trades) / total_closed
        else:
            avg_pnl = avg_win = avg_loss = avg_duration = 0
        
        # Profit factor
        gross_profit = sum(t.net_pnl for t in self.closed_trades if t.net_pnl > 0)
        gross_loss = abs(sum(t.net_pnl for t in self.closed_trades if t.net_pnl < 0))
        profit_factor = gross_profit / gross_loss if gross_loss > 0 else float('inf')
        
        # Return on capital
        total_return = (self.balance - self.initial_balance) / self.initial_balance * 100
        
        return {
            # ---- Genel ----
            "initial_balance": self.initial_balance,
            "current_balance": self.balance,
            "total_return_pct": total_return,
            "total_pnl": self.total_pnl,
            "total_fees": self.total_fees,
            
            # ---- Trade Sayıları ----
            "total_trades": self.total_trades,
            "open_trades": len(self.open_trades),
            "closed_trades": total_closed,
            "winning_trades": self.winning_trades,
            "losing_trades": self.losing_trades,
            
            # ---- Oranlar ----
            "win_rate_pct": win_rate,
            "profit_factor": profit_factor,
            
            # ---- Ortalamalar ----
            "avg_pnl": avg_pnl,
            "avg_win": avg_win,
            "avg_loss": avg_loss,
            "avg_duration_min": avg_duration,
            
            # ---- Risk Metrikleri ----
            "peak_balance": self.peak_balance,
            "max_drawdown_pct": self.max_drawdown,
            
            # ---- Günlük ----
            "daily_pnl": self.daily_pnl,
            "daily_trades": self.daily_trades,
        }

    def get_open_positions_summary(self) -> List[Dict[str, Any]]:
        """Açık pozisyonların özetini döndür."""
        positions = []
        for trade in self.open_trades.values():
            positions.append({
                "trade_id": trade.trade_id,
                "symbol": trade.symbol,
                "direction": trade.direction,
                "entry_price": trade.entry_price,
                "stop_loss": trade.stop_loss,
                "take_profit": trade.take_profit,
                "position_size": trade.position_size,
                "leverage": trade.leverage,
                "opened_at": trade.opened_at,
                "ic_confidence": trade.ic_confidence,
            })
        return positions

    def print_summary(self) -> None:
        """Konsola güzel formatlanmış özet yazdır."""
        s = self.get_summary()
        
        print(f"\n{'='*60}")
        print(f"📊 PAPER TRADING PERFORMANS ÖZETİ")
        print(f"{'='*60}")
        
        print(f"\n💰 BAKİYE")
        print(f"   Başlangıç: ${s['initial_balance']:.2f}")
        print(f"   Güncel:    ${s['current_balance']:.2f}")
        print(f"   Getiri:    {s['total_return_pct']:+.1f}%")
        print(f"   Net PnL:   ${s['total_pnl']:+.2f}")
        print(f"   Ücretler:  ${s['total_fees']:.2f}")
        
        print(f"\n📈 TRADE İSTATİSTİKLERİ")
        print(f"   Toplam:    {s['total_trades']}")
        print(f"   Açık:      {s['open_trades']}")
        print(f"   Kapalı:    {s['closed_trades']}")
        print(f"   Kazanan:   {s['winning_trades']}")
        print(f"   Kaybeden:  {s['losing_trades']}")
        
        print(f"\n📊 PERFORMANS METRİKLERİ")
        print(f"   Win Rate:       {s['win_rate_pct']:.1f}%")
        print(f"   Profit Factor:  {s['profit_factor']:.2f}")
        print(f"   Ort. PnL:       ${s['avg_pnl']:.2f}")
        print(f"   Ort. Kazanç:    ${s['avg_win']:.2f}")
        print(f"   Ort. Kayıp:     ${s['avg_loss']:.2f}")
        print(f"   Ort. Süre:      {s['avg_duration_min']:.0f} dakika")
        
        print(f"\n⚠️ RİSK METRİKLERİ")
        print(f"   Peak Bakiye:    ${s['peak_balance']:.2f}")
        print(f"   Max Drawdown:   {s['max_drawdown_pct']:.1f}%")
        
        print(f"{'='*60}\n")

    # =========================================================================
    # KAYIT/YÜKLEME
    # =========================================================================

    def _save_trades(self) -> None:
        """Tüm trade'leri JSON dosyasına kaydet."""
        filepath = self.log_dir / "paper_trades.json"
        
        data = {
            "meta": {
                "initial_balance": self.initial_balance,
                "current_balance": self.balance,
                "total_trades": self.total_trades,
                "last_updated": datetime.now().isoformat(),
            },
            "open_trades": [t.to_dict() for t in self.open_trades.values()],
            "closed_trades": [t.to_dict() for t in self.closed_trades],
            "daily_pnl": self.daily_pnl,
            "daily_trades": self.daily_trades,
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

            # ── Excel otomatik export (Madde 3) ──
            try:
                self.export_to_xlsx()                  # Her kayıtta xlsx de güncelle
            except Exception as xlsx_err:
                logger.debug(f"Excel export hatası (kritik değil): {xlsx_err}")
        
        logger.debug(f"💾 Trade'ler kaydedildi: {filepath}")

    def _load_existing_trades(self) -> None:
        """Mevcut trade loglarını yükle (varsa)."""
        filepath = self.log_dir / "paper_trades.json"
        
        if not filepath.exists():
            return
        
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Meta bilgileri yükle
            meta = data.get("meta", {})
            self.balance = meta.get("current_balance", self.initial_balance)
            self.total_trades = meta.get("total_trades", 0)
            
            # Açık trade'leri yükle
            for trade_dict in data.get("open_trades", []):
                trade = self._dict_to_trade(trade_dict)
                self.open_trades[trade.trade_id] = trade
                self.all_trades.append(trade)
            
            # Kapalı trade'leri yükle
            for trade_dict in data.get("closed_trades", []):
                trade = self._dict_to_trade(trade_dict)
                self.closed_trades.append(trade)
                self.all_trades.append(trade)
                
                # İstatistikleri güncelle
                if trade.net_pnl and trade.net_pnl > 0:
                    self.winning_trades += 1
                elif trade.net_pnl:
                    self.losing_trades += 1
                if trade.net_pnl:
                    self.total_pnl += trade.net_pnl
                if trade.fees:
                    self.total_fees += trade.fees
            
            # Günlük verileri yükle
            self.daily_pnl = data.get("daily_pnl", {})
            self.daily_trades = data.get("daily_trades", {})
            
            logger.info(
                f"📂 Mevcut loglar yüklendi: "
                f"{len(self.open_trades)} açık, {len(self.closed_trades)} kapalı trade"
            )
            
        except Exception as e:
            logger.warning(f"⚠️ Trade log yükleme hatası: {e}")

    def _dict_to_trade(self, d: Dict) -> PaperTrade:
        """Sözlükten PaperTrade objesi oluştur."""
        return PaperTrade(
            trade_id=d.get("trade_id", ""),
            symbol=d.get("symbol", ""),
            full_symbol=d.get("full_symbol", ""),
            direction=d.get("direction", ""),
            entry_price=d.get("entry_price", 0),
            position_size=d.get("position_size", 0),
            position_value=d.get("position_value", 0),
            leverage=d.get("leverage", 1),
            stop_loss=d.get("stop_loss", 0),
            take_profit=d.get("take_profit", 0),
            risk_amount=d.get("risk_amount", 0),
            risk_reward=d.get("risk_reward", 0),
            ic_confidence=d.get("ic_confidence", 0),
            ic_direction=d.get("ic_direction", ""),
            best_timeframe=d.get("best_timeframe", ""),
            market_regime=d.get("market_regime", ""),
            ai_decision=d.get("ai_decision"),
            ai_confidence=d.get("ai_confidence"),
            opened_at=d.get("opened_at", ""),
            closed_at=d.get("closed_at"),
            duration_minutes=d.get("duration_minutes"),
            status=d.get("status", TradeStatus.OPEN.value),
            exit_price=d.get("exit_price"),
            exit_reason=d.get("exit_reason"),
            pnl_absolute=d.get("pnl_absolute"),
            pnl_percent=d.get("pnl_percent"),
            fees=d.get("fees", 0),
            net_pnl=d.get("net_pnl"),
            notes=d.get("notes", ""),
            tags=d.get("tags", []),
        )

    def export_to_csv(self, filepath: Optional[Path] = None) -> Path:
        """
        Tüm kapalı trade'leri CSV'ye export et.
        
        Parameters:
        ----------
        filepath : Path, optional
            Çıktı dosya yolu (varsayılan: log_dir/paper_trades.csv)
            
        Returns:
        -------
        Path
            Oluşturulan CSV dosyasının yolu
        """
        filepath = filepath or self.log_dir / "paper_trades.csv"
        
        if not self.closed_trades:
            logger.warning("⚠️ Export için kapalı trade yok")
            return filepath
        
        # CSV alanları
        fieldnames = [
            "trade_id", "symbol", "direction", "entry_price", "exit_price",
            "position_size", "leverage", "stop_loss", "take_profit",
            "pnl_absolute", "pnl_percent", "net_pnl", "fees",
            "status", "exit_reason", "duration_minutes",
            "ic_confidence", "ic_direction", "best_timeframe", "market_regime",
            "ai_decision", "ai_confidence",
            "opened_at", "closed_at", "notes"
        ]
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for trade in self.closed_trades:
                row = {k: getattr(trade, k, '') for k in fieldnames}
                writer.writerow(row)
        
        logger.info(f"📄 CSV export: {filepath} ({len(self.closed_trades)} trade)")
        return filepath


    def export_to_xlsx(self, filepath: Optional[Path] = None) -> Path:
        """
        Tüm trade'leri profesyonel formatlı Excel dosyasına export eder.
        
        2 Sheet oluşturur:
        - Trades: Her trade satır satır (açık + kapalı)
        - Summary: Performans metrikleri özeti
        
        Her _save_trades() çağrısında otomatik çalışır.
        Mevcut JSON kaydını bozmaz — ek çıktı olarak .xlsx üretir.
        
        İstatistiksel not:
        Excel çıktısı walk-forward analiz ve parametre optimizasyonu
        için pandas ile tekrar okunabilir (pd.read_excel).
        
        Parameters:
        ----------
        filepath : Path, optional
            Çıktı dosya yolu (varsayılan: log_dir/paper_trades.xlsx)
            
        Returns:
        -------
        Path
            Oluşturulan Excel dosyasının yolu
        """
        try:
            from openpyxl import Workbook                         # Excel dosya oluşturma
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side, numbers  # Hücre stilleri
            )
            from openpyxl.utils import get_column_letter           # Sütun harf dönüşümü
        except ImportError:
            logger.warning("⚠️ openpyxl yüklü değil — pip install openpyxl")
            return filepath or self.log_dir / "paper_trades.xlsx"
        
        filepath = filepath or self.log_dir / "paper_trades.xlsx"  # Varsayılan kayıt yolu
        
        wb = Workbook()                                            # Yeni Excel workbook
        
        # ─────────────────────────────────────────────────────────
        # SHEET 1: TRADES (Tüm işlemler satır satır)
        # ─────────────────────────────────────────────────────────
        ws_trades = wb.active                                      # İlk sheet
        ws_trades.title = "Trades"                                 # Sheet adı
        
        # ── Sütun tanımları: (başlık, genişlik, sayı formatı) ──
        columns = [
            ("Trade ID",       14, None),                          # Benzersiz kimlik
            ("Tarih (Açılış)", 18, None),                          # Trade açılış zamanı
            ("Tarih (Kapanış)",18, None),                          # Trade kapanış zamanı
            ("Coin",           8,  None),                          # Sembol (BTC, ETH vb.)
            ("Yön",            8,  None),                          # LONG / SHORT
            ("Giriş ($)",      12, '#,##0.00'),                    # Entry price
            ("Çıkış ($)",      12, '#,##0.00'),                    # Exit price
            ("Lot (Coin)",     12, '#,##0.000000'),                # Position size
            ("Hacim ($)",      12, '#,##0.00'),                    # Position value
            ("Kaldıraç",       10, '0x'),                          # Leverage
            ("SL ($)",         12, '#,##0.00'),                    # Stop-loss fiyatı
            ("TP ($)",         12, '#,##0.00'),                    # Take-profit fiyatı
            ("R:R",            8,  '0.00'),                        # Risk/Reward oranı
            ("PnL ($)",        10, '#,##0.00;(#,##0.00);"-"'),     # Kâr/zarar ($)
            ("PnL (%)",        10, '0.00%'),                       # Kâr/zarar (%)
            ("Fee ($)",        10, '#,##0.00'),                    # İşlem ücreti
            ("Net PnL ($)",    12, '#,##0.00;(#,##0.00);"-"'),     # Net kâr/zarar
            ("IC Güven",       10, '0.0'),                         # IC confidence score
            ("IC Yön",         8,  None),                          # IC direction
            ("TF",             6,  None),                          # En iyi timeframe
            ("Rejim",          14, None),                          # Market regime
            ("AI Karar",       10, None),                          # AI decision
            ("Durum",          14, None),                          # Trade status
            ("Çıkış Nedeni",   14, None),                          # Exit reason
            ("Süre (dk)",      10, '#,##0'),                       # Duration in minutes
        ]
        
        # ── Stil tanımları ──
        header_font = Font(                                        # Başlık fontu
            name='Arial', bold=True, color='FFFFFF', size=10
        )
        header_fill = PatternFill(                                 # Başlık arka plan (koyu mavi)
            start_color='2F5496', end_color='2F5496', fill_type='solid'
        )
        header_alignment = Alignment(                              # Başlık hizalama
            horizontal='center', vertical='center', wrap_text=True
        )
        
        data_font = Font(name='Arial', size=9)                    # Veri fontu
        
        # Koşullu renkler (PnL için)
        green_font = Font(name='Arial', size=9, color='006100')   # Kâr (yeşil)
        red_font = Font(name='Arial', size=9, color='9C0006')     # Zarar (kırmızı)
        green_fill = PatternFill(                                  # Kâr arka plan
            start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'
        )
        red_fill = PatternFill(                                    # Zarar arka plan
            start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'
        )
        
        long_fill = PatternFill(                                   # LONG arka plan (açık yeşil)
            start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'
        )
        short_fill = PatternFill(                                  # SHORT arka plan (açık kırmızı)
            start_color='FCE4EC', end_color='FCE4EC', fill_type='solid'
        )
        
        thin_border = Border(                                      # İnce kenarlık
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9'),
        )
        
        # ── Başlık satırı yaz ──
        for col_idx, (title, width, _) in enumerate(columns, 1):
            cell = ws_trades.cell(row=1, column=col_idx, value=title)
            cell.font = header_font                                # Beyaz kalın font
            cell.fill = header_fill                                # Koyu mavi arka plan
            cell.alignment = header_alignment                      # Ortala
            ws_trades.column_dimensions[get_column_letter(col_idx)].width = width
        
        # ── Freeze panes: başlık satırı sabit kalsın ──
        ws_trades.freeze_panes = 'A2'                              # Scroll'da başlık kaybolmaz
        
        # ── Trade verilerini yaz (tüm tradeler: önce açık, sonra kapalı) ──
        all_trade_list = list(self.open_trades.values()) + self.closed_trades
        
        for row_idx, trade in enumerate(all_trade_list, 2):        # 2'den başla (1 = başlık)
            # Her sütuna karşılık gelen trade alanı
            row_data = [
                trade.trade_id[:12] if trade.trade_id else "",     # Trade ID (kısa)
                trade.opened_at or "",                             # Açılış tarihi
                trade.closed_at or "",                             # Kapanış tarihi
                trade.symbol or "",                                # Coin adı
                trade.direction or "",                             # LONG/SHORT
                trade.entry_price or 0,                            # Giriş fiyatı
                trade.exit_price if trade.exit_price else None,    # Çıkış fiyatı
                trade.position_size or 0,                          # Lot miktarı
                trade.position_value or 0,                         # Pozisyon hacmi ($)
                trade.leverage or 1,                               # Kaldıraç
                trade.stop_loss or 0,                              # Stop-loss
                trade.take_profit or 0,                            # Take-profit
                trade.risk_reward or 0,                            # R:R oranı
                trade.pnl_absolute if trade.pnl_absolute else 0,  # PnL ($)
                (trade.pnl_percent or 0) / 100 if trade.pnl_percent else 0,  # PnL (%) — Excel formatı
                trade.fees or 0,                                   # Fee
                trade.net_pnl if trade.net_pnl else 0,            # Net PnL
                trade.ic_confidence or 0,                          # IC güven
                trade.ic_direction or "",                          # IC yön
                trade.best_timeframe or "",                        # TF
                trade.market_regime or "",                         # Rejim
                trade.ai_decision or "",                           # AI karar
                trade.status or "",                                # Durum
                trade.exit_reason or "",                           # Çıkış nedeni
                trade.duration_minutes if trade.duration_minutes else None,
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_trades.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font                              # Veri fontu
                cell.border = thin_border                          # Kenarlık
                
                # Sayı formatı uygula (sütun bazlı)
                _, _, fmt = columns[col_idx - 1]
                if fmt and value is not None:
                    cell.number_format = fmt
            
            # ── Koşullu formatlama: Yön rengi ──
            direction_cell = ws_trades.cell(row=row_idx, column=5)  # Yön sütunu (E)
            if trade.direction == "LONG":
                direction_cell.fill = long_fill                    # LONG → açık yeşil
            elif trade.direction == "SHORT":
                direction_cell.fill = short_fill                   # SHORT → açık kırmızı
            
            # ── Koşullu formatlama: PnL renklendirme ──
            pnl_val = trade.net_pnl or trade.pnl_absolute or 0
            for pnl_col in [14, 17]:                               # PnL($) ve Net PnL($) sütunları
                pnl_cell = ws_trades.cell(row=row_idx, column=pnl_col)
                if pnl_val > 0:
                    pnl_cell.font = green_font                     # Kâr → yeşil font
                    pnl_cell.fill = green_fill                     # Yeşil arka plan
                elif pnl_val < 0:
                    pnl_cell.font = red_font                       # Zarar → kırmızı font
                    pnl_cell.fill = red_fill                       # Kırmızı arka plan
            
            # ── Durum sütunu renklendirme ──
            status_cell = ws_trades.cell(row=row_idx, column=22)   # Durum sütunu
            if "tp" in (trade.status or "").lower():
                status_cell.fill = green_fill                      # TP → yeşil
            elif "sl" in (trade.status or "").lower():
                status_cell.fill = red_fill                        # SL → kırmızı
        
        # ── AutoFilter ekle (sıralama/filtreleme için) ──
        if len(all_trade_list) > 0:
            last_col = get_column_letter(len(columns))
            ws_trades.auto_filter.ref = f"A1:{last_col}{len(all_trade_list) + 1}"
        
        # ─────────────────────────────────────────────────────────
        # SHEET 2: SUMMARY (Performans özeti)
        # ─────────────────────────────────────────────────────────
        ws_summary = wb.create_sheet("Summary")                    # İkinci sheet
        
        # Performans metriklerini hesapla
        summary = self.get_summary()                               # Mevcut summary metodu
        
        # ── Başlık stili ──
        title_font = Font(name='Arial', bold=True, size=14, color='2F5496')
        section_font = Font(name='Arial', bold=True, size=11, color='2F5496')
        label_font = Font(name='Arial', size=10)
        value_font = Font(name='Arial', bold=True, size=10)
        
        # ── Başlık ──
        ws_summary['A1'] = '📊 PAPER TRADING PERFORMANS RAPORU'
        ws_summary['A1'].font = title_font
        ws_summary.merge_cells('A1:D1')
        
        ws_summary['A2'] = f'Oluşturulma: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws_summary['A2'].font = Font(name='Arial', size=9, italic=True, color='808080')
        
        # ── Sütun genişlikleri ──
        ws_summary.column_dimensions['A'].width = 22               # Metrik adı
        ws_summary.column_dimensions['B'].width = 15               # Değer
        ws_summary.column_dimensions['C'].width = 22               # Metrik adı (2. sütun)
        ws_summary.column_dimensions['D'].width = 15               # Değer (2. sütun)
        
        # ── BAKİYE BİLGİLERİ (Satır 4'ten başla) ──
        row = 4
        ws_summary.cell(row=row, column=1, value='💰 BAKİYE').font = section_font
        row += 1
        
        balance_items = [
            ('Başlangıç Bakiye', f"${summary.get('initial_balance', 0):.2f}"),
            ('Güncel Bakiye', f"${summary.get('current_balance', 0):.2f}"),
            ('Toplam Getiri', f"{summary.get('total_return_pct', 0):+.1f}%"),
            ('Net PnL', f"${summary.get('total_pnl', 0):+.2f}"),
            ('Toplam Fee', f"${summary.get('total_fees', 0):.2f}"),
        ]
        
        for label, value in balance_items:
            ws_summary.cell(row=row, column=1, value=label).font = label_font
            val_cell = ws_summary.cell(row=row, column=2, value=value)
            val_cell.font = value_font
            # Getiri rengini ayarla
            if 'Getiri' in label or 'PnL' in label:
                if '+' in value or (value.replace('$','').replace('%','').strip().startswith('-') is False and float(value.replace('$','').replace('%','').replace('+','').strip() or 0) > 0):
                    val_cell.font = Font(name='Arial', bold=True, size=10, color='006100')
                elif '-' in value:
                    val_cell.font = Font(name='Arial', bold=True, size=10, color='9C0006')
            row += 1
        
        # ── TRADE İSTATİSTİKLERİ ──
        row += 1
        ws_summary.cell(row=row, column=1, value='📈 TRADE İSTATİSTİKLERİ').font = section_font
        row += 1
        
        trade_items = [
            ('Toplam Trade', str(summary.get('total_trades', 0))),
            ('Açık Pozisyon', str(summary.get('open_trades', 0))),
            ('Kapalı Trade', str(summary.get('closed_trades', 0))),
            ('Kazanan', str(summary.get('winning_trades', 0))),
            ('Kaybeden', str(summary.get('losing_trades', 0))),
        ]
        
        for label, value in trade_items:
            ws_summary.cell(row=row, column=1, value=label).font = label_font
            ws_summary.cell(row=row, column=2, value=value).font = value_font
            row += 1
        
        # ── PERFORMANS METRİKLERİ ──
        row += 1
        ws_summary.cell(row=row, column=1, value='📊 PERFORMANS METRİKLERİ').font = section_font
        row += 1
        
        perf_items = [
            ('Win Rate', f"{summary.get('win_rate_pct', 0):.1f}%"),
            ('Profit Factor', f"{summary.get('profit_factor', 0):.2f}"),
            ('Ort. PnL', f"${summary.get('avg_pnl', 0):.2f}"),
            ('Ort. Kazanç', f"${summary.get('avg_win', 0):.2f}"),
            ('Ort. Kayıp', f"${summary.get('avg_loss', 0):.2f}"),
            ('Ort. Süre', f"{summary.get('avg_duration_min', 0):.0f} dk"),
        ]
        
        for label, value in perf_items:
            ws_summary.cell(row=row, column=1, value=label).font = label_font
            ws_summary.cell(row=row, column=2, value=value).font = value_font
            row += 1
        
        # ── RİSK METRİKLERİ ──
        row += 1
        ws_summary.cell(row=row, column=1, value='⚠️ RİSK METRİKLERİ').font = section_font
        row += 1
        
        risk_items = [
            ('Peak Bakiye', f"${summary.get('peak_balance', 0):.2f}"),
            ('Max Drawdown', f"{summary.get('max_drawdown_pct', 0):.1f}%"),
        ]
        
        for label, value in risk_items:
            ws_summary.cell(row=row, column=1, value=label).font = label_font
            ws_summary.cell(row=row, column=2, value=value).font = value_font
            row += 1
        
        # ── GÜNLÜK PNL TABLOSU (varsa) ──
        if self.daily_pnl:
            row += 2
            ws_summary.cell(row=row, column=1, value='📅 GÜNLÜK PNL').font = section_font
            row += 1
            
            # Başlıklar
            for col, title in enumerate(['Tarih', 'PnL ($)', 'Trade Sayısı'], 1):
                cell = ws_summary.cell(row=row, column=col, value=title)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            # Günlük veriler (tarihe göre sıralı)
            for date_str in sorted(self.daily_pnl.keys()):
                ws_summary.cell(row=row, column=1, value=date_str).font = label_font
                
                pnl_val = self.daily_pnl[date_str]
                pnl_cell = ws_summary.cell(row=row, column=2, value=f"${pnl_val:+.2f}")
                pnl_cell.font = Font(
                    name='Arial', size=10, bold=True,
                    color='006100' if pnl_val >= 0 else '9C0006'   # Yeşil/kırmızı
                )
                
                trade_count = self.daily_trades.get(date_str, 0)
                ws_summary.cell(row=row, column=3, value=trade_count).font = label_font
                row += 1
        
        # ── Kaydet ──
        try:
            wb.save(str(filepath))                                 # Excel dosyasını diske yaz
            logger.info(f"📊 Excel export: {filepath} ({len(all_trade_list)} trade)")
        except PermissionError:
            # Dosya açıksa (Excel'de izliyorsa) alternatif isimle kaydet
            alt_path = filepath.with_stem(filepath.stem + f"_{datetime.now().strftime('%H%M%S')}")
            wb.save(str(alt_path))
            logger.warning(f"⚠️ Orijinal dosya kilitli, alternatif kaydedildi: {alt_path}")
            filepath = alt_path
        
        return filepath

    def reset(self) -> None:
        """Tüm verileri sıfırla (yeni başlangıç)."""
        self.balance = self.initial_balance
        self.open_trades.clear()
        self.closed_trades.clear()
        self.all_trades.clear()
        self.total_trades = 0
        self.winning_trades = 0
        self.losing_trades = 0
        self.total_pnl = 0.0
        self.total_fees = 0.0
        self.peak_balance = self.initial_balance
        self.max_drawdown = 0.0
        self.daily_pnl.clear()
        self.daily_trades.clear()
        
        if self.auto_save:
            self._save_trades()
        
        logger.info("🔄 PaperTrader sıfırlandı")


# =============================================================================
# MODÜL TESTİ
# =============================================================================

if __name__ == "__main__":
    # Basit test
    import tempfile
    
    with tempfile.TemporaryDirectory() as tmpdir:
        pt = PaperTrader(initial_balance=100.0, log_dir=Path(tmpdir))
        
        # Trade aç
        trade = pt.open_trade(
            symbol="BTC",
            full_symbol="BTC/USDT:USDT",
            direction="LONG",
            entry_price=95000,
            position_size=0.01,
            stop_loss=94000,
            take_profit=97000,
            leverage=5,
            ic_confidence=75,
            ic_direction="LONG",
            best_timeframe="4h",
            market_regime="trending_up",
        )
        
        print(f"Trade açıldı: {trade.trade_id}")
        
        # TP tetiklenmesini simüle et
        closed = pt.check_exits({"BTC": 97500})
        
        if closed:
            print(f"Trade kapandı: {closed[0].exit_reason}")
        
        # Özet
        pt.print_summary()
